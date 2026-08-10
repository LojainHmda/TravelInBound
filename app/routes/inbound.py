from flask import Blueprint, render_template, render_template_string, redirect, url_for, flash, request, jsonify, abort, send_file, current_app
from flask_login import login_required, current_user
from datetime import datetime, timedelta, date
import calendar
import time
from typing import cast, Any
import json
import os
import sys
from sqlalchemy.orm import selectinload

from app.extensions import db, csrf
from app.models.inbound import (
    InboundRequest, ItineraryRow, InboundHotel, InboundTransport,
    InboundMeal, InboundGuide, InboundCashExpense, InboundDocument,
    HotelRoom, COST_UNIT_PER_PERSON, COST_UNIT_PER_GROUP,
    ArrivalBatch, DepartureBatch, InboundRepresentative,
    itinerary_row_guide_supplier_id_list, HotelCategory,
)
from werkzeug.utils import secure_filename
import uuid
from urllib.parse import urlencode
from app.models import STATUS_REQUEST, STATUS_QUOTED, STATUS_RESERVED, STATUS_BOOKED, STATUS_IN_PROGRESS, STATUS_CONFIRMED
from app.models.service import ServiceItem
from app.models.booking import Booking
from app.models.invoice import Invoice
from app.models.customer import Customer
from app.models.supplier import Supplier
from app.services.proforma_doc_generator import ProformaDocGenerator
from app.services.voucher_trip_plan_generator import VoucherTripPlanGenerator
from app.services.storage import document_storage
from app.utils import is_valid_phone, PHONE_ERROR

# Create blueprint for inbound tour operator routes
inbound_bp = Blueprint('inbound', __name__, url_prefix='/inbound')

# Guide languages available in the creation form
GUIDE_LANGUAGES = [
    'Arabic',
    'English',
    'French',
    'German',
    'Spanish',
    'Italian',
    'Russian',
    'Chinese',
    'Japanese',
    'Korean',
    'Turkish',
    'Hebrew',
    'Persian',
    'Hindi',
    'Portuguese',
    'Dutch',
    'Greek',
    'Swedish',
    'Norwegian',
    'Danish',
    'Polish',
    'Czech',
    'Hungarian',
    'Romanian',
    'Bulgarian',
    'Croatian',
    'Serbian',
    'Other',
]

_SUPPLIER_DROPDOWN_CACHE = {
    'expires_at': 0.0,
    'data': None,
}


def _invalidate_supplier_dropdown_cache():
    _SUPPLIER_DROPDOWN_CACHE['expires_at'] = 0.0
    _SUPPLIER_DROPDOWN_CACHE['data'] = None


class _RepresentativeOption:
    """Lightweight object for template dropdowns (id + name)."""
    __slots__ = ('id', 'name')

    def __init__(self, rep_id, name):
        self.id = rep_id
        self.name = name


def _normalize_meet_assist_name(name):
    return str(name or '').strip()


def _find_inbound_representative_by_name(name):
    n = _normalize_meet_assist_name(name)
    if not n:
        return None
    return InboundRepresentative.query.filter(
        db.func.lower(InboundRepresentative.name) == n.lower()
    ).first()


def _find_ground_handler_supplier_by_name(name):
    from app.models.supplier import Supplier

    n = _normalize_meet_assist_name(name)
    if not n:
        return None
    return Supplier.query.filter(
        Supplier.supplier_type == 'GROUND_HANDLER',
        db.func.lower(Supplier.name) == n.lower(),
    ).first()


def _generate_ground_handler_code():
    from app.models.supplier import Supplier

    count = Supplier.query.filter(Supplier.code.like('GHD-%')).count()
    return f'GHD-{count + 1:03d}'


def _ensure_inbound_representative(name):
    n = _normalize_meet_assist_name(name)
    if not n:
        return None
    existing = _find_inbound_representative_by_name(n)
    if existing:
        return existing
    rep = InboundRepresentative(name=n)
    db.session.add(rep)
    return rep


def _resolve_supplier_bank_fields(data):
    """Map payment_method (or legacy payment_terms Cliq) to stored bank_name/bank_account."""
    payment_method = (data.get('payment_method') or '').strip()
    payment_terms = (data.get('payment_terms') or '').strip()
    cliq_alias = (data.get('cliq_alias') or '').strip()
    bank_name = (data.get('bank_name') or '').strip()
    bank_account = (data.get('bank_account') or '').strip()

    if payment_method == 'Cliq':
        return 'Cliq', cliq_alias or None
    if payment_method == 'Bank':
        return bank_name or None, bank_account or None
    if payment_terms == 'Cliq':
        return 'Cliq', cliq_alias or None
    return bank_name or None, bank_account or None


def _merge_payment_method_into_notes(notes, payment_method):
    """Prepend payment method line to notes when set."""
    method = (payment_method or '').strip()
    if not method:
        return notes
    base = (notes or '').strip()
    if base.startswith('Payment Method:'):
        return base or None
    merged = f"Payment Method: {method}\n{base}".strip() if base else f"Payment Method: {method}"
    return merged or None


def _ensure_ground_handler_supplier(name, data=None):
    """Create GROUND_HANDLER supplier if missing; optional extra fields from add form."""
    from app.models.supplier import Supplier

    n = _normalize_meet_assist_name(name)
    if not n:
        return None
    existing = _find_ground_handler_supplier_by_name(n)
    if existing:
        return existing

    payload = data or {}
    languages_val = (payload.get('languages') or '').strip() or None
    payment_terms = (payload.get('payment_terms') or '').strip() or None
    payment_method = (payload.get('payment_method') or '').strip()
    entity_type = (payload.get('supplier_type') or 'COMPANY').strip().upper() or 'COMPANY'
    bank_name_val, bank_account_val = _resolve_supplier_bank_fields(payload)
    notes_value = _merge_payment_method_into_notes(
        (payload.get('notes') or '').strip() or None,
        payment_method,
    )

    new_supplier = Supplier(
        name=n,
        code=_generate_ground_handler_code(),
        supplier_type='GROUND_HANDLER',
        entity_type=entity_type,
        languages=languages_val,
        phone=(payload.get('phone') or '').strip() or None,
        contact_person=(payload.get('contact_person') or '').strip() or None,
        email=(payload.get('email') or '').strip() or None,
        website=(payload.get('website') or '').strip() or None,
        city=(payload.get('city') or '').strip() or None,
        country=(payload.get('country') or '').strip() or None,
        payment_terms=payment_terms,
        default_currency=(payload.get('default_currency') or 'USD') or 'USD',
        address=(payload.get('address') or '').strip() or None,
        bank_name=bank_name_val,
        bank_account=bank_account_val,
        tax_number=(payload.get('tax_number') or '').strip() or None,
        notes=notes_value,
        is_active=True,
    )
    db.session.add(new_supplier)
    return new_supplier


def _sync_meet_assist_representative_pair(name, supplier_data=None):
    """Keep representative dropdown and Meet & Assist hub list aligned by name."""
    n = _normalize_meet_assist_name(name)
    if not n:
        return None, None
    rep = _ensure_inbound_representative(n)
    supplier = _ensure_ground_handler_supplier(n, supplier_data)
    return rep, supplier


def _get_merged_representatives_for_dropdown():
    """Union of InboundRepresentative names and GROUND_HANDLER supplier names."""
    from app.models.supplier import Supplier

    by_key = {}
    for r in InboundRepresentative.query.order_by(InboundRepresentative.name).all():
        n = _normalize_meet_assist_name(r.name)
        if n:
            by_key[n.lower()] = {'id': r.id, 'name': n}
    for s in Supplier.query.filter(
        Supplier.is_active == True,
        Supplier.supplier_type == 'GROUND_HANDLER',
    ).order_by(Supplier.name).all():
        n = _normalize_meet_assist_name(s.name)
        if not n or n.lower() in by_key:
            continue
        rep = _find_inbound_representative_by_name(n)
        by_key[n.lower()] = {'id': rep.id if rep else 0, 'name': n}
    return sorted(by_key.values(), key=lambda x: x['name'].lower())


def sync_all_meet_assist_representative_pairs():
    """Backfill missing rows so hub and dropdown always share the same names."""
    from app.models.supplier import Supplier

    changed = False
    rep_names = {
        _normalize_meet_assist_name(r.name).lower()
        for r in InboundRepresentative.query.all()
        if _normalize_meet_assist_name(r.name)
    }
    suppliers = Supplier.query.filter(
        Supplier.is_active == True,
        Supplier.supplier_type == 'GROUND_HANDLER',
    ).all()
    supplier_names = {
        _normalize_meet_assist_name(s.name).lower()
        for s in suppliers
        if _normalize_meet_assist_name(s.name)
    }

    for r in InboundRepresentative.query.all():
        n = _normalize_meet_assist_name(r.name)
        if n and n.lower() not in supplier_names:
            _ensure_ground_handler_supplier(n)
            changed = True

    for s in suppliers:
        n = _normalize_meet_assist_name(s.name)
        if n and n.lower() not in rep_names:
            _ensure_inbound_representative(n)
            changed = True

    if changed:
        try:
            db.session.commit()
            _invalidate_supplier_dropdown_cache()
        except Exception:
            db.session.rollback()
            raise


def _get_supplier_dropdown_data(cache_ttl_seconds: int = 120):
    """Load supplier dropdown data once and reuse briefly across requests."""
    now = time.time()
    cached = _SUPPLIER_DROPDOWN_CACHE.get('data')
    if cached is not None and now < float(_SUPPLIER_DROPDOWN_CACHE.get('expires_at', 0.0)):
        return cached

    from app.models.supplier import Supplier

    suppliers = Supplier.query.filter(
        Supplier.is_active == True,
        Supplier.supplier_type.in_(['HOTEL', 'ACCOMMODATION', 'TRANSPORT', 'GUIDE', 'RESTAURANT', 'GROUND_HANDLER']),
    ).order_by(Supplier.supplier_type, Supplier.city, Supplier.name).all()
    representatives = [
        _RepresentativeOption(r['id'], r['name'])
        for r in _get_merged_representatives_for_dropdown()
    ]

    hotels = []
    transports = []
    guides = []
    restaurants = []
    ground_handlers = []
    for s in suppliers:
        t = (s.supplier_type or '').upper()
        if t in ('HOTEL', 'ACCOMMODATION'):
            hotels.append(s)
        elif t == 'TRANSPORT':
            transports.append(s)
        elif t == 'GUIDE':
            guides.append(s)
        elif t == 'RESTAURANT':
            restaurants.append(s)
        elif t == 'GROUND_HANDLER':
            ground_handlers.append(s)

    hotels_by_city = {}
    city_order = ['Amman', 'Aqaba', 'Petra', 'Dead Sea', 'Other']
    for hotel in hotels:
        city = hotel.city or 'Other'
        hotels_by_city.setdefault(city, []).append(hotel)
    sorted_hotels_by_city = {
        city: hotels_by_city.get(city, [])
        for city in city_order if city in hotels_by_city
    }

    # Extract unique cities and categories from hotels
    accommodation_cities = []
    accommodation_categories = set()
    for hotel in hotels:
        city = hotel.city or 'Other'
        if city not in accommodation_cities:
            accommodation_cities.append(city)
        category = hotel.accommodation_category
        if category:
            accommodation_categories.add(category)

    # Order cities: put standard ones first, then others
    city_order = ['Amman', 'Aqaba', 'Petra', 'Dead Sea', 'Other']
    ordered_cities = [c for c in city_order if c in accommodation_cities]
    ordered_cities.extend([c for c in accommodation_cities if c not in city_order])

    # Sort categories alphabetically
    ordered_categories = sorted(list(accommodation_categories))

    data = {
        'hotel_suppliers': hotels,
        'hotels_by_city': sorted_hotels_by_city,
        'accommodation_cities': ordered_cities,
        'accommodation_categories': ordered_categories,
        'transport_suppliers': transports,
        'guide_suppliers': guides,
        'representatives': representatives,
        'restaurant_suppliers': restaurants,
        'ground_handler_suppliers': ground_handlers,
    }
    _SUPPLIER_DROPDOWN_CACHE['data'] = data
    _SUPPLIER_DROPDOWN_CACHE['expires_at'] = now + cache_ttl_seconds
    return data


def _parse_needs_transport(val, default=True):
    """Parse Yes/No from arrival/departure Individual Transport dropdown."""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ('no', 'false', '0', 'off'):
        return False
    return True


def _transport_flight_stub_complete(t: InboundTransport) -> bool:
    """Green chip when supplier and vehicle type are set (execution layer complete)."""
    return bool(t.supplier_id and (t.vehicle_type or '').strip())


def _is_individual_transport_from_flight(t: InboundTransport) -> bool:
    """True for arrival/departure individual-transfer stubs (chip flow), including legacy rows missing batch FK."""
    if t.source_arrival_batch_id is not None or t.source_departure_batch_id is not None:
        return True
    if t.is_airport_transfer and (t.is_arrival or t.is_departure):
        return True
    return False


def _include_transport_in_trip_summary(t: InboundTransport) -> bool:
    """Individual transports from flights appear in trip summary / export only when supplier + vehicle are saved."""
    if _is_individual_transport_from_flight(t):
        return _transport_flight_stub_complete(t)
    return True


def _trip_summary_transports(transports) -> list:
    """Ordered list of transports shown in Trip Summary and Word tour file."""
    lst = list(transports or [])
    lst.sort(key=lambda x: (x.date or date.min, x.id))
    return [t for t in lst if _include_transport_in_trip_summary(t)]


def _detach_individual_transport_from_batch(transport: InboundTransport) -> None:
    """Turn off Individual Transport on the source flight batch so stubs are not recreated."""
    if transport.source_arrival_batch_id:
        arr = ArrivalBatch.query.filter_by(
            id=transport.source_arrival_batch_id, request_id=transport.request_id
        ).first()
        if arr:
            arr.needs_transport = False
    if transport.source_departure_batch_id:
        dep = DepartureBatch.query.filter_by(
            id=transport.source_departure_batch_id, request_id=transport.request_id
        ).first()
        if dep:
            dep.needs_transport = False


def _sync_transport_from_arrival(arrival: ArrivalBatch) -> None:
    """Create/update or remove InboundTransport stub for an arrival batch (after flush)."""
    if not arrival or not arrival.id or not arrival.arrival_date:
        return
    req_id = arrival.request_id
    need = arrival.needs_transport
    if need is None:
        need = True
    if not need:
        InboundTransport.query.filter_by(
            request_id=req_id, source_arrival_batch_id=arrival.id
        ).delete(synchronize_session=False)
        return
    t = InboundTransport.query.filter_by(
        request_id=req_id, source_arrival_batch_id=arrival.id
    ).first()
    if not t:
        t = InboundTransport(
            request_id=req_id,
            date=arrival.arrival_date,
            end_date=arrival.arrival_date,
            pax=arrival.pax_count or 0,
            source_arrival_batch_id=arrival.id,
        )
        db.session.add(t)
    t.date = arrival.arrival_date
    t.end_date = arrival.arrival_date
    t.pax = arrival.pax_count or 0
    t.pickup_time = arrival.arrival_time
    t.pickup_location = arrival.arrival_point
    t.is_airport_transfer = True
    t.is_arrival = True
    t.is_departure = False


def _sync_transport_from_departure(departure: DepartureBatch) -> None:
    """Create/update or remove InboundTransport stub for a departure batch (after flush)."""
    if not departure or not departure.id or not departure.departure_date:
        return
    req_id = departure.request_id
    need = departure.needs_transport
    if need is None:
        need = True
    if not need:
        InboundTransport.query.filter_by(
            request_id=req_id, source_departure_batch_id=departure.id
        ).delete(synchronize_session=False)
        return
    t = InboundTransport.query.filter_by(
        request_id=req_id, source_departure_batch_id=departure.id
    ).first()
    if not t:
        t = InboundTransport(
            request_id=req_id,
            date=departure.departure_date,
            end_date=departure.departure_date,
            pax=departure.pax_count or 0,
            source_departure_batch_id=departure.id,
        )
        db.session.add(t)
    t.date = departure.departure_date
    t.end_date = departure.departure_date
    t.pax = departure.pax_count or 0
    t.pickup_time = departure.departure_time
    t.dropoff_location = departure.departure_point
    t.is_airport_transfer = True
    t.is_arrival = False
    t.is_departure = True


def _reconcile_flight_linked_transports(request_id: int) -> None:
    """Create/remove InboundTransport stubs from arrival/departure batches (idempotent).

    Call from read endpoints if saves ran without sync (e.g. server not restarted after deploy).
    """
    for arr in ArrivalBatch.query.filter_by(request_id=request_id).order_by(ArrivalBatch.id).all():
        _sync_transport_from_arrival(arr)
    for dep in DepartureBatch.query.filter_by(request_id=request_id).order_by(DepartureBatch.id).all():
        _sync_transport_from_departure(dep)


def _hub_status_scope_label(status_val):
    """Human label for Hub status filter (?status=) — matches inbound list filter mapping."""
    if not status_val:
        return None
    u = str(status_val).strip().upper()
    if u == 'REQUEST':
        return 'Request'
    if u in ('SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS'):
        return 'Confirmed'
    if u in ('INVOICE', 'COMPLETED', 'INVOICED'):
        return 'Invoiced'
    return None


def _inbound_list_filter_summary():
    """Human-readable active filters for list export/print headers."""
    parts = []
    request_number = (request.args.get('request_number') or '').strip()
    if request_number:
        parts.append(('Request Number', request_number))
    agent = (request.args.get('agent') or '').strip()
    if agent:
        parts.append(('Agent', agent))
    filter_year = (request.args.get('filter_year') or '').strip()
    if filter_year:
        parts.append(('Year', filter_year))
    filter_month_raw = (request.args.get('filter_month') or '').strip()
    if filter_month_raw:
        try:
            parts.append(('Month', calendar.month_name[int(filter_month_raw)]))
        except (ValueError, IndexError):
            parts.append(('Month', filter_month_raw))
    status = (request.args.get('status') or '').strip()
    if status:
        parts.append(('Status', _hub_status_scope_label(status) or status))
    queue = (request.args.get('queue') or '').strip().lower()
    if queue == 'deleted':
        parts.append(('Queue', 'Deleted'))
    return parts


def _inbound_row_status_label(inb, queue):
    if queue == 'deleted' or getattr(inb, 'pending_invoice_queue', False):
        return 'Deleted'
    status_val = str(inb.status or '').upper()
    if status_val == 'REQUEST':
        return 'Request'
    if status_val in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS']:
        return 'Confirmed'
    if status_val in ['INVOICE', 'COMPLETED', 'INVOICED']:
        return 'Invoiced'
    return status_val.title() if status_val else ''


def _inbound_list_context():
    """Build template context for inbound list views (main index and /all-files)."""
    path_tail = (request.path or '').rstrip('/')
    all_files_view = path_tail.endswith('all-files')
    query, queue = _build_inbound_list_query(all_files_view)

    now_y = datetime.now().year
    list_filter_years = list(range(now_y + 1, now_y - 16, -1))
    list_filter_months = [(str(i), calendar.month_name[i]) for i in range(1, 13)]

    # List order: newest first by default (req_sort=desc); req_sort=asc = oldest first
    req_sort = (request.args.get('req_sort') or 'desc').strip().lower()
    if req_sort not in ('asc', 'desc'):
        req_sort = 'desc'
    if req_sort == 'desc':
        query = query.order_by(InboundRequest.created_at.desc(), InboundRequest.id.desc())
    else:
        query = query.order_by(InboundRequest.created_at.asc(), InboundRequest.id.asc())

    page_raw = (request.args.get('page') or '1').strip()
    try:
        page = max(1, int(page_raw))
    except ValueError:
        page = 1
    per_page = 50
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    requests = pagination.items

    def _inbound_list_qs(updates):
        d = request.args.to_dict(flat=True)
        d.update(updates)
        if 'page' not in updates:
            d['page'] = '1'
        return '?' + urlencode(d)

    req_sort_qs_desc = _inbound_list_qs({'req_sort': 'desc'})
    req_sort_qs_asc = _inbound_list_qs({'req_sort': 'asc'})
    export_qs = _inbound_list_qs({})
    page_qs = {p: _inbound_list_qs({'page': str(p)}) for p in range(1, pagination.pages + 1)}
    prev_page_qs = _inbound_list_qs({'page': str(pagination.prev_num)}) if pagination.has_prev else None
    next_page_qs = _inbound_list_qs({'page': str(pagination.next_num)}) if pagination.has_next else None
    page_numbers = list(pagination.iter_pages(left_edge=1, left_current=1, right_current=2, right_edge=1))
    search_filter_fields = ('request_number', 'agent', 'filter_year', 'filter_month')
    has_search_filters = any((request.args.get(f) or '').strip() for f in search_filter_fields)
    all_filter_fields = ('request_number', 'agent', 'filter_year', 'filter_month', 'status')
    has_active_filters = any((request.args.get(f) or '').strip() for f in all_filter_fields)
    has_submitted_search = (request.args.get('search') or '').strip() == '1'
    # Status-specific pages (Request/Confirmed/Invoiced/Deleted) show no data until Search is clicked
    is_status_page = (bool((request.args.get('status') or '').strip()) or queue == 'deleted') and not all_files_view
    if is_status_page:
        show_results_table = has_search_filters or has_submitted_search
    else:
        show_results_table = (not all_files_view) or has_active_filters or has_submitted_search

    # Run-down plan: expensive (loads confirmed/booked requests + itinerary). The template only
    # shows it on the main list without a Hub status filter and not on /inbound/all-files.
    status_arg = (request.args.get('status') or '').strip()
    base_page_title = 'All inbound files' if all_files_view else 'Inbound Tour Requests'
    scope_label = _hub_status_scope_label(request.args.get('status'))
    if queue == 'deleted':
        inbound_page_title = 'All inbound files — Deleted' if all_files_view else 'Inbound tour Deleted'
    elif scope_label:
        inbound_page_title = f'{base_page_title} — {scope_label}' if all_files_view else f'Inbound tour {scope_label}'
    else:
        inbound_page_title = base_page_title
    need_run_down = (
        queue != 'deleted'
        and not status_arg
        and not all_files_view
    )
    run_down_data = get_run_down_data_by_date() if need_run_down else []

    return {
        'requests': requests,
        'run_down_data': run_down_data,
        'is_deleted_queue': queue == 'deleted',
        'list_filter_years': list_filter_years,
        'list_filter_months': list_filter_months,
        'req_sort': req_sort,
        'req_sort_qs_desc': req_sort_qs_desc,
        'req_sort_qs_asc': req_sort_qs_asc,
        'pagination': pagination,
        'page_numbers': page_numbers,
        'page_qs': page_qs,
        'prev_page_qs': prev_page_qs,
        'next_page_qs': next_page_qs,
        'export_qs': export_qs,
        'show_results_table': show_results_table,
        'inbound_page_title': inbound_page_title,
        'all_files_view': all_files_view,
    }


def _build_inbound_list_query(all_files_view: bool):
    """Shared inbound list query (filters only, no sort/pagination)."""
    from sqlalchemy import or_, func, case, and_

    query = InboundRequest.query

    # Always exclude unsaved draft requests (temporary IN-NEW- numbers)
    query = query.filter(
        or_(
            InboundRequest.is_saved == True,
            ~InboundRequest.request_number.like('IN-NEW-%')
        )
    )

    # Main list vs. Deleted queue (removed via trash on list; Hub tile "Deleted")
    queue = (request.args.get('queue') or '').strip().lower()
    if queue == 'to_invoice':
        queue = 'deleted'
    if queue == 'deleted':
        query = query.filter(InboundRequest.pending_invoice_queue.is_(True))
    elif not all_files_view:
        query = query.filter(InboundRequest.pending_invoice_queue.isnot(True))

    request_number = request.args.get('request_number', '')
    if request_number:
        query = query.filter(InboundRequest.request_number.contains(request_number))

    agent = (request.args.get('agent') or '').strip()
    if agent:
        like = f'%{agent}%'
        cust_full = func.trim(
            func.concat(
                Customer.first_name,
                ' ',
                func.coalesce(Customer.last_name, ''),
            )
        )
        display_agent = case(
            (InboundRequest.customer_id.isnot(None), cust_full),
            else_=func.coalesce(InboundRequest.contact_name, ''),
        )
        query = query.outerjoin(Customer, InboundRequest.customer_id == Customer.id)
        if agent.lower() == 'tba':
            blank_agent_placeholder = and_(
                InboundRequest.customer_id.is_(None),
                or_(
                    InboundRequest.contact_name.is_(None),
                    func.trim(InboundRequest.contact_name) == '',
                ),
            )
            query = query.filter(or_(display_agent.ilike(like), blank_agent_placeholder))
        else:
            query = query.filter(display_agent.ilike(like))

    filter_year_raw = (request.args.get('filter_year') or '').strip()
    filter_month_raw = (request.args.get('filter_month') or '').strip()
    y = None
    m = None
    if filter_year_raw:
        try:
            y = int(filter_year_raw)
        except ValueError:
            y = None
    if filter_month_raw:
        try:
            m = int(filter_month_raw)
            if m < 1 or m > 12:
                m = None
        except ValueError:
            m = None
    if y is not None and m is not None:
        month_start = date(y, m, 1)
        month_end = date(y, m, calendar.monthrange(y, m)[1])
        query = query.filter(
            InboundRequest.from_date <= month_end,
            InboundRequest.to_date >= month_start,
        )
    elif y is not None:
        year_start = date(y, 1, 1)
        year_end = date(y, 12, 31)
        query = query.filter(
            InboundRequest.from_date <= year_end,
            InboundRequest.to_date >= year_start,
        )

    status = request.args.get('status', '')
    if status:
        mapped_filter_status = _map_status_for_filter(status)
        if mapped_filter_status == 'REQUEST':
            query = query.filter(
                InboundRequest.status == 'REQUEST',
                InboundRequest.pending_invoice_queue.isnot(True),
            )
        elif mapped_filter_status == 'CONFIRMED':
            query = query.filter(
                InboundRequest.status.in_(['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS']),
                InboundRequest.pending_invoice_queue.isnot(True),
            )
        elif mapped_filter_status == 'INVOICED':
            query = query.filter(
                InboundRequest.status.in_(['INVOICE', 'COMPLETED', 'INVOICED']),
                InboundRequest.pending_invoice_queue.isnot(True),
            )
        elif mapped_filter_status == 'DELETED':
            query = query.filter(InboundRequest.pending_invoice_queue.is_(True))

    return query, queue


def _map_status_for_filter(status_val):
    """Map old statuses to new 3-state system."""
    if not status_val:
        return None
    status_upper = str(status_val).upper()
    if status_upper in ['REQUEST']:
        return 'REQUEST'
    if status_upper in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS']:
        return 'CONFIRMED'
    if status_upper in ['INVOICE', 'COMPLETED', 'INVOICED']:
        return 'INVOICED'
    if status_upper in ['DELETED']:
        return 'DELETED'
    return None


@inbound_bp.route('/')
@login_required
def index():
    """List all inbound requests with filtering and run-down plan"""
    ctx = _inbound_list_context()
    return render_template(
        'inbound/index.html',
        inbound_hide_run_down=False,
        **ctx,
    )


@inbound_bp.route('/export-list-excel')
@login_required
def export_list_excel():
    """Export inbound list using current filters."""
    import io

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Excel export requires openpyxl package', 'error')
        return redirect(request.referrer or url_for('inbound.index'))

    all_files_view = (request.args.get('all_files_view') or '').strip() == '1'
    query, queue = _build_inbound_list_query(all_files_view)

    req_sort = (request.args.get('req_sort') or 'desc').strip().lower()
    if req_sort not in ('asc', 'desc'):
        req_sort = 'desc'
    if req_sort == 'desc':
        query = query.order_by(InboundRequest.created_at.desc(), InboundRequest.id.desc())
    else:
        query = query.order_by(InboundRequest.created_at.asc(), InboundRequest.id.asc())

    rows = query.all()

    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Inbound Requests"

    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    headers = ['Request No.', 'Agent', 'Contact', 'Nationality', 'Pax', 'Travel Dates', 'Days']
    if all_files_view:
        headers.append('Status')

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    def _status_label(inb):
        if queue == 'deleted' or getattr(inb, 'pending_invoice_queue', False):
            return 'Deleted'
        status_val = str(inb.status or '').upper()
        if status_val == 'REQUEST':
            return 'Request'
        if status_val in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS']:
            return 'Confirmed'
        if status_val in ['INVOICE', 'COMPLETED', 'INVOICED']:
            return 'Invoiced'
        return status_val.title() if status_val else ''

    row_idx = 2
    for inb in rows:
        values = [
            inb.request_number,
            inb.agent or 'TBA',
            inb.contact_name or 'TBA',
            inb.nationality or 'TBA',
            inb.pax or 0,
            (
                f"{inb.from_date.strftime('%d %b')} - {inb.to_date.strftime('%d %b %Y')}"
                if inb.from_date and inb.to_date else ''
            ),
            inb.no_of_days or 0,
        ]
        if all_files_view:
            values.append(_status_label(inb))

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
        row_idx += 1

    widths = [18, 24, 22, 15, 8, 24, 8, 14]
    for idx, width in enumerate(widths[:len(headers)], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Inbound_Requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )


@inbound_bp.route('/print-list')
@login_required
def print_list():
    """Print-friendly view of inbound list using current filters."""
    all_files_view = (request.args.get('all_files_view') or '').strip() == '1'
    query, queue = _build_inbound_list_query(all_files_view)

    req_sort = (request.args.get('req_sort') or 'desc').strip().lower()
    if req_sort not in ('asc', 'desc'):
        req_sort = 'desc'
    if req_sort == 'desc':
        query = query.order_by(InboundRequest.created_at.desc(), InboundRequest.id.desc())
    else:
        query = query.order_by(InboundRequest.created_at.asc(), InboundRequest.id.asc())

    rows = query.all()
    scope_label = _hub_status_scope_label(request.args.get('status'))
    if queue == 'deleted':
        page_title = 'All inbound files — Deleted' if all_files_view else 'Inbound tour — Deleted'
    elif scope_label:
        page_title = (
            f'All inbound files — {scope_label}'
            if all_files_view
            else f'Inbound tour — {scope_label}'
        )
    else:
        page_title = 'All inbound files' if all_files_view else 'Inbound Tour Requests'

    is_invoiced_status_view = (request.args.get('status') or '').strip().upper() == 'INVOICED'

    return render_template(
        'inbound/list_print.html',
        rows=rows,
        all_files_view=all_files_view,
        queue=queue,
        filter_summary=_inbound_list_filter_summary(),
        page_title=page_title,
        printed_at=datetime.now(),
        is_invoiced_status_view=is_invoiced_status_view,
        row_status_label=_inbound_row_status_label,
    )


def get_run_down_data_by_date():
    """Get confirmed itineraries grouped by date with activities"""
    from app.models.customer import Customer
    from sqlalchemy.orm import joinedload

    # Get date range (next 30 days)
    today = datetime.now().date()
    date_to = today + timedelta(days=30)

    # Get all confirmed requests with their itinerary rows (exclude Deleted queue)
    # Use eager loading to fetch customers in a single query
    confirmed_requests = InboundRequest.query.filter(
        InboundRequest.status.in_(['CONFIRMED', 'BOOKED']),
        InboundRequest.pending_invoice_queue.isnot(True),
    ).options(
        joinedload(InboundRequest.customer),
        joinedload(InboundRequest.itinerary_rows)
    ).all()

    # Group activities by date
    activities_by_date = {}

    for req in confirmed_requests:
        # Get customer info (already loaded via eager loading)
        customer_name = "TBA"
        if req.customer:
            customer_name = req.customer.name
        elif req.contact_name:
            customer_name = req.contact_name

        # Process itinerary rows
        for row in req.itinerary_rows:
            if row.date < today or row.date > date_to:
                continue

            date_key = row.date.strftime('%Y-%m-%d')

            if date_key not in activities_by_date:
                activities_by_date[date_key] = {
                    'date': row.date,
                    'date_formatted': row.date.strftime('%A, %B %d, %Y'),
                    'activities': []
                }

            # Build activity info with detailed service data
            services = []
            base_cost = row.base_cost or 0

            if row.flag_hotel:
                # Build room breakdown
                room_details = []
                if row.hotel_single_rooms > 0:
                    room_details.append(f"{row.hotel_single_rooms} Single")
                if row.hotel_double_rooms > 0:
                    room_details.append(f"{row.hotel_double_rooms} Double")
                if row.hotel_triple_rooms > 0:
                    room_details.append(f"{row.hotel_triple_rooms} Triple")
                if row.hotel_other_rooms > 0:
                    room_details.append(f"{row.hotel_other_rooms} Other")

                services.append({
                    'type': 'HOTEL',
                    'icon': 'fa-hotel',
                    'description': row.description or 'Hotel Service',
                    'cost': base_cost,
                    'rooms': ', '.join(room_details) if room_details else 'Rooms TBA',
                    'cost_unit': row.cost_unit
                })
            if row.flag_transport:
                services.append({
                    'type': 'TRANSPORT',
                    'icon': 'fa-bus',
                    'description': row.description or 'Transport Service',
                    'cost': base_cost,
                    'cost_unit': row.cost_unit,
                    'pax': req.pax
                })
            if row.flag_meal:
                services.append({
                    'type': 'MEAL',
                    'icon': 'fa-utensils',
                    'description': row.description or 'Meal Service',
                    'cost': base_cost,
                    'cost_unit': row.cost_unit,
                    'pax': req.pax
                })
            if row.flag_guide:
                services.append({
                    'type': 'GUIDE',
                    'icon': 'fa-user-tie',
                    'description': row.description or 'Guide Service',
                    'cost': base_cost,
                    'cost_unit': row.cost_unit,
                    'pax': req.pax
                })
            if row.flag_airport:
                services.append({
                    'type': 'AIRPORT',
                    'icon': 'fa-plane',
                    'description': row.description or 'Airport Service',
                    'cost': 0,
                    'pax': req.pax
                })

            if services:  # Only add if there are services
                activities_by_date[date_key]['activities'].append({
                    'request_number': req.request_number,
                    'request_id': req.id,
                    'customer_name': customer_name,
                    'pax': req.pax,
                    'services': services,
                    'status': req.status,
                    'status_color': get_status_color(req.status)
                })

    # Sort by date
    sorted_data = sorted(activities_by_date.values(), key=lambda x: x['date'])
    return sorted_data

@inbound_bp.route('/new')
@login_required
def new_request():
    """Create new inbound request and go directly to itinerary creation"""
    try:
        return _new_request_impl()
    except Exception as e:
        import traceback
        from flask import current_app
        current_app.logger.error(f"new_request FAILED: {e}\n{traceback.format_exc()}")
        traceback.print_exc()
        flash(f'Failed to create new request: {str(e)}', 'error')
        return redirect(url_for('inbound.index'))

def _ensure_default_inbound_user():
    from app.models.user import User, create_test_data

    if User.query.get(1) is None:
        create_test_data()
        db.session.commit()
    if User.query.get(1) is None:
        raise RuntimeError("Default user (id=1) is required but could not be created")


def _create_default_itinerary_rows(request_obj, from_date, to_date):
    current_date = from_date
    day_counter = 1
    total_days = (to_date - from_date).days + 1

    while current_date <= to_date:
        if day_counter == 1:
            description = "Arrival Day"
        elif current_date == to_date:
            description = "Departure Day"
        else:
            description = f"Day {day_counter}"

        row = ItineraryRow(
            request_id=request_obj.id,
            date=current_date,
            description=description,
            flag_hotel=(day_counter != total_days),
            flag_transport=True,
            flag_guide=(day_counter > 1 and day_counter < total_days),
            flag_meal=(day_counter > 1 and day_counter < total_days),
            flag_airport=(day_counter == 1 or day_counter == total_days),
        )
        db.session.add(row)
        current_date += timedelta(days=1)
        day_counter += 1


def _create_inbound_request_draft(parent_request=None, link_note=None):
    """Create a new inbound request draft, optionally linked to a main file."""
    import uuid

    _ensure_default_inbound_user()

    is_linked = parent_request is not None
    if is_linked:
        from_date = parent_request.from_date or datetime.now().date()
        to_date = parent_request.to_date or (from_date + timedelta(days=3))
        request_number = InboundRequest.generate_linked_request_number(parent_request)
        is_saved = True
    else:
        from_date = datetime.now().date()
        to_date = (datetime.now() + timedelta(days=3)).date()
        request_number = f"IN-NEW-{str(uuid.uuid4())[:6].upper()}"
        is_saved = False

    request_obj = InboundRequest(
        request_number=request_number,
        from_date=from_date,
        to_date=to_date,
        customer_type='GROUP',
        contact_name='TBA',
        nationality='TBA',
        pax=1,
        user_id=1,
        status=STATUS_REQUEST,
        is_saved=is_saved,
    )

    if parent_request:
        request_obj.parent_request_id = parent_request.id
        request_obj.link_note = (link_note or '').strip() or None
        request_obj.customer_id = parent_request.customer_id
        request_obj.contact_name = parent_request.contact_name or 'TBA'
        request_obj.nationality = parent_request.nationality or 'TBA'
        request_obj.pax = parent_request.pax or 1
        request_obj.customer_type = parent_request.customer_type or 'GROUP'
        request_obj.agent_ref = parent_request.agent_ref

    request_obj.calculate_days()
    db.session.add(request_obj)
    db.session.flush()
    _create_default_itinerary_rows(request_obj, from_date, to_date)
    db.session.commit()
    return request_obj


def _main_files_for_link_query():
    """Saved main inbound files eligible as link parents."""
    from sqlalchemy import or_

    return InboundRequest.query.filter(
        InboundRequest.pending_invoice_queue.isnot(True),
        InboundRequest.parent_request_id.is_(None),
        or_(
            InboundRequest.is_saved.is_(True),
            ~InboundRequest.request_number.like('IN-NEW-%'),
        ),
    )


def _new_request_impl():
    """Implementation of new inbound request creation"""
    request_obj = _create_inbound_request_draft()
    return redirect(url_for('inbound.view_request', id=request_obj.id))


@inbound_bp.route('/api/main-files-for-link')
def api_main_files_for_link():
    """Search main inbound files that can be linked as parents."""
    from sqlalchemy import or_, func

    q = (request.args.get('q') or '').strip()
    query = _main_files_for_link_query()
    if q:
        like = f'%{q}%'
        cust_full = func.trim(
            func.concat(
                Customer.first_name,
                ' ',
                func.coalesce(Customer.last_name, ''),
            )
        )
        query = query.outerjoin(Customer, InboundRequest.customer_id == Customer.id).filter(
            or_(
                InboundRequest.request_number.ilike(like),
                InboundRequest.contact_name.ilike(like),
                InboundRequest.agent_ref.ilike(like),
                cust_full.ilike(like),
            )
        )
    results = query.order_by(InboundRequest.updated_at.desc()).limit(25).all()
    return jsonify([
        {
            'id': row.id,
            'request_number': row.request_number,
            'document_sequence': row.document_sequence,
            'contact_name': row.contact_name,
            'agent': row.agent,
            'status': row.status,
            'from_date': row.from_date.isoformat() if row.from_date else None,
            'to_date': row.to_date.isoformat() if row.to_date else None,
            'pax': row.pax,
        }
        for row in results
    ])


def _query_linked_attachment_rows(search_query=''):
    """Return (child, parent) rows for linked attachment listings."""
    from sqlalchemy import or_
    from sqlalchemy.orm import aliased

    Parent = aliased(InboundRequest)
    query = (
        db.session.query(InboundRequest, Parent)
        .join(Parent, InboundRequest.parent_request_id == Parent.id)
        .filter(
            InboundRequest.parent_request_id.isnot(None),
            InboundRequest.pending_invoice_queue.isnot(True),
        )
    )
    search_query = (search_query or '').strip()
    if search_query:
        like = f'%{search_query}%'
        query = query.filter(
            or_(
                InboundRequest.request_number.ilike(like),
                Parent.request_number.ilike(like),
                InboundRequest.contact_name.ilike(like),
                InboundRequest.link_note.ilike(like),
            )
        )
    return (
        query.order_by(Parent.request_number.asc(), InboundRequest.created_at.desc())
        .limit(200)
        .all()
    )


def _group_linked_attachment_rows(rows):
    """Group linked attachment rows by parent request."""
    groups = []
    by_parent = {}
    for child, parent in rows:
        bucket = by_parent.get(parent.id)
        if not bucket:
            bucket = {'parent': parent, 'attachments': []}
            by_parent[parent.id] = bucket
            groups.append(bucket)
        bucket['attachments'].append(child)
    return groups


@inbound_bp.route('/linked-attachments')
@login_required
def linked_attachments_page():
    """Dedicated page listing files linked to main inbound records."""
    q = (request.args.get('q') or '').strip()
    rows = _query_linked_attachment_rows(q)
    groups = _group_linked_attachment_rows(rows)
    return render_template(
        'inbound/linked_attachments.html',
        groups=groups,
        query=q,
        total_count=sum(len(g['attachments']) for g in groups),
    )


@inbound_bp.route('/linked-attachments/create', methods=['POST'])
@login_required
def linked_attachments_create():
    """Create a linked attachment from the dedicated page form."""
    parent_id = request.form.get('parent_id', type=int)
    link_note = (request.form.get('link_note') or '').strip()

    if not parent_id:
        flash('Please select a main file.', 'error')
        return redirect(url_for('inbound.linked_attachments_page'))

    parent = _main_files_for_link_query().filter(InboundRequest.id == parent_id).first()
    if not parent:
        flash('Main file not found or not eligible for linking.', 'error')
        return redirect(url_for('inbound.linked_attachments_page'))

    try:
        child = _create_inbound_request_draft(
            parent_request=parent,
            link_note=link_note,
        )
        flash(f'Linked file {child.request_number} created.', 'success')
        return redirect(url_for('inbound.view_request', id=child.id, mode='edit'))
    except Exception as e:
        db.session.rollback()
        flash(f'Could not create linked file: {e}', 'error')
        return redirect(url_for('inbound.linked_attachments_page'))


@inbound_bp.route('/api/linked-attachments')
def api_linked_attachments():
    """List inbound requests linked to main files (attached / related files)."""
    q = (request.args.get('q') or '').strip()
    rows = _query_linked_attachment_rows(q)
    return jsonify([
        {
            'id': child.id,
            'request_number': child.request_number,
            'link_note': child.link_note,
            'status': child.status,
            'contact_name': child.contact_name,
            'created_at': child.created_at.isoformat() if child.created_at else None,
            'parent_id': parent.id,
            'parent_request_number': parent.request_number,
            'parent_document_sequence': parent.document_sequence,
            'view_url': url_for('inbound.view_request', id=child.id, mode='view'),
            'edit_url': url_for('inbound.view_request', id=child.id, mode='edit'),
            'parent_view_url': url_for('inbound.view_request', id=parent.id, mode='view'),
        }
        for child, parent in rows
    ])


@inbound_bp.route('/api/create-linked-request', methods=['POST'])
@csrf.exempt
def api_create_linked_request():
    """Create a new inbound request linked to an existing main file."""
    data = request.get_json(silent=True) or {}
    parent_id = data.get('parent_id')
    link_note = (data.get('link_note') or '').strip()

    if not parent_id:
        return jsonify({'success': False, 'message': 'Please select a main file'}), 400

    parent = _main_files_for_link_query().filter(InboundRequest.id == int(parent_id)).first()
    if not parent:
        return jsonify({'success': False, 'message': 'Main file not found or not eligible for linking'}), 404

    try:
        child = _create_inbound_request_draft(
            parent_request=parent,
            link_note=link_note,
        )
        return jsonify({
            'success': True,
            'request_id': child.id,
            'request_number': child.request_number,
            'redirect_url': url_for('inbound.view_request', id=child.id, mode='edit'),
            'parent_request_number': parent.request_number,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/<int:id>/edit')
@login_required
def edit_request(id):
    """Redirect to unified view/edit page"""
    request_obj = InboundRequest.query.get_or_404(id)

    # Check ownership
    if request_obj.user_id != 1:
        flash('Access denied.', 'error')
        return redirect(url_for('inbound.index'))

    # Redirect to the unified view_request page in edit mode
    return redirect(url_for('inbound.view_request', id=id, mode='edit'))

@inbound_bp.route('/<int:id>/view')
@login_required
def view_request(id):
    """View inbound request details with unified edit functionality"""
    from flask import request as flask_request

    # Use eager loading for all service relationships to ensure they're loaded for template
    # Include subqueryload for hotel rooms to avoid N+1 queries
    from sqlalchemy.orm import subqueryload
    request_obj = InboundRequest.query.options(
        selectinload(InboundRequest.inbound_hotels).subqueryload(InboundHotel.rooms),
        selectinload(InboundRequest.inbound_transports),
        selectinload(InboundRequest.inbound_guides),
        selectinload(InboundRequest.inbound_meals),
        selectinload(InboundRequest.inbound_optionals),
        selectinload(InboundRequest.itinerary_rows),
        selectinload(InboundRequest.arrival_batches),
        selectinload(InboundRequest.departure_batches)
    ).get_or_404(id)

    # Get mode parameter (view or edit)
    mode = flask_request.args.get('mode', 'edit')  # Default to edit for backward compatibility
    view_only = (mode == 'view')

    # Load supplier dropdown data from short-lived cache to keep navigation fast.
    try:
        dropdowns = _get_supplier_dropdown_data()
        hotel_suppliers = dropdowns['hotel_suppliers']
        sorted_hotels_by_city = dropdowns['hotels_by_city']
        accommodation_cities = dropdowns['accommodation_cities']
        accommodation_categories = dropdowns['accommodation_categories']
        transport_suppliers = dropdowns['transport_suppliers']
        guide_suppliers = dropdowns['guide_suppliers']
        representatives = dropdowns['representatives']
        restaurant_suppliers = dropdowns['restaurant_suppliers']
        ground_handler_suppliers = dropdowns['ground_handler_suppliers']
    except Exception as e:
        # Fallback to empty lists if query fails
        import traceback
        print(f"[ERROR] Failed to load suppliers: {e}")
        traceback.print_exc()
        hotel_suppliers = []
        sorted_hotels_by_city = {}
        accommodation_cities = []
        accommodation_categories = []
        transport_suppliers = []
        guide_suppliers = []
        representatives = []
        restaurant_suppliers = []
        ground_handler_suppliers = []

    # Sort itinerary rows ensuring child rows appear directly below their parent
    sorted_itinerary_rows = sort_itinerary_rows_with_children(request_obj.itinerary_rows)
    
    itinerary_guide_slot_saved = _itinerary_guide_slot_saved_map(request_obj)
    itin_guide_slots = sum(
        len(itinerary_row_guide_supplier_id_list(r)) for r in (request_obj.itinerary_rows or [])
    )
    guide_tab_badge_count = max(len(request_obj.inbound_guides or []), itin_guide_slots)
    trip_summary_transports = _trip_summary_transports(request_obj.inbound_transports)
    parent_request = request_obj.parent_request if request_obj.parent_request_id else None
    linked_children = []
    if not request_obj.parent_request_id:
        linked_children = request_obj.linked_requests.order_by(InboundRequest.created_at.desc()).all()
    return render_template('inbound/view_request.html',
                           request=request_obj,
                           inbound_request=request_obj,  # Explicit variable to avoid Flask request confusion
                           parent_request=parent_request,
                           linked_children=linked_children,
                           view_only=view_only,
                           rows=sorted_itinerary_rows,
                           hotel_suppliers=hotel_suppliers,
                           hotels_by_city=sorted_hotels_by_city,
                           accommodation_cities=accommodation_cities,
                           accommodation_categories=accommodation_categories,
                           transport_suppliers=transport_suppliers,
                           guide_suppliers=guide_suppliers,
                           representatives=representatives,
                           restaurant_suppliers=restaurant_suppliers,
                           ground_handler_suppliers=ground_handler_suppliers,
                           itinerary_guide_slot_saved=itinerary_guide_slot_saved,
                           guide_tab_badge_count=guide_tab_badge_count,
                           trip_summary_transports=trip_summary_transports)


@inbound_bp.route('/<int:id>/delete', methods=['GET', 'POST'])
@login_required
@csrf.exempt
def delete_request(id):
    """Permanently delete a request and all of its linked inbound records.
    Returns JSON when called with AJAX (X-Requested-With: XMLHttpRequest or Accept: application/json).
    """
    request_obj = InboundRequest.query.get_or_404(id)
    is_ajax = (
        request.headers.get('X-Requested-With') == 'XMLHttpRequest'
        or 'application/json' in request.headers.get('Accept', '')
    )

    try:
        # Ensure invoices no longer reference this request before delete.
        Invoice.query.filter_by(inbound_request_id=request_obj.id).update(
            {'inbound_request_id': None},
            synchronize_session=False
        )
        db.session.delete(request_obj)
        db.session.commit()
        if is_ajax:
            return jsonify({'success': True})
        flash('Request deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        if is_ajax:
            return jsonify({'success': False, 'message': str(e)}), 500
        flash(f'Error deleting request: {str(e)}', 'error')

    return redirect(url_for('inbound.index'))


@inbound_bp.route('/api/<int:request_id>/trash', methods=['POST'])
@csrf.exempt
def api_trash_inbound_request(request_id):
    """Move request to Deleted queue (soft delete from main list)."""
    request_obj = InboundRequest.query.get(request_id)
    if not request_obj:
        return jsonify({'success': False, 'message': 'Request not found'}), 404
    data = request.get_json(silent=True) or {}
    reason = (data.get('reason') or '').strip()
    if not reason:
        return jsonify({'success': False, 'message': 'Deletion reason is required'}), 400
    try:
        request_obj.pending_invoice_queue = True
        request_obj.deleted_reason = reason
        db.session.commit()
        return jsonify({
            'success': True,
            'message': 'Request moved to Deleted queue',
            'pending_invoice_queue': True,
            'deleted_reason': reason,
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500


@inbound_bp.route('/<int:id>/restore-queue')
@login_required
def restore_from_invoice_queue(id):
    """Put request back on the main inbound list (undo trash / Deleted queue)."""
    request_obj = InboundRequest.query.get_or_404(id)

    try:
        request_obj.pending_invoice_queue = False
        request_obj.deleted_reason = None
        db.session.commit()
        flash(f'Request {request_obj.request_number} is back on the main list.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error restoring request: {str(e)}', 'error')

    return redirect(url_for('inbound.index', queue='deleted'))


# API Route for updating request details
@inbound_bp.route('/api/<int:request_id>/update', methods=['POST'])
@csrf.exempt

def api_update_request(request_id):
    """Update inbound request master details"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        data = request.get_json()

        # Update master details
        # Properly handle customer_id - convert to int if provided, keep existing if not provided
        customer_id_value = data.get('customer_id')
        if customer_id_value and str(customer_id_value).strip():
            try:
                request_obj.customer_id = int(customer_id_value)
            except (ValueError, TypeError):
                pass  # Keep existing value if conversion fails
        elif customer_id_value == '' or customer_id_value is None:
            # Only reset if explicitly set to empty
            if 'customer_id' in data:
                request_obj.customer_id = None
        request_obj.customer_type = data.get('customer_type', request_obj.customer_type)
        request_obj.contact_name = data.get('contact_name', request_obj.contact_name)
        request_obj.agent_ref = data.get('agent_ref', request_obj.agent_ref)
        request_obj.nationality = data.get('nationality', request_obj.nationality)
        request_obj.pax = int(data.get('pax', request_obj.pax))
        request_obj.special_note = data.get('special_note', request_obj.special_note)

        # Update arrival/departure details
        request_obj.arrival_point = data.get('arrival_point', request_obj.arrival_point)
        request_obj.departure_point = data.get('departure_point', request_obj.departure_point)

        # Handle arrival/departure time updates
        if data.get('arrival_time'):
            try:
                request_obj.arrival_time = datetime.strptime(data.get('arrival_time'), '%H:%M').time()
            except:
                pass  # Invalid time format, skip
        elif data.get('arrival_time') == '':
            request_obj.arrival_time = None

        if data.get('departure_time'):
            try:
                request_obj.departure_time = datetime.strptime(data.get('departure_time'), '%H:%M').time()
            except:
                pass  # Invalid time format, skip
        elif data.get('departure_time') == '':
            request_obj.departure_time = None

        # Update dates
        from_date_str = data.get('from_date')
        to_date_str = data.get('to_date')
        if from_date_str:
            request_obj.from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        if to_date_str:
            request_obj.to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

        # Recalculate days if dates changed
        if from_date_str or to_date_str:
            request_obj.calculate_days()

        # Generate document sequence on save if not already assigned
        request_obj.assign_document_sequence()

        db.session.commit()
        return jsonify({'success': True, 'message': 'Request updated successfully', 'document_sequence': request_obj.document_sequence})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/update-restaurant-voucher-note', methods=['POST'])
@csrf.exempt
def api_update_restaurant_voucher_note(request_id):
    """Update Note and Special Request for restaurant voucher only (per-voucher isolation)"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403
    try:
        data = request.get_json() or {}
        request_obj.restaurant_voucher_note = data.get('note', '')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Restaurant voucher note saved'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/update-hotel-voucher-note', methods=['POST'])
@csrf.exempt
def api_update_hotel_voucher_note(request_id):
    """Update Note and Special Request for hotel voucher only (per-voucher isolation)"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403
    try:
        data = request.get_json() or {}
        request_obj.hotel_voucher_note = data.get('note', '')
        db.session.commit()
        return jsonify({'success': True, 'message': 'Hotel voucher note saved'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/advance-expense-sheet', methods=['GET'])
def api_get_advance_expense_sheet(request_id):
    """Get saved advance expense sheet data for request"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.advance_expense_sheet_data:
        try:
            return jsonify(json.loads(request_obj.advance_expense_sheet_data))
        except (ValueError, TypeError):
            return jsonify({})
    return jsonify({})

@inbound_bp.route('/api/<int:request_id>/save-advance-expense-sheet', methods=['POST'])
@csrf.exempt
def api_save_advance_expense_sheet(request_id):
    """Save advance expense sheet data for request (no user check)"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    try:
        data = request.get_json() or {}
        request_obj.advance_expense_sheet_data = json.dumps(data)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Advance expense sheet saved'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/closing-guide-payment-sheet', methods=['GET'])
def api_get_closing_guide_payment_sheet(request_id):
    """Get saved closing guide payment sheet data for request"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.closing_guide_payment_sheet_data:
        try:
            return jsonify(json.loads(request_obj.closing_guide_payment_sheet_data))
        except (ValueError, TypeError):
            return jsonify({})
    return jsonify({})

@inbound_bp.route('/api/<int:request_id>/save-closing-guide-payment-sheet', methods=['POST'])
@csrf.exempt
def api_save_closing_guide_payment_sheet(request_id):
    """Save closing guide payment sheet data for request (no user check)"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    try:
        data = request.get_json() or {}
        if not hasattr(request_obj, 'closing_guide_payment_sheet_data'):
            return jsonify({'success': False, 'message': 'Database schema outdated. Please restart the server to add the closing_guide_payment_sheet_data column.'}), 500
        request_obj.closing_guide_payment_sheet_data = json.dumps(data)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Closing guide payment sheet saved'})
    except Exception as e:
        db.session.rollback()
        err_msg = str(e)
        if 'no such column' in err_msg.lower() or 'closing_guide_payment' in err_msg.lower():
            err_msg = 'Database schema outdated. Please restart the server to apply the latest migration.'
        return jsonify({'success': False, 'message': err_msg}), 500

# API Route for saving itinerary
@inbound_bp.route('/api/<int:request_id>/save-itinerary', methods=['POST'])
@csrf.exempt

def api_save_itinerary(request_id):
    """Save itinerary data for inbound request - auto-generates days if empty"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        data = request.get_json()
        rows_data = data.get('rows', [])

        print(f"DEBUG: Saving itinerary for request {request_id}")
        print(f"DEBUG: Received {len(rows_data)} rows")

        # Auto-generate days if no rows provided but we have dates
        if not rows_data and request_obj.from_date and request_obj.to_date:
            print("DEBUG: No rows provided, auto-generating days from date range")

            # Generate one row per day
            current_date = request_obj.from_date
            day_counter = 1

            while current_date <= request_obj.to_date:
                row_data = {
                    'date': current_date.strftime('%Y-%m-%d'),
                    'description': f'Day {day_counter} - {current_date.strftime("%A, %B %d")}',
                    'base_cost': 0.0,
                    'currency': request_obj.total_currency or 'USD',
                    'cost_unit': 'PER_PERSON',
                    'flag_hotel': False,
                    'flag_guide': False,
                    'flag_transport': False,
                    'flag_meal': False,
                    'flag_airport': False,
                    'hotel_single_rooms': 0,
                    'hotel_double_rooms': 0,
                    'hotel_triple_rooms': 0,
                    'hotel_other_rooms': 0
                }
                rows_data.append(row_data)

                current_date += timedelta(days=1)
                day_counter += 1

            print(f"DEBUG: Auto-generated {len(rows_data)} rows")

        # Clear existing itinerary rows
        deleted_count = ItineraryRow.query.filter_by(request_id=request_id).delete()
        print(f"DEBUG: Deleted {deleted_count} existing rows")

        # Add new rows
        for i, row_data in enumerate(rows_data):
            print(f"DEBUG: Processing row {i}: {row_data}")
            row = ItineraryRow(
                request_id=request_id,
                date=datetime.strptime(row_data['date'], '%Y-%m-%d').date(),
                description=row_data['description'],
                base_cost=float(row_data['base_cost']) if row_data['base_cost'] else 0.0,
                currency=row_data['currency'],
                cost_unit=row_data['cost_unit'],
                flag_hotel=row_data.get('flag_hotel', False),
                flag_guide=row_data.get('flag_guide', False),
                flag_transport=row_data.get('flag_transport', False),
                flag_meal=row_data.get('flag_meal', False),
                flag_airport=row_data.get('flag_airport', False),
                hotel_single_rooms=int(row_data.get('hotel_single_rooms', 0)),
                hotel_double_rooms=int(row_data.get('hotel_double_rooms', 0)),
                hotel_triple_rooms=int(row_data.get('hotel_triple_rooms', 0)),
                hotel_other_rooms=int(row_data.get('hotel_other_rooms', 0))
            )
            db.session.add(row)
            print(f"DEBUG: Added row {i} to session")

        # Recalculate totals
        request_obj.calculate_total()

        db.session.commit()
        print("DEBUG: Successfully saved itinerary")
        return jsonify({'success': True, 'message': 'Itinerary saved successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# API Routes for AJAX operations
@inbound_bp.route('/api/<int:request_id>/itinerary', methods=['GET'])
@csrf.exempt

def api_get_itinerary(request_id):
    """Get itinerary rows for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    rows = []
    for row in request_obj.itinerary_rows:
        rows.append({
            'id': row.id,
            'date': row.date.isoformat(),
            'description': row.description,
            'base_cost': row.base_cost,
            'cost_unit': row.cost_unit,
            'currency': row.currency,
            'flag_hotel': row.flag_hotel,
            'flag_guide': row.flag_guide,
            'flag_transport': row.flag_transport,
            'flag_meal': row.flag_meal,
            'flag_airport': row.flag_airport,
            'flag_drive': getattr(row, 'flag_drive', False),
            'hotel_single_rooms': getattr(row, 'hotel_single_rooms', 0),
            'hotel_double_rooms': getattr(row, 'hotel_double_rooms', 0),
            'hotel_triple_rooms': getattr(row, 'hotel_triple_rooms', 0),
            'hotel_other_rooms': getattr(row, 'hotel_other_rooms', 0),
            'row_cost': row.calculate_row_cost(request_obj.pax)
        })

    # Handle arrival/departure times - could be datetime.time, string, or None
    arrival_time_str = ''
    if request_obj.arrival_time:
        if hasattr(request_obj.arrival_time, 'strftime'):
            arrival_time_str = request_obj.arrival_time.strftime('%H:%M')
        elif isinstance(request_obj.arrival_time, str):
            arrival_time_str = request_obj.arrival_time

    departure_time_str = ''
    if request_obj.departure_time:
        if hasattr(request_obj.departure_time, 'strftime'):
            departure_time_str = request_obj.departure_time.strftime('%H:%M')
        elif isinstance(request_obj.departure_time, str):
            departure_time_str = request_obj.departure_time

    return jsonify({
        'rows': rows,
        'total': request_obj.calculate_total(),
        'arrival_point': request_obj.arrival_point or '',
        'arrival_time': arrival_time_str,
        'departure_point': request_obj.departure_point or '',
        'departure_time': departure_time_str
    })

@inbound_bp.route('/api/<int:request_id>/itinerary', methods=['POST'])
@csrf.exempt

def api_save_itinerary_original(request_id):
    """Save itinerary rows for a request (original version)"""
    print(f"[DEBUG] api_save_itinerary_original called for request_id: {request_id}")

    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        if request_obj.user_id != 1:
            print(f"[DEBUG] Access denied: user {1} != owner {request_obj.user_id}")
            return jsonify({'error': 'Access denied'}), 403

        data = request.get_json()
        if not data:
            print("[DEBUG] No JSON data received")
            return jsonify({'success': False, 'message': 'No data received'}), 400

        print(f"[DEBUG] Received data keys: {data.keys() if data else 'None'}")
        rows_data = data.get('rows', [])
        print(f"[DEBUG] Number of rows to save: {len(rows_data)}")

        # Update arrival/departure details if provided
        if 'arrival_point' in data:
            request_obj.arrival_point = data.get('arrival_point') or None
        if 'departure_point' in data:
            request_obj.departure_point = data.get('departure_point') or None
        if 'arrival_time' in data and data.get('arrival_time'):
            try:
                request_obj.arrival_time = datetime.strptime(data.get('arrival_time'), '%H:%M').time()
            except:
                request_obj.arrival_time = None
        elif 'arrival_time' in data and not data.get('arrival_time'):
            request_obj.arrival_time = None
        if 'departure_time' in data and data.get('departure_time'):
            try:
                request_obj.departure_time = datetime.strptime(data.get('departure_time'), '%H:%M').time()
            except:
                request_obj.departure_time = None
        elif 'departure_time' in data and not data.get('departure_time'):
            request_obj.departure_time = None

        # Update new arrival/departure fields
        if 'visa_type' in data:
            request_obj.visa_type = data.get('visa_type') or 'NOT_INCLUDED'
        if 'arrival_driver_name' in data:
            request_obj.arrival_driver_name = data.get('arrival_driver_name') or None
        if 'meeting_assistance' in data:
            # Properly parse boolean from various input types
            ma_value = data.get('meeting_assistance')
            if isinstance(ma_value, bool):
                request_obj.meeting_assistance = ma_value
            elif isinstance(ma_value, str):
                request_obj.meeting_assistance = ma_value.lower() in ('true', '1', 'yes')
            elif isinstance(ma_value, (int, float)):
                request_obj.meeting_assistance = bool(ma_value)
            else:
                request_obj.meeting_assistance = False
        if 'departure_tax' in data:
            request_obj.departure_tax = data.get('departure_tax') or 'NOT_INCLUDED'

        # Update or create rows using row IDs for matching (handles multiple rows per date)
        # Get existing rows indexed by ID
        existing_rows_dict = {row.id: row for row in ItineraryRow.query.filter_by(request_id=request_id).all()}
        submitted_ids = set()

        # Import models needed for service deletion
        from app.models.inbound import InboundHotel, InboundTransport, InboundMeal, InboundGuide

        # Process each row from the submitted data
        for row_data in rows_data:
            row_date = datetime.strptime(row_data['date'], '%Y-%m-%d').date()
            row_id = row_data.get('id')  # May be None for new rows

            # Update existing row or create new one
            if row_id and row_id in existing_rows_dict:
                # Update existing row by ID
                submitted_ids.add(row_id)
                row = existing_rows_dict[row_id]
                # Update all fields
                row.date = row_date
                row.description = row_data['description']
                row.restaurant = row_data.get('restaurant', '')
                row.cash_expense = float(row_data.get('cash_expense', 0))
                row.comment = row_data.get('comment', '')
                row.base_cost = float(row_data.get('base_cost', 0))
                row.cost_unit = row_data.get('cost_unit', COST_UNIT_PER_PERSON)
                row.currency = row_data.get('currency', 'USD')
                row.flag_hotel = row_data.get('flag_hotel', False)
                row.flag_guide = row_data.get('flag_guide', False)
                row.flag_transport = row_data.get('flag_transport', False)
                row.flag_meal = row_data.get('flag_meal', False)
                row.flag_airport = row_data.get('flag_airport', False)
                row.flag_drive = row_data.get('flag_drive', False)
                row.hotel_single_rooms = int(row_data.get('hotel_single_rooms', 0))
                row.hotel_double_rooms = int(row_data.get('hotel_double_rooms', 0))
                row.hotel_triple_rooms = int(row_data.get('hotel_triple_rooms', 0))
                row.hotel_other_rooms = int(row_data.get('hotel_other_rooms', 0))
            else:
                # Create new row (no ID or ID not found)
                row = ItineraryRow(
                    request_id=request_id,
                    date=row_date,
                    description=row_data['description'],
                    restaurant=row_data.get('restaurant', ''),
                    cash_expense=float(row_data.get('cash_expense', 0)),
                    comment=row_data.get('comment', ''),
                    base_cost=float(row_data.get('base_cost', 0)),
                    cost_unit=row_data.get('cost_unit', COST_UNIT_PER_PERSON),
                    currency=row_data.get('currency', 'USD'),
                    flag_hotel=row_data.get('flag_hotel', False),
                    flag_guide=row_data.get('flag_guide', False),
                    flag_transport=row_data.get('flag_transport', False),
                    flag_meal=row_data.get('flag_meal', False),
                    flag_airport=row_data.get('flag_airport', False),
                    flag_drive=row_data.get('flag_drive', False),
                    hotel_single_rooms=int(row_data.get('hotel_single_rooms', 0)),
                    hotel_double_rooms=int(row_data.get('hotel_double_rooms', 0)),
                    hotel_triple_rooms=int(row_data.get('hotel_triple_rooms', 0)),
                    hotel_other_rooms=int(row_data.get('hotel_other_rooms', 0))
                )
                db.session.add(row)

            db.session.flush()  # Get the ID for new rows

            # Delete existing auto-generated services for this row to avoid duplicates
            InboundHotel.query.filter_by(source_itinerary_id=row.id).delete()
            InboundTransport.query.filter_by(source_itinerary_id=row.id).delete()
            InboundMeal.query.filter_by(source_itinerary_id=row.id).delete()
            InboundGuide.query.filter_by(source_itinerary_id=row.id).delete()

            # Regenerate service records based on current flags
            _auto_generate_services(request_obj, row)

        # Delete orphaned rows (rows whose IDs are no longer in the submitted data)
        orphaned_rows = [row for row in existing_rows_dict.values() if row.id not in submitted_ids]
        for orphaned_row in orphaned_rows:
            # Delete services first to avoid foreign key violations
            InboundHotel.query.filter_by(source_itinerary_id=orphaned_row.id).delete()
            InboundTransport.query.filter_by(source_itinerary_id=orphaned_row.id).delete()
            InboundMeal.query.filter_by(source_itinerary_id=orphaned_row.id).delete()
            InboundGuide.query.filter_by(source_itinerary_id=orphaned_row.id).delete()
            # Now delete the row itself
            db.session.delete(orphaned_row)

        # Recalculate total
        request_obj.calculate_total()

        db.session.commit()

        return jsonify({'success': True, 'total': request_obj.total_amount})

    except Exception as e:
        db.session.rollback()
        print(f"[DEBUG] Error saving itinerary: {str(e)}")
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/generate-days', methods=['POST'])
@csrf.exempt

def api_generate_by_days(request_id):
    """Generate itinerary rows by days"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    # Clear existing rows
    ItineraryRow.query.filter_by(request_id=request_id).delete()

    # Generate one row per day
    current_date = request_obj.from_date
    day_counter = 1

    while current_date <= request_obj.to_date:
        row = ItineraryRow(
            request_id=request_id,
            date=current_date,
            description=f'Day {day_counter} - {current_date.strftime("%A, %B %d")}',
            base_cost=0.0,
            cost_unit=COST_UNIT_PER_PERSON,
            currency=request_obj.total_currency
        )
        db.session.add(row)

        current_date += timedelta(days=1)
        day_counter += 1

    db.session.commit()

    return jsonify({'success': True})

@inbound_bp.route('/api/<int:request_id>/generate-sections', methods=['POST'])
@csrf.exempt

def api_generate_by_sections(request_id):
    """Generate itinerary rows by service sections"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    # Clear existing rows
    ItineraryRow.query.filter_by(request_id=request_id).delete()

    # Generate rows grouped by service type
    services = [
        ('Accommodation', True, False, False, False, False),  # Hotel flag
        ('Transportation', False, False, True, False, False),  # Transport flag
        ('Meals & Dining', False, False, False, True, False),  # Meal flag
        ('Guide Services', False, True, False, False, False),  # Guide flag
        ('Airport Services', False, False, False, False, True)  # Airport flag
    ]

    for service_name, hotel, guide, transport, meal, airport in services:
        row = ItineraryRow(
            request_id=request_id,
            date=request_obj.from_date,
            description=f'{service_name} - {request_obj.from_date.strftime("%B %d")} to {request_obj.to_date.strftime("%B %d")}',
            base_cost=0.0,
            cost_unit=COST_UNIT_PER_PERSON,
            currency=request_obj.total_currency,
            flag_hotel=hotel,
            flag_guide=guide,
            flag_transport=transport,
            flag_meal=meal,
            flag_airport=airport
        )
        db.session.add(row)

    db.session.commit()

    return jsonify({'success': True})

def _auto_generate_services(request_obj, itinerary_row):
    """Auto-generate service records based on itinerary row flags"""

    if itinerary_row.flag_hotel:
        # Generate hotel record - auto-inherit check-in/out from request dates
        check_in = request_obj.from_date
        check_out = request_obj.to_date
        nights = (check_out - check_in).days

        hotel = InboundHotel(
            request_id=request_obj.id,
            source_itinerary_id=itinerary_row.id,
            check_in_date=check_in,
            check_out_date=check_out,
            nights=nights,
            meal_plan='BB',
            total_cost=itinerary_row.calculate_row_cost(request_obj.pax),
            currency=itinerary_row.currency
        )
        db.session.add(hotel)

    if itinerary_row.flag_transport:
        # Generate transport record
        # Default vehicle type based on pax size
        if request_obj.pax <= 4:
            vehicle_type = 'Sedan'
        elif request_obj.pax <= 8:
            vehicle_type = 'Van'
        else:
            vehicle_type = 'Bus'

        transport = InboundTransport(
            request_id=request_obj.id,
            source_itinerary_id=itinerary_row.id,
            date=itinerary_row.date,
            vehicle_type=vehicle_type,
            cost=itinerary_row.calculate_row_cost(request_obj.pax),
            currency=itinerary_row.currency
        )
        db.session.add(transport)

    if itinerary_row.flag_airport:
        # Generate airport transfer (special transport)
        transport = InboundTransport(
            request_id=request_obj.id,
            source_itinerary_id=itinerary_row.id,
            date=itinerary_row.date,
            vehicle_type='Airport Transfer',
            is_airport_transfer=True,
            cost=itinerary_row.calculate_row_cost(request_obj.pax),
            currency=itinerary_row.currency
        )
        db.session.add(transport)

    if itinerary_row.flag_meal:
        # Generate meal record
        meal = InboundMeal(
            request_id=request_obj.id,
            source_itinerary_id=itinerary_row.id,
            date=itinerary_row.date,
            meal_type='Lunch',  # Default
            cost_per_person=itinerary_row.base_cost if itinerary_row.cost_unit == COST_UNIT_PER_PERSON else itinerary_row.base_cost / request_obj.pax,
            total_cost=itinerary_row.calculate_row_cost(request_obj.pax),
            currency=itinerary_row.currency
        )
        db.session.add(meal)

    if itinerary_row.flag_guide:
        # Generate guide record
        # Default language from nationality mapping
        language_map = {
            'German': 'German',
            'French': 'French',
            'Spanish': 'Spanish',
            'Italian': 'Italian',
            'Russian': 'Russian',
            'Chinese': 'Mandarin',
            'Japanese': 'Japanese',
            'Korean': 'Korean',
            'Arabic': 'Arabic'
        }

        language = language_map.get(request_obj.nationality, 'English')

        guide = InboundGuide(
            request_id=request_obj.id,
            source_itinerary_id=itinerary_row.id,
            date=itinerary_row.date,
            language=language,
            service_type='Meet & Greet',
            duration_hours=4.0,  # Default 4 hours
            cost=itinerary_row.calculate_row_cost(request_obj.pax),
            currency=itinerary_row.currency
        )
        db.session.add(guide)

@inbound_bp.route('/api/<int:request_id>/update-master-details', methods=['POST'])
@csrf.exempt

def api_update_master_details(request_id):
    """Update master details"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()

    # Update master details
    request_obj.agent = data.get('agent', request_obj.agent)
    request_obj.contact_name = data.get('contact_name', request_obj.contact_name)
    request_obj.agent_ref = data.get('agent_ref', request_obj.agent_ref)
    request_obj.nationality = data.get('nationality', request_obj.nationality)
    request_obj.pax = data.get('pax', request_obj.pax)
    request_obj.special_note = data.get('special_note', request_obj.special_note)
    # Properly handle customer_id - convert to int if provided
    customer_id_value = data.get('customer_id')
    if customer_id_value and str(customer_id_value).strip():
        try:
            request_obj.customer_id = int(customer_id_value)
        except (ValueError, TypeError):
            pass  # Keep existing value if conversion fails
    elif customer_id_value == '' or customer_id_value is None:
        if 'customer_id' in data:
            request_obj.customer_id = None

    # Update arrival/departure details
    request_obj.arrival_point = data.get('arrival_point', request_obj.arrival_point)
    request_obj.departure_point = data.get('departure_point', request_obj.departure_point)

    # Handle date updates
    if data.get('from_date'):
        request_obj.from_date = datetime.strptime(data.get('from_date'), '%Y-%m-%d').date()
    if data.get('to_date'):
        request_obj.to_date = datetime.strptime(data.get('to_date'), '%Y-%m-%d').date()

    # Handle time updates
    if data.get('arrival_time'):
        try:
            request_obj.arrival_time = datetime.strptime(data.get('arrival_time'), '%H:%M').time()
        except:
            pass  # Invalid time format, skip
    elif data.get('arrival_time') == '':
        request_obj.arrival_time = None

    if data.get('departure_time'):
        try:
            request_obj.departure_time = datetime.strptime(data.get('departure_time'), '%H:%M').time()
        except:
            pass  # Invalid time format, skip
    elif data.get('departure_time') == '':
        request_obj.departure_time = None

    # Recalculate days
    request_obj.calculate_days()

    db.session.commit()

    return jsonify({
        'success': True, 
        'no_of_days': request_obj.no_of_days,
        'message': 'Master details updated successfully'
    })

@inbound_bp.route('/api/<int:request_id>/auto-save-and-regenerate', methods=['POST'])
@csrf.exempt
def api_auto_save_and_regenerate(request_id):
    """Auto-save master details and regenerate itinerary rows"""
    request_obj = InboundRequest.query.options(
        selectinload(InboundRequest.arrival_batches),
        selectinload(InboundRequest.departure_batches)
    ).get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()

    # Track if dates changed
    dates_changed = False
    old_from_date = request_obj.from_date
    old_to_date = request_obj.to_date

    # Update master details
    if data.get('from_date'):
        new_from_date = datetime.strptime(data.get('from_date'), '%Y-%m-%d').date()
        if new_from_date != old_from_date:
            request_obj.from_date = new_from_date
            dates_changed = True
    if data.get('to_date'):
        new_to_date = datetime.strptime(data.get('to_date'), '%Y-%m-%d').date()
        if new_to_date != old_to_date:
            request_obj.to_date = new_to_date
            dates_changed = True

    if data.get('pax'):
        request_obj.pax = int(data.get('pax'))
    if data.get('customer_type'):
        request_obj.customer_type = data.get('customer_type')
    if data.get('contact_name'):
        request_obj.contact_name = data.get('contact_name')
    if data.get('nationality'):
        request_obj.nationality = data.get('nationality')

    # Recalculate days
    request_obj.calculate_days()

    db.session.commit()

    # Regenerate itinerary if dates changed
    if dates_changed and request_obj.from_date and request_obj.to_date:
        # Clear existing rows
        ItineraryRow.query.filter_by(request_id=request_id).delete()

        # Generate one row per day
        current_date = request_obj.from_date
        day_counter = 1

        while current_date <= request_obj.to_date:
            row = ItineraryRow(
                request_id=request_id,
                date=current_date,
                description=f'Day {day_counter} - {current_date.strftime("%A, %B %d")}',
                base_cost=0.0,
                cost_unit=COST_UNIT_PER_PERSON,
                currency=request_obj.total_currency
            )
            db.session.add(row)

            current_date += timedelta(days=1)
            day_counter += 1

        # IMPORTANT: Do NOT auto-update ArrivalBatch and DepartureBatch dates when request dates change
        # Arrival and Departure dates must ONLY be set from their respective tabs (Arrival/Departure)
        # Request dates (from_date/to_date) are used ONLY for:
        # 1. Generating itinerary rows (date range)
        # 2. Calculating trip duration
        # They should NEVER override actual arrival/departure service dates entered by the user

        db.session.commit()

    # Render the itinerary rows HTML using the component template
    sorted_rows = sort_itinerary_rows_with_children(request_obj.itinerary_rows)
    rows_html = render_template('components/itinerary_rows.html', 
                                rows=sorted_rows,
                                request=request_obj,
                                inbound_request=request_obj,  # Explicit variable to avoid Flask request confusion
                                guide_suppliers=_guide_suppliers_for_itinerary_ui(),
                                itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(request_obj),
                                view_only=False)

    return jsonify({
        'success': True,
        'dates_changed': dates_changed,
        'no_of_days': request_obj.no_of_days,
        'itinerary_html': rows_html
    })

@inbound_bp.route('/api/<int:request_id>/save-request', methods=['POST'])
@csrf.exempt
def api_save_request(request_id):
    """Save request and assign final sequence number based on from_date month"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    
    # Track original dates to detect changes
    original_from = request_obj.from_date
    original_to = request_obj.to_date

    # Update from_date if provided
    if data.get('from_date'):
        request_obj.from_date = datetime.strptime(data.get('from_date'), '%Y-%m-%d').date()
    if data.get('to_date'):
        request_obj.to_date = datetime.strptime(data.get('to_date'), '%Y-%m-%d').date()
    if data.get('pax'):
        request_obj.pax = int(data.get('pax'))
    if data.get('customer_type'):
        request_obj.customer_type = data.get('customer_type')
    if data.get('contact_name'):
        request_obj.contact_name = data.get('contact_name')
    if data.get('nationality'):
        request_obj.nationality = data.get('nationality')
    if data.get('customer_id'):
        request_obj.customer_id = int(data.get('customer_id'))
    if data.get('agent_ref'):
        request_obj.agent_ref = data.get('agent_ref')

    # Check if dates changed
    dates_changed = (original_from != request_obj.from_date) or (original_to != request_obj.to_date)

    # Recalculate days
    request_obj.calculate_days()

    # Only generate sequence number if not already saved AND has placeholder number
    # Linked attachments keep their {parent}-N number and must not get a new YYMM### id.
    if request_obj.parent_request_id:
        if not request_obj.is_saved:
            request_obj.is_saved = True
    elif not request_obj.is_saved and request_obj.request_number.startswith('IN-NEW-'):
        # Generate sequence number based on from_date month
        request_obj.request_number = InboundRequest.generate_request_number(request_obj.from_date)
        request_obj.is_saved = True
    elif not request_obj.is_saved:
        # Legacy request without is_saved flag but with valid number - just mark as saved
        request_obj.is_saved = True

    # If dates changed, regenerate itinerary
    itinerary_html = None
    if dates_changed and request_obj.from_date and request_obj.to_date:
        # Delete existing itinerary rows
        ItineraryRow.query.filter_by(request_id=request_id).delete()
        
        # Generate new itinerary rows (day_number is calculated property, only set date and description)
        current_date = request_obj.from_date
        while current_date <= request_obj.to_date:
            new_row = ItineraryRow(
                request_id=request_id,
                date=current_date,
                description=''
            )
            db.session.add(new_row)
            current_date += timedelta(days=1)
        
        db.session.flush()
        
        # Render updated itinerary HTML
        sorted_rows = sort_itinerary_rows_with_children(request_obj.itinerary_rows)
        itinerary_html = render_template('components/itinerary_rows.html', 
                                        rows=sorted_rows,
                                        request=request_obj,
                                        inbound_request=request_obj,  # Explicit variable to avoid Flask request confusion
                                        guide_suppliers=_guide_suppliers_for_itinerary_ui(),
                                        itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(request_obj),
                                        view_only=False)

    db.session.commit()

    response = {
        'success': True,
        'request_number': request_obj.request_number,
        'message': 'Request saved successfully',
        'dates_changed': dates_changed,
        'no_of_days': request_obj.no_of_days
    }
    
    if itinerary_html:
        response['itinerary_html'] = itinerary_html
    
    return jsonify(response)


def _check_guide_assigned_on_date(supplier_id, check_date, exclude_guide_id=None):
    """Check if a guide (supplier) is already assigned to another trip on the given date.
    Returns (conflict_found, conflicting_request_number) or (False, None).
    Each guide can have only one trip per day. Excludes cancelled assignments.
    When editing, exclude_guide_id is the row being updated so we don't count it.
    A guide covers check_date if: date <= check_date <= (end_date or date).
    """
    if not supplier_id or not check_date:
        return False, None
    from sqlalchemy import or_
    # Guide covers check_date when: date <= check_date AND (end_date >= check_date OR (end_date IS NULL AND date = check_date))
    q = InboundGuide.query.filter(
        InboundGuide.supplier_id == int(supplier_id),
        InboundGuide.date <= check_date,
        InboundGuide.is_cancelled == False,
        or_(
            (InboundGuide.end_date != None) & (InboundGuide.end_date >= check_date),
            (InboundGuide.end_date == None) & (InboundGuide.date == check_date)
        )
    )
    if exclude_guide_id is not None:
        q = q.filter(InboundGuide.id != exclude_guide_id)
    existing = q.first()
    if existing:
        req = InboundRequest.query.get(existing.request_id)
        req_num = req.request_number if req else f"Request #{existing.request_id}"
        return True, req_num
    return False, None


def _itinerary_guide_slot_saved_map(request_obj):
    """Map 'itinerary_row_id:supplier_id' -> InboundGuide.id for itinerary-integrated guides."""
    m = {}
    guides = list(getattr(request_obj, 'inbound_guides', None) or [])
    for g in guides:
        if g.source_itinerary_id and g.supplier_id:
            m[f'{g.source_itinerary_id}:{g.supplier_id}'] = g.id
    # Backfill: supplier is on the itinerary row for that day but guide.source_itinerary_id was never set
    for row in getattr(request_obj, 'itinerary_rows', None) or []:
        row_date = getattr(row, 'date', None)
        if not row_date:
            continue
        for sid in row.get_itinerary_guide_supplier_id_list():
            try:
                sid_i = int(sid)
            except (TypeError, ValueError):
                continue
            k = f'{row.id}:{sid_i}'
            if k in m:
                continue
            for g in guides:
                if not g.supplier_id or not g.date:
                    continue
                try:
                    if int(g.supplier_id) != sid_i:
                        continue
                except (TypeError, ValueError):
                    continue
                if g.date != row_date:
                    continue
                m[k] = g.id
                break
    return m


def _ensure_guide_linked_to_itinerary_row(request_id, guide):
    """If guide has no source_itinerary_id but matches an itinerary day that lists this supplier, link it."""
    if not guide or guide.source_itinerary_id or not guide.supplier_id or not guide.date:
        return
    try:
        sid = int(guide.supplier_id)
    except (TypeError, ValueError):
        return
    for row in ItineraryRow.query.filter_by(request_id=request_id, date=guide.date).order_by(ItineraryRow.id):
        rids = row.get_itinerary_guide_supplier_id_list()
        rids_int = []
        for x in rids:
            try:
                rids_int.append(int(x))
            except (TypeError, ValueError):
                continue
        if sid in rids_int:
            guide.source_itinerary_id = row.id
            return


def _relocate_itinerary_linked_guide_for_new_start_date(request_id, guide):
    """
    If a guide is tied to an itinerary row but its start date (guide.date) no longer matches
    that row's day, move the supplier chip to the itinerary row for the new date and repoint
    guide.source_itinerary_id. Returns (ok, error_message).
    """
    if not guide or not guide.source_itinerary_id or not guide.supplier_id or not guide.date:
        return True, None
    try:
        sid = int(guide.supplier_id)
    except (TypeError, ValueError):
        return True, None
    old_row = ItineraryRow.query.filter_by(
        id=guide.source_itinerary_id, request_id=request_id
    ).first()
    if not old_row:
        return True, None
    if old_row.date == guide.date:
        return True, None
    target = ItineraryRow.query.filter_by(
        request_id=request_id, date=guide.date
    ).order_by(ItineraryRow.id).first()
    if not target:
        return False, (
            'No itinerary day exists for the guide start date you entered. '
            'Pick a date that matches a day on the itinerary, or add that day first.'
        )
    old_ids = old_row.get_itinerary_guide_supplier_id_list()
    old_ids = [int(x) for x in old_ids if x is not None]
    old_ids = [x for x in old_ids if x != sid]
    old_row.set_itinerary_guide_supplier_id_list(old_ids)
    new_ids = target.get_itinerary_guide_supplier_id_list()
    new_ids = [int(x) for x in new_ids if x is not None]
    if sid not in new_ids:
        new_ids.append(sid)
    target.set_itinerary_guide_supplier_id_list(new_ids)
    guide.source_itinerary_id = target.id
    return True, None


def _guide_suppliers_for_itinerary_ui():
    from app.models.supplier import Supplier
    return Supplier.query.filter(
        Supplier.supplier_type == 'GUIDE',
        Supplier.is_active == True
    ).order_by(Supplier.name).limit(500).all()


@inbound_bp.route('/api/<int:request_id>/check-guide-availability', methods=['GET'])
def api_check_guide_availability(request_id):
    """Check if a guide can be assigned for given dates (one trip per day rule)."""
    guide_supplier_id = request.args.get('guide_supplier_id', '').strip()
    from_date_str = request.args.get('from_date', '').strip()
    to_date_str = request.args.get('to_date', '').strip()
    exclude_row_id = request.args.get('exclude_row_id')  # When editing, exclude current guide row
    if not guide_supplier_id or not from_date_str:
        return jsonify({'available': True})  # No guide/date selected - skip check
    try:
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else from_date
        if to_date < from_date:
            to_date = from_date
    except (ValueError, TypeError):
        return jsonify({'available': True})
    exclude_guide_id = int(exclude_row_id) if exclude_row_id else None
    current = from_date
    while current <= to_date:
        conflict, req_num = _check_guide_assigned_on_date(
            int(guide_supplier_id), current,
            exclude_guide_id=exclude_guide_id
        )
        if conflict:
            return jsonify({
                'available': False,
                'error': f'This guide cannot have more than one trip on the same day. They are already assigned to {req_num} on {current.strftime("%d %b %Y")}.'
            }), 200
        current += timedelta(days=1)
    return jsonify({'available': True})


@inbound_bp.route('/api/representatives', methods=['GET'])
def api_list_representatives():
    """List all representatives for dropdown population (merged with Meet & Assist suppliers)."""
    merged = _get_merged_representatives_for_dropdown()
    return jsonify({'representatives': merged})


@inbound_bp.route('/api/representatives', methods=['POST'])
@csrf.exempt
def api_add_representative():
    """Add a new representative; also ensures a Meet & Assist (GROUND_HANDLER) supplier exists."""
    data = request.get_json() or {}
    name = _normalize_meet_assist_name(data.get('name'))
    if not name:
        return jsonify({'success': False, 'error': 'Name is required'}), 400
    try:
        rep, _supplier = _sync_meet_assist_representative_pair(name)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()
        return jsonify({
            'success': True,
            'representative': {'id': rep.id, 'name': rep.name},
        }), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<int:request_id>/save-service-data', methods=['POST'])
@csrf.exempt
def api_save_service_data(request_id):
    """Save service data (hotel, transport, guide, meal) for itinerary"""
    from datetime import date as date_type

    # Initialize variables at function level to avoid UnboundLocalError
    from_date = None
    to_date = None

    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    service_type = data.get('service_type')
    form_data = data.get('data', {})
    row_id = data.get('row_id')
    is_global = data.get('is_global', False)

    # Debug logging - use print for immediate visibility
    print(f"[SAVE SERVICE] type={service_type}, is_global={is_global}, row_id={row_id}")
    print(f"[SAVE SERVICE] form_data keys: {list(form_data.keys())}")
    print(f"[SAVE SERVICE] hotel_name value: '{form_data.get('hotel_name', 'NOT_FOUND')}'")
    print(f"[SAVE SERVICE] full form_data: {form_data}")

    # Validate context
    if not service_type:
        return jsonify({'success': False, 'error': 'Missing service_type'}), 400

    if not is_global and not row_id:
        return jsonify({'success': False, 'error': 'Missing row_id for day-specific save'}), 400

    # Validate row exists for day-specific saves
    # Skip validation for service types that use their own table IDs (not ItineraryRow IDs)
    service_uses_own_table = ['arrival', 'departure', 'hotel', 'transport', 'guide', 'meal']
    if not is_global and service_type not in service_uses_own_table:
        row = ItineraryRow.query.get(row_id)
        if not row or row.request_id != request_id:
            return jsonify({'success': False, 'error': 'Invalid row_id'}), 400

    # Validate required fields based on service type
    validation_errors = []
    if service_type == 'hotel':
        if not form_data.get('hotel_name', '').strip():
            validation_errors.append('Hotel Name is required')
    elif service_type == 'transport':
        if not form_data.get('transport_vehicle', '').strip():
            validation_errors.append('Vehicle Type is required')
    elif service_type == 'meal':
        if not form_data.get('meal_restaurant', '').strip():
            validation_errors.append('Restaurant is required')

    if validation_errors:
        return jsonify({
            'success': False, 
            'error': ', '.join(validation_errors),
            'validation_errors': validation_errors
        }), 400

    try:
        guide_record_id_for_json = None
        if service_type == 'hotel':
            # Remember the Hotel form's City/Category filter selections on the request
            # so they are restored on reopen. UI-only filters; committed with the hotel save below.
            if 'hotel_city_filter' in form_data:
                request_obj.hotel_filter_city = (form_data.get('hotel_city_filter') or '').strip() or None
            if 'hotel_category_filter' in form_data:
                request_obj.hotel_filter_category = (form_data.get('hotel_category_filter') or '').strip() or None
            # Check if editing an existing hotel (row_id or hotel_id provided) or adding new
            # row_id is passed when editing from Trip Summary
            hotel_id = row_id or data.get('hotel_id')
            if hotel_id:
                # Update existing hotel
                hotel = InboundHotel.query.filter_by(id=hotel_id, request_id=request_id).first()
                if not hotel:
                    return jsonify({'success': False, 'error': 'Hotel not found'}), 404
                print(f"[SAVE SERVICE] Updating existing hotel id={hotel_id}")
            else:
                # Create new hotel entry - use only service-specific dates
                check_in_str = form_data.get('hotel_check_in', '').strip()
                check_out_str = form_data.get('hotel_check_out', '').strip()
                hotel_name_value = form_data.get('hotel_name', '').strip()
                
                if not check_in_str or not check_out_str:
                    field_errors = {}
                    if not check_in_str:
                        field_errors['hotel_check_in'] = 'Check-in date is required'
                    if not check_out_str:
                        field_errors['hotel_check_out'] = 'Check-out date is required'
                    return jsonify({
                        'success': False,
                        'error': 'Hotel check-in and check-out dates are required',
                        'field_errors': field_errors
                    }), 400
                
                try:
                    check_in_date = datetime.strptime(check_in_str, '%Y-%m-%d').date()
                    check_out_date = datetime.strptime(check_out_str, '%Y-%m-%d').date()
                except ValueError as e:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid date format: {str(e)}',
                        'field_errors': {'hotel_check_in': 'Invalid date format', 'hotel_check_out': 'Invalid date format'}
                    }), 400
                
                # Check for duplicate hotel entry before creating
                # Prevent creating duplicate hotels with same name, check-in, and check-out dates
                if hotel_name_value:
                    existing_hotel = InboundHotel.query.filter_by(
                        request_id=request_id,
                        hotel_name=hotel_name_value,
                        check_in_date=check_in_date,
                        check_out_date=check_out_date
                    ).first()
                    
                    if existing_hotel:
                        print(f"[SAVE SERVICE] Duplicate hotel detected - using existing hotel id={existing_hotel.id}")
                        hotel = existing_hotel
                    else:
                        hotel = InboundHotel(
                            request_id=request_id,
                            check_in_date=check_in_date,
                            check_out_date=check_out_date
                        )
                        db.session.add(hotel)
                        print(f"[SAVE SERVICE] Creating new hotel entry")
                else:
                    hotel = InboundHotel(
                        request_id=request_id,
                        check_in_date=check_in_date,
                        check_out_date=check_out_date
                    )
                    db.session.add(hotel)
                    print(f"[SAVE SERVICE] Creating new hotel entry (no hotel name yet)")

            # Process hotel data (both new and existing)
            hotel_name_value = form_data.get('hotel_name', '').strip()
            print(f"[SAVE SERVICE] Assigning hotel_name: '{hotel_name_value}' to hotel id: {hotel.id if hotel.id else 'NEW'}")
            hotel.hotel_name = hotel_name_value
            hotel.hotel_category = form_data.get('hotel_category', '')
            hotel.meal_plan = form_data.get('hotel_board', 'BB')
            hotel.status = form_data.get('hotel_status', 'REQUEST')

            # Dates are already set when creating new hotel, but update them if provided in form_data
            # (This handles the case when editing an existing hotel)
            check_in_str = form_data.get('hotel_check_in', '')
            check_out_str = form_data.get('hotel_check_out', '')
            if check_in_str:
                hotel.check_in_date = datetime.strptime(check_in_str, '%Y-%m-%d').date()
            if check_out_str:
                hotel.check_out_date = datetime.strptime(check_out_str, '%Y-%m-%d').date()
            cut_off_str = form_data.get('hotel_cut_off_date', '')
            hotel.cut_off_date = datetime.strptime(cut_off_str, '%Y-%m-%d').date() if cut_off_str else None

            # Calculate nights
            if hotel.check_in_date and hotel.check_out_date:
                hotel.nights = (hotel.check_out_date - hotel.check_in_date).days

            hotel.total_cost = float(form_data.get('hotel_cost', 0) or 0)

            # Room distribution
            hotel.single_rooms = int(form_data.get('hotel_single_rooms', 0) or 0)
            hotel.double_rooms = int(form_data.get('hotel_double_rooms', 0) or 0)
            hotel.triple_rooms = int(form_data.get('hotel_triple_rooms', 0) or 0)
            hotel.notes = form_data.get('hotel_notes', '')

            # Flush to get hotel ID before creating rooms
            db.session.flush()

            # Check if room_list data was provided from the Room List tab
            room_list_data = data.get('room_list', [])

            if room_list_data:
                # Use room list data provided by user (with guest names and new fields)
                HotelRoom.query.filter_by(hotel_id=hotel.id).delete()
                default_board_basis = form_data.get('hotel_board', 'BB')

                for room_data in room_list_data:
                    room = HotelRoom(
                        hotel_id=hotel.id,
                        room_type=room_data.get('room_type', 'Double'),
                        room_count=1,
                        room_category=room_data.get('room_category', ''),
                        room_option=room_data.get('room_option', ''),
                        board_basis=room_data.get('board_basis', default_board_basis),
                        dietary_requirements=room_data.get('dietary_requirements', ''),
                        adults=room_data.get('adults', 2),
                        children=room_data.get('children', 0),
                        guest_names=room_data.get('guest_names', '')
                    )
                    db.session.add(room)

                # Update room distribution counts based on room list
                hotel.single_rooms = sum(1 for r in room_list_data if r.get('room_type') == 'Single')
                hotel.double_rooms = sum(1 for r in room_list_data if r.get('room_type') in ['Double', 'Twin'])
                hotel.triple_rooms = sum(1 for r in room_list_data if r.get('room_type') in ['Triple', 'Suite'])
            else:
                # Use distribution counts to create rooms (without guest names)
                # Get room categories from distribution if provided
                room_categories = data.get('room_categories', {})
                sgl_category = room_categories.get('single', '')
                dbl_category = room_categories.get('double', '')
                trp_category = room_categories.get('triple', '')

                total_rooms = hotel.single_rooms + hotel.double_rooms + hotel.triple_rooms
                if total_rooms > 0:
                    # Check if existing rooms match new distribution or categories changed
                    existing_rooms = HotelRoom.query.filter_by(hotel_id=hotel.id).all()
                    existing_count = len(existing_rooms)

                    # Recreate if distribution changed OR if room categories were provided
                    needs_recreate = existing_count != total_rooms or (sgl_category or dbl_category or trp_category)
                    if needs_recreate:
                        HotelRoom.query.filter_by(hotel_id=hotel.id).delete()

                        # Create individual room records based on distribution
                        board_basis = form_data.get('hotel_board', 'BB')
                        for i in range(hotel.single_rooms):
                            room = HotelRoom(hotel_id=hotel.id, room_type='Single', room_count=1, board_basis=board_basis, adults=1, room_category=sgl_category)
                            db.session.add(room)
                        for i in range(hotel.double_rooms):
                            room = HotelRoom(hotel_id=hotel.id, room_type='Double', room_count=1, board_basis=board_basis, adults=2, room_category=dbl_category)
                            db.session.add(room)
                        for i in range(hotel.triple_rooms):
                            room = HotelRoom(hotel_id=hotel.id, room_type='Triple', room_count=1, board_basis=board_basis, adults=3, room_category=trp_category)
                            db.session.add(room)

            # Save confirmation number to first room (hotel-level field in UI)
            confirmation = form_data.get('hotel_confirmation_number', '').strip()
            rooms = HotelRoom.query.filter_by(hotel_id=hotel.id).all()
            if rooms:
                rooms[0].confirmation = confirmation
            elif confirmation:
                room = HotelRoom(hotel_id=hotel.id, room_type='Double', room_count=1, confirmation=confirmation)
                db.session.add(room)

        elif service_type == 'transport':
            drv_phone = form_data.get('transport_driver_phone', '').strip()
            if drv_phone and not is_valid_phone(drv_phone):
                return jsonify({'success': False, 'error': f'Driver phone: {PHONE_ERROR}'}), 400
            # Check if editing an existing transport record via row_id
            if row_id:
                transport = InboundTransport.query.filter_by(id=row_id, request_id=request_id).first()
                if transport:
                    # Update existing transport
                    print(f"[SAVE SERVICE] Updating existing transport id={row_id}")
                    transport.vehicle_type = form_data.get('transport_vehicle', '')
                    transport.pickup_location = form_data.get('transport_pickup', '')
                    transport.dropoff_location = form_data.get('transport_dropoff', '')
                    transport.driver_name = form_data.get('transport_driver_name', '')
                    transport.driver_phone = form_data.get('transport_driver_phone', '')
                    transport.license_number = form_data.get('transport_license_number', '')
                    transport.status = form_data.get('transport_status', STATUS_REQUEST)
                    transport.cost = float(form_data.get('transport_cost', 0) or 0)
                    transport.currency = form_data.get('transport_cost_currency', 'USD')
                    transport.note = form_data.get('transport_notes', '')
                    supplier_id = form_data.get('transport_supplier')
                    transport.supplier_id = int(supplier_id) if supplier_id else None
                    if form_data.get('transport_from_date'):
                        try:
                            transport.date = datetime.strptime(form_data['transport_from_date'], '%Y-%m-%d').date()
                        except (ValueError, TypeError, OSError):
                            pass  # Keep existing date on parse error
                    if form_data.get('transport_to_date'):
                        try:
                            transport.end_date = datetime.strptime(form_data['transport_to_date'], '%Y-%m-%d').date()
                        except (ValueError, TypeError, OSError):
                            pass
                    _cut_off = form_data.get('transport_cut_off_date', '')
                    transport.cut_off_date = datetime.strptime(_cut_off, '%Y-%m-%d').date() if _cut_off else None
                else:
                    return jsonify({'success': False, 'error': 'Transport record not found'}), 404
            else:
                # Create a single transport entry with date range (from_date → to_date)
                from_date_str = form_data.get('transport_from_date', '')
                to_date_str = form_data.get('transport_to_date', '')

                # Use only service-specific dates, no fallback to request dates
                if not from_date_str or not to_date_str:
                    return jsonify({
                        'success': False,
                        'error': 'Transport dates are required',
                        'field_errors': {'transport_from_date': 'From date is required', 'transport_to_date': 'To date is required'}
                    }), 400
                
                try:
                    from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError, OSError) as e:
                    return jsonify({
                        'success': False,
                        'error': 'Invalid date format. Please use YYYY-MM-DD.',
                        'field_errors': {'transport_from_date': 'Invalid date format', 'transport_to_date': 'Invalid date format'}
                    }), 400

                # Create a single transport entry with date range
                supplier_id = form_data.get('transport_supplier')
                supplier_id = int(supplier_id) if supplier_id else None

                transport = InboundTransport(
                    request_id=request_id,
                    date=from_date,  # Start date
                    end_date=to_date  # End date for multi-day service
                )
                _cut_off = form_data.get('transport_cut_off_date', '')
                transport.cut_off_date = datetime.strptime(_cut_off, '%Y-%m-%d').date() if _cut_off else None
                transport.vehicle_type = form_data.get('transport_vehicle', '')
                transport.pickup_location = form_data.get('transport_pickup', '')
                transport.dropoff_location = form_data.get('transport_dropoff', '')
                transport.driver_name = form_data.get('transport_driver_name', '')
                transport.driver_phone = form_data.get('transport_driver_phone', '')
                transport.license_number = form_data.get('transport_license_number', '')
                transport.status = form_data.get('transport_status', 'REQUEST')
                transport.cost = float(form_data.get('transport_cost', 0) or 0)
                transport.currency = form_data.get('transport_cost_currency', 'USD')
                transport.note = form_data.get('transport_notes', '')
                transport.supplier_id = supplier_id
                db.session.add(transport)

                print(f"[SAVE SERVICE] Created single transport entry from {from_date} to {to_date}")

        elif service_type == 'guide':
            from app.models.supplier import Supplier

            source_itinerary_id = data.get('source_itinerary_id')
            try:
                source_itinerary_id = int(source_itinerary_id) if source_itinerary_id not in (None, '') else None
            except (TypeError, ValueError):
                source_itinerary_id = None

            itinerary_slot_supplier_id = data.get('itinerary_slot_supplier_id')
            try:
                itinerary_slot_supplier_id = int(itinerary_slot_supplier_id) if itinerary_slot_supplier_id not in (None, '') else None
            except (TypeError, ValueError):
                itinerary_slot_supplier_id = None

            guide_supplier_id_val = form_data.get('guide_supplier_id')
            try:
                guide_supplier_id_val = int(guide_supplier_id_val) if guide_supplier_id_val and str(guide_supplier_id_val) != '__ADD_NEW__' else None
            except (TypeError, ValueError):
                guide_supplier_id_val = None

            from_date_str = form_data.get('guide_from_date', '').strip()
            to_date_str = (form_data.get('guide_to_date', '').strip() or from_date_str)

            exclude_gid = None
            if row_id:
                try:
                    exclude_gid = int(row_id)
                except (TypeError, ValueError):
                    exclude_gid = None
            elif source_itinerary_id and guide_supplier_id_val:
                existing_for_slot = InboundGuide.query.filter_by(
                    request_id=request_id,
                    source_itinerary_id=source_itinerary_id,
                    supplier_id=guide_supplier_id_val
                ).first()
                if existing_for_slot:
                    exclude_gid = existing_for_slot.id

            if guide_supplier_id_val and from_date_str:
                try:
                    fd = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    td = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else fd
                    if td < fd:
                        td = fd
                    d = fd
                    while d <= td:
                        conflict, req_num = _check_guide_assigned_on_date(
                            guide_supplier_id_val, d,
                            exclude_guide_id=exclude_gid
                        )
                        if conflict:
                            return jsonify({
                                'success': False,
                                'error': f'This guide cannot have more than one trip on the same day. They are already assigned to {req_num} on {d.strftime("%d %b %Y")}.'
                            }), 400
                        d += timedelta(days=1)
                except (ValueError, TypeError):
                    pass

            def _apply_guide_form_fields(g):
                gsup = form_data.get('guide_supplier_id')
                gsup = int(gsup) if gsup and str(gsup) != '__ADD_NEW__' else None
                if gsup:
                    supplier = Supplier.query.get(gsup)
                    g.guide_name = supplier.name if supplier else ''
                else:
                    g.guide_name = form_data.get('guide_name', '')
                g.supplier_id = gsup
                g.language = form_data.get('guide_language', '')
                g.telephone_number = form_data.get('guide_phone', '')
                g.cost = float(form_data.get('guide_cost', 0) or 0)
                g.currency = form_data.get('guide_cost_currency', 'USD')
                g.is_cancelled = form_data.get('guide_cancelled') in ['true', 'True', True, 'on', '1']
                g.additional_comments = form_data.get('guide_notes', '')
                g.status = form_data.get('guide_status', 'REQUESTED')
                if form_data.get('guide_from_date'):
                    g.date = datetime.strptime(form_data['guide_from_date'], '%Y-%m-%d').date()
                if form_data.get('guide_to_date'):
                    g.end_date = datetime.strptime(form_data['guide_to_date'], '%Y-%m-%d').date()
                elif g.date:
                    g.end_date = g.date

            guide = None
            if row_id:
                guide = InboundGuide.query.filter_by(id=row_id, request_id=request_id).first()
                if not guide:
                    return jsonify({'success': False, 'error': 'Guide record not found'}), 404
                print(f"[SAVE SERVICE] Updating existing guide id={row_id}")
                _apply_guide_form_fields(guide)
            elif source_itinerary_id:
                it_row = ItineraryRow.query.filter_by(id=source_itinerary_id, request_id=request_id).first()
                if not it_row:
                    return jsonify({'success': False, 'error': 'Invalid itinerary row'}), 400
                if guide_supplier_id_val and itinerary_slot_supplier_id and itinerary_slot_supplier_id != guide_supplier_id_val:
                    return jsonify({'success': False, 'error': 'Guide supplier mismatch'}), 400
                if not from_date_str or not to_date_str:
                    return jsonify({
                        'success': False,
                        'error': 'Guide dates are required',
                        'field_errors': {'guide_from_date': 'From date is required', 'guide_to_date': 'To date is required'}
                    }), 400
                try:
                    from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError) as e:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid date format: {str(e)}',
                        'field_errors': {'guide_from_date': 'Invalid date format', 'guide_to_date': 'Invalid date format'}
                    }), 400
                guide = InboundGuide.query.filter_by(
                    request_id=request_id,
                    source_itinerary_id=source_itinerary_id,
                    supplier_id=guide_supplier_id_val
                ).first()
                if guide:
                    print(f"[SAVE SERVICE] Updating itinerary-linked guide id={guide.id}")
                    _apply_guide_form_fields(guide)
                else:
                    supplier = Supplier.query.get(guide_supplier_id_val)
                    guide_name_text = supplier.name if supplier else (form_data.get('guide_name', '').strip() or '')
                    guide = InboundGuide(
                        request_id=request_id,
                        date=from_date,
                        end_date=to_date,
                        source_itinerary_id=source_itinerary_id
                    )
                    guide.guide_name = guide_name_text
                    guide.language = form_data.get('guide_language', '')
                    guide.telephone_number = form_data.get('guide_phone', '')
                    guide.cost = float(form_data.get('guide_cost', 0) or 0)
                    guide.currency = form_data.get('guide_cost_currency', 'USD')
                    guide.is_cancelled = form_data.get('guide_cancelled') in ['true', 'True', True, 'on', '1']
                    guide.additional_comments = form_data.get('guide_notes', '')
                    guide.status = form_data.get('guide_status', 'REQUEST')
                    guide.supplier_id = guide_supplier_id_val
                    db.session.add(guide)
                    print(f"[SAVE SERVICE] Created itinerary-linked guide for row {source_itinerary_id}")
            else:
                from_date_str = form_data.get('guide_from_date', '')
                to_date_str = form_data.get('guide_to_date', '')
                if not from_date_str or not to_date_str:
                    return jsonify({
                        'success': False,
                        'error': 'Guide dates are required',
                        'field_errors': {'guide_from_date': 'From date is required', 'guide_to_date': 'To date is required'}
                    }), 400
                try:
                    from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()
                except (ValueError, TypeError) as e:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid date format: {str(e)}',
                        'field_errors': {'guide_from_date': 'Invalid date format', 'guide_to_date': 'Invalid date format'}
                    }), 400

                guide_supplier_id = form_data.get('guide_supplier_id')
                guide_supplier_id = int(guide_supplier_id) if guide_supplier_id and str(guide_supplier_id) != '__ADD_NEW__' else None
                guide_name_text = form_data.get('guide_name', '').strip()
                if guide_supplier_id:
                    supplier = Supplier.query.get(guide_supplier_id)
                    if supplier:
                        guide_name_text = supplier.name

                guide = InboundGuide(
                    request_id=request_id,
                    date=from_date,
                    end_date=to_date
                )
                guide.guide_name = guide_name_text
                guide.language = form_data.get('guide_language', '')
                guide.telephone_number = form_data.get('guide_phone', '')
                guide.cost = float(form_data.get('guide_cost', 0) or 0)
                guide.currency = form_data.get('guide_cost_currency', 'USD')
                guide.is_cancelled = form_data.get('guide_cancelled') in ['true', 'True', True, 'on', '1']
                guide.additional_comments = form_data.get('guide_notes', '')
                guide.status = form_data.get('guide_status', 'REQUEST')
                guide.supplier_id = guide_supplier_id
                db.session.add(guide)
                print(f"[SAVE SERVICE] Created single guide entry from {from_date} to {to_date}")

            if guide:
                _ensure_guide_linked_to_itinerary_row(request_id, guide)
            if guide and guide.source_itinerary_id:
                ok_reloc, reloc_err = _relocate_itinerary_linked_guide_for_new_start_date(request_id, guide)
                if not ok_reloc:
                    db.session.rollback()
                    return jsonify({'success': False, 'error': reloc_err}), 400

        elif service_type == 'meal':
            # Check if editing an existing meal record via row_id
            if row_id:
                meal = InboundMeal.query.filter_by(id=row_id, request_id=request_id).first()
                if meal:
                    # Update existing meal
                    print(f"[SAVE SERVICE] Updating existing meal id={row_id}")
                    from app.models.supplier import Supplier
                    # meal_restaurant select value is the supplier_id
                    restaurant_val = form_data.get('meal_restaurant', '')
                    supplier_id = form_data.get('meal_supplier') or restaurant_val
                    if supplier_id:
                        try:
                            meal.supplier_id = int(supplier_id)
                            supplier = Supplier.query.get(int(supplier_id))
                            meal.restaurant = supplier.name if supplier else restaurant_val
                        except (ValueError, TypeError):
                            meal.supplier_id = None
                            meal.restaurant = restaurant_val
                    else:
                        meal.supplier_id = None
                        meal.restaurant = restaurant_val
                    meal.meal_type = form_data.get('meal_type', '')
                    meal.meal_note = form_data.get('meal_notes', '')
                    meal.total_cost = float(form_data.get('meal_cost', 0) or 0)
                    meal.currency = form_data.get('meal_cost_currency', 'JOD')
                    meal.location = form_data.get('meal_location', '')
                    meal.status = form_data.get('meal_status', STATUS_REQUEST)
                    # Save restaurant PAX count independently from request PAX
                    meal_pax = form_data.get('meal_pax', '')
                    meal.pax_count = int(meal_pax) if meal_pax and str(meal_pax).strip() else None
                    print(f"[SAVE MEAL UPDATE] meal_pax from form: {repr(meal_pax)}, meal.pax_count set to: {meal.pax_count}")
                    if form_data.get('meal_from_date'):
                        meal.date = datetime.strptime(form_data['meal_from_date'], '%Y-%m-%d').date()
                    meal.end_date = None  # Restaurant uses single date only
                    _cut_off = form_data.get('meal_cut_off_date', '')
                    meal.cut_off_date = datetime.strptime(_cut_off, '%Y-%m-%d').date() if _cut_off else None
                else:
                    return jsonify({'success': False, 'error': 'Meal record not found'}), 404
            else:
                # Create a single meal entry with single date
                from_date_str = form_data.get('meal_from_date', '')

                if not from_date_str:
                    return jsonify({
                        'success': False,
                        'error': 'Meal date is required',
                        'field_errors': {'meal_from_date': 'Date is required'}
                    }), 400

                try:
                    from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
                    if not isinstance(from_date, date_type):
                        raise ValueError('Invalid date object created')
                except (ValueError, TypeError) as e:
                    return jsonify({
                        'success': False,
                        'error': f'Invalid date format: {str(e)}',
                        'field_errors': {'meal_from_date': 'Invalid date format'}
                    }), 400

                # Create a single meal entry with single date
                from app.models.supplier import Supplier
                # meal_restaurant select value is the supplier_id
                restaurant_val = form_data.get('meal_restaurant', '')
                supplier_id_val = form_data.get('meal_supplier') or restaurant_val
                resolved_supplier_id = None
                resolved_restaurant = restaurant_val
                if supplier_id_val:
                    try:
                        resolved_supplier_id = int(supplier_id_val)
                        supplier = Supplier.query.get(int(supplier_id_val))
                        resolved_restaurant = supplier.name if supplier else restaurant_val
                    except (ValueError, TypeError):
                        resolved_supplier_id = None
                        resolved_restaurant = restaurant_val

                try:
                    meal = InboundMeal(
                        request_id=request_id,
                        date=from_date,
                        end_date=None  # Restaurant uses single date only
                    )
                    _cut_off = form_data.get('meal_cut_off_date', '')
                    meal.cut_off_date = datetime.strptime(_cut_off, '%Y-%m-%d').date() if _cut_off else None
                    meal.restaurant = resolved_restaurant
                    meal.meal_type = form_data.get('meal_type', '')
                    meal.meal_note = form_data.get('meal_notes', '')
                    meal.total_cost = float(form_data.get('meal_cost', 0) or 0)
                    meal.currency = form_data.get('meal_cost_currency', 'JOD')
                    meal.location = form_data.get('meal_location', '')
                    meal.status = form_data.get('meal_status', STATUS_REQUEST)
                    meal.supplier_id = resolved_supplier_id
                    # Save restaurant PAX count independently from request PAX
                    meal_pax = form_data.get('meal_pax', '')
                    meal.pax_count = int(meal_pax) if meal_pax and str(meal_pax).strip() else None
                    print(f"[SAVE MEAL CREATE] meal_pax from form: {repr(meal_pax)}, meal.pax_count set to: {meal.pax_count}")

                    db.session.add(meal)
                    db.session.flush()  # Flush to get the ID and catch any errors early
                    print(f"[SAVE SERVICE] Created single meal entry for date {from_date}")
                except Exception as db_error:
                    db.session.rollback()
                    import traceback
                    error_trace = traceback.format_exc()
                    print(f"[SAVE SERVICE] Error creating meal: {str(db_error)}")
                    print(f"[SAVE SERVICE] Traceback: {error_trace}")
                    return jsonify({
                        'success': False,
                        'error': f'Database error: {str(db_error)}',
                        'field_errors': {}
                    }), 500

        elif service_type == 'arrival':
            from app.models.inbound import ArrivalBatch

            # Validate required fields - date and point are both required
            arrival_date_str = form_data.get('arrival_date', '').strip()
            arrival_point = form_data.get('arrival_point', '').strip()
            
            field_errors = {}
            if not arrival_date_str:
                field_errors['arrival_date'] = 'Arrival Date is required'
            if not arrival_point:
                field_errors['arrival_point'] = 'Arrival Point is required'
            
            if field_errors:
                return jsonify({
                    'success': False, 
                    'error': 'Arrival Date and Arrival Point are required',
                    'field_errors': field_errors
                }), 400

            # Parse date and time - date is required, time is optional
            arrival_date = None
            arrival_time = None
            try:
                arrival_date = datetime.strptime(arrival_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError) as e:
                return jsonify({
                    'success': False,
                    'error': f'Invalid arrival date format: {str(e)}',
                    'field_errors': {'arrival_date': 'Invalid date format'}
                }), 400
            
            if form_data.get('arrival_time'):
                try:
                    arrival_time = datetime.strptime(form_data['arrival_time'], '%H:%M').time()
                except (ValueError, TypeError):
                    pass  # Time is optional, so we don't fail if it's invalid

            # Check if we're updating an existing batch or creating new
            # Use row_id from main data (when editing from Trip Summary) or batch_id from form_data
            batch_id = row_id or form_data.get('arrival_id') or form_data.get('batch_id')
            if batch_id:
                arrival = ArrivalBatch.query.filter_by(id=batch_id, request_id=request_id).first()
            else:
                arrival = None

            if not arrival:
                # Create new batch
                arrival = ArrivalBatch(request_id=request_id)
                db.session.add(arrival)
                print(f"[SAVE SERVICE] Creating new arrival batch")
            else:
                print(f"[SAVE SERVICE] Updating existing arrival batch id={batch_id}")

            # Update fields - use only service-specific data, no fallbacks
            arrival.arrival_date = arrival_date
            arrival.arrival_point = arrival_point
            arrival.arrival_time = arrival_time
            arrival.flight_number = form_data.get('arrival_flight_number', '')
            arrival.vehicle_details = form_data.get('arrival_vehicle_type', '')
            # Use only arrival_pax_count, no fallback to request.pax
            pax_val = form_data.get('arrival_pax_count', '')
            arrival.pax_count = int(pax_val) if pax_val else None
            arrival.driver_name = form_data.get('arrival_driver_name', '')

            # New fields: visa_status, meet_assist, representative_name, notes
            arrival.visa_status = form_data.get('arrival_visa_status', 'NOT_INCLUDED')
            meet_assist_val = form_data.get('arrival_meet_assist', 'no')
            arrival.meet_assist = meet_assist_val in ['yes', 'true', 'True', True, 1, '1', 'on']
            arrival.representative_name = form_data.get('arrival_representative_name', '')
            arrival.status = form_data.get('arrival_status', 'REQUESTED')
            arrival.needs_transport = _parse_needs_transport(
                form_data.get('arrival_needs_transport'), default=True
            )
            # Notes: Use same simple pattern as arrival_point - strip and assign directly
            arrival.notes = form_data.get('arrival_notes', '').strip() or None
            print(f"[SAVE SERVICE] ===== ARRIVAL NOTES DEBUG =====")
            print(f"[SAVE SERVICE] Form data arrival_notes: {repr(form_data.get('arrival_notes'))}")
            print(f"[SAVE SERVICE] Saved to arrival.notes: {repr(arrival.notes)}")
            print(f"[SAVE SERVICE] Arrival ID: {arrival.id if arrival.id else 'NEW'}")
            print(f"[SAVE SERVICE] Arrival Point (for comparison): {repr(arrival.arrival_point)}")
            print(f"[SAVE SERVICE] ================================")

        elif service_type == 'departure':
            from app.models.inbound import DepartureBatch

            # Validate required fields - date and point are both required
            departure_date_str = form_data.get('departure_date', '').strip()
            departure_point = form_data.get('departure_point', '').strip()
            
            field_errors = {}
            if not departure_date_str:
                field_errors['departure_date'] = 'Departure Date is required'
            if not departure_point:
                field_errors['departure_point'] = 'Departure Point is required'
            
            if field_errors:
                return jsonify({
                    'success': False, 
                    'error': 'Departure Date and Departure Point are required',
                    'field_errors': field_errors
                }), 400

            # Parse date and time - date is required, time is optional
            departure_date = None
            departure_time = None
            try:
                departure_date = datetime.strptime(departure_date_str, '%Y-%m-%d').date()
            except (ValueError, TypeError) as e:
                return jsonify({
                    'success': False,
                    'error': f'Invalid departure date format: {str(e)}',
                    'field_errors': {'departure_date': 'Invalid date format'}
                }), 400
            
            if form_data.get('departure_time'):
                try:
                    departure_time = datetime.strptime(form_data['departure_time'], '%H:%M').time()
                except (ValueError, TypeError):
                    pass  # Time is optional, so we don't fail if it's invalid

            # Check if we're updating an existing batch or creating new
            # Use row_id from main data (when editing from Trip Summary) or batch_id from form_data
            batch_id = row_id or form_data.get('departure_id') or form_data.get('batch_id')
            if batch_id:
                departure = DepartureBatch.query.filter_by(id=batch_id, request_id=request_id).first()
            else:
                departure = None

            if not departure:
                # Create new batch
                departure = DepartureBatch(request_id=request_id)
                db.session.add(departure)
                print(f"[SAVE SERVICE] Creating new departure batch")
            else:
                print(f"[SAVE SERVICE] Updating existing departure batch id={batch_id}")

            # Update fields - use only service-specific data, no fallbacks
            departure.departure_date = departure_date
            departure.departure_point = departure_point
            departure.departure_time = departure_time
            departure.flight_number = form_data.get('departure_flight_number', '')
            # Use only departure_pax_count, no fallback to request.pax
            pax_val = form_data.get('departure_pax_count', '')
            departure.pax_count = int(pax_val) if pax_val else None
            departure.driver_name = form_data.get('departure_driver_name', '')
            departure.departure_tax = form_data.get('departure_tax', 'NOT_INCLUDED')

            # New fields: meet_assist, representative_name, notes
            meet_assist_val = form_data.get('departure_meet_assist', 'no')
            departure.meet_assist = meet_assist_val in ['yes', 'true', 'True', True, 1, '1', 'on']
            departure.representative_name = form_data.get('departure_representative_name', '')
            departure.status = form_data.get('departure_status', 'REQUESTED')
            departure.notes = form_data.get('departure_notes', '')
            departure.needs_transport = _parse_needs_transport(
                form_data.get('departure_needs_transport'), default=True
            )

            # Parse program date
            if form_data.get('departure_program_date'):
                try:
                    departure.program_date = datetime.strptime(form_data['departure_program_date'], '%Y-%m-%d').date()
                except:
                    pass

        # Ensure all changes are flushed before commit
        db.session.flush()
        if service_type == 'guide':
            guide_record_id_for_json = guide.id
        if service_type == 'arrival':
            _sync_transport_from_arrival(arrival)
        elif service_type == 'departure':
            _sync_transport_from_departure(departure)
        
        # Debug: Check notes value before commit
        if service_type == 'arrival':
            print(f"[SAVE SERVICE] Before commit - arrival.notes: {repr(getattr(arrival, 'notes', 'NOT_FOUND'))}")
            print(f"[SAVE SERVICE] Before commit - arrival.id: {arrival.id if arrival.id else 'NEW'}")
        
        db.session.commit()
        print(f"[SAVE SERVICE] Commit successful for {service_type}")
        
        # Debug: Verify notes after commit
        if service_type == 'arrival':
            db.session.refresh(arrival)
            print(f"[SAVE SERVICE] After commit - arrival.notes: {repr(getattr(arrival, 'notes', 'NOT_FOUND'))}")
            print(f"[SAVE SERVICE] After commit - arrival.id: {arrival.id}")

        # Expire cached data and re-query with eager loading
        db.session.expire_all()

        # Re-query fresh request with all service relationships eagerly loaded
        # Use filter_by().first() instead of .get() to ensure fresh query
        # Include arrival_batches and departure_batches for flights summary
        from app.models.inbound import ArrivalBatch, DepartureBatch
        fresh_request = InboundRequest.query.filter_by(id=request_id).options(
            selectinload(InboundRequest.inbound_hotels),
            selectinload(InboundRequest.inbound_transports),
            selectinload(InboundRequest.inbound_guides),
            selectinload(InboundRequest.inbound_meals),
            selectinload(InboundRequest.itinerary_rows),
            selectinload(InboundRequest.arrival_batches),
            selectinload(InboundRequest.departure_batches)
        ).first()
        
        if not fresh_request:
            return jsonify({'success': False, 'error': 'Request not found after save'}), 404

        # Build service lookup maps for template
        hotel_map = {h.source_itinerary_id: h for h in fresh_request.inbound_hotels}
        transport_map = {t.source_itinerary_id: t for t in fresh_request.inbound_transports}
        guide_map = {g.source_itinerary_id: g for g in fresh_request.inbound_guides}
        meal_map = {m.source_itinerary_id: m for m in fresh_request.inbound_meals}

        # Global fallbacks (source_itinerary_id=None)
        global_hotel = hotel_map.get(None)
        global_transport = transport_map.get(None)
        global_guide = guide_map.get(None)
        global_meal = meal_map.get(None)

        # Render updated HTML partials for instant DOM update
        sorted_itin = sort_itinerary_rows_with_children(fresh_request.itinerary_rows)
        itinerary_html = render_template(
            'components/itinerary_rows.html',
            rows=sorted_itin,
            request=fresh_request,
            inbound_request=fresh_request,
            guide_suppliers=_guide_suppliers_for_itinerary_ui(),
            itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(fresh_request),
            view_only=False,
            hotel_map=hotel_map,
            transport_map=transport_map,
            guide_map=guide_map,
            meal_map=meal_map,
            global_hotel=global_hotel,
            global_transport=global_transport,
            global_guide=global_guide,
            global_meal=global_meal
        )

        # Also render the service-specific entries table
        service_entries_html = None
        summary_entries_html = None

        if service_type == 'hotel':
            # Always use direct query to ensure fresh data and avoid relationship duplicates
            # This prevents duplicate entries that can occur with SQLAlchemy relationships
            hotels_list = InboundHotel.query.filter_by(request_id=request_id).order_by(InboundHotel.check_in_date).all()
            print(f"[SAVE SERVICE] Direct query found {len(hotels_list)} hotels")
            
            # Remove duplicates by hotel ID (safety check) - use dictionary for O(1) lookup
            seen_ids = {}
            unique_hotels = []
            duplicate_count = 0
            for hotel in hotels_list:
                if hotel.id not in seen_ids:
                    seen_ids[hotel.id] = True
                    unique_hotels.append(hotel)
                else:
                    duplicate_count += 1
                    print(f"[SAVE SERVICE] WARNING: Duplicate hotel detected (ID: {hotel.id}, Name: {hotel.hotel_name}) - skipping")
            
            if duplicate_count > 0:
                print(f"[SAVE SERVICE] Removed {duplicate_count} duplicate hotels")
            
            hotels_list = unique_hotels
            print(f"[SAVE SERVICE] After deduplication: {len(hotels_list)} unique hotels")
            
            # Final verification - ensure no duplicate IDs
            hotel_ids = [h.id for h in hotels_list]
            if len(hotel_ids) != len(set(hotel_ids)):
                print(f"[SAVE SERVICE] ERROR: Still have duplicate IDs after deduplication!")
                # Force unique by keeping first occurrence
                seen = set()
                hotels_list = [h for h in hotels_list if h.id not in seen and not seen.add(h.id)]
                print(f"[SAVE SERVICE] Force deduplicated to {len(hotels_list)} hotels")
            
            # Log all hotels for debugging
            for idx, h in enumerate(hotels_list):
                print(f"[SAVE SERVICE] Hotel {idx+1}: id={h.id}, name={h.hotel_name}, check_in={h.check_in_date}")
            
            service_entries_html = render_template(
                'components/hotel_entries.html',
                hotels=hotels_list,
                view_only=False
            )
            # Use Trip Summary format for summary_entries_html
            req_status = fresh_request.status
            mapped_status = 'CONFIRMED' if req_status in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS'] else ('INVOICED' if req_status in ['INVOICE', 'COMPLETED', 'INVOICED'] else req_status)
            summary_entries_html = render_template(
                'components/hotel_summary_entries.html',
                hotels=hotels_list,
                view_only=False,
                mapped_status=mapped_status
            )
            print(f"[SAVE SERVICE] summary_entries_html length: {len(summary_entries_html) if summary_entries_html else 0}")
            print(f"[SAVE SERVICE] summary_entries_html preview: {summary_entries_html[:200] if summary_entries_html else 'None'}")

        elif service_type == 'transport':
            # Always use direct query to ensure fresh data (same pattern as hotels)
            transports_list = InboundTransport.query.filter_by(request_id=request_id).order_by(InboundTransport.date).all()
            print(f"[SAVE SERVICE] Direct query found {len(transports_list)} transports")
            for t in transports_list:
                print(f"[SAVE SERVICE] Transport id={t.id}, note={repr(t.note)}")
            service_entries_html = render_template(
                'components/transport_entries.html',
                transports=transports_list,
                view_only=False
            )
            # Consolidate transports by vehicle+pickup+dropoff for summary
            transport_groups = {}
            for t in transports_list:
                key = f"{t.vehicle_type or ''}-{t.pickup_location or ''}-{t.dropoff_location or ''}"
                if key not in transport_groups:
                    transport_groups[key] = {
                        'vehicle_type': t.vehicle_type,
                        'pickup_location': t.pickup_location,
                        'dropoff_location': t.dropoff_location,
                        'status': t.status,
                        'dates': []
                    }
                if t.date:
                    transport_groups[key]['dates'].append(t.date)

            # Format date ranges for summary
            for key in transport_groups:
                dates = transport_groups[key]['dates']
                if dates:
                    dates.sort()
                    if len(dates) > 1:
                        transport_groups[key]['date_range'] = f"{dates[0].strftime('%d %b %Y')} - {dates[-1].strftime('%d %b %Y')}"
                    else:
                        transport_groups[key]['date_range'] = dates[0].strftime('%d %b %Y')
                else:
                    transport_groups[key]['date_range'] = '-'

            # Use Trip Summary format for summary_entries_html (same as hotel pattern)
            summary_entries_html = render_template(
                'components/transport_summary_entries.html',
                transports=_trip_summary_transports(transports_list),
                view_only=False
            )

        elif service_type == 'guide':
            # Always use direct query to ensure fresh data (same pattern as hotels)
            guides_list = InboundGuide.query.filter_by(request_id=request_id).order_by(InboundGuide.date).all()
            print(f"[SAVE SERVICE] Direct query found {len(guides_list)} guides")
            for g in guides_list:
                print(f"[SAVE SERVICE] Guide id={g.id}, name={g.guide_name}, additional_comments={repr(g.additional_comments)}")

            service_entries_html = render_template(
                'components/guide_entries.html',
                guides=guides_list,
                view_only=False
            )
            # Consolidate guides by name for summary
            guide_groups = {}
            for g in guides_list:
                key = g.guide_name or ''
                if key not in guide_groups:
                    guide_groups[key] = {
                        'guide_name': g.guide_name,
                        'language': g.language,
                        'telephone': g.telephone_number,
                        'dates': []
                    }
                if g.date:
                    guide_groups[key]['dates'].append(g.date)

            # Format date ranges for summary
            for key in guide_groups:
                dates = guide_groups[key]['dates']
                if dates:
                    dates.sort()
                    if len(dates) > 1:
                        guide_groups[key]['date_range'] = f"{dates[0].strftime('%d %b %Y')} - {dates[-1].strftime('%d %b %Y')}"
                    else:
                        guide_groups[key]['date_range'] = dates[0].strftime('%d %b %Y')
                else:
                    guide_groups[key]['date_range'] = '-'

            # Generate individual guide records for Trip Summary (not grouped)
            # Use properties: service_date, telephone
            # Use Trip Summary format for summary_entries_html (same as hotel pattern)
            # CRITICAL: Always generate summary_entries_html for guides (never None)
            summary_entries_html = render_template(
                'components/guide_summary_entries.html',
                guides=guides_list,
                view_only=False
            )
            # Ensure summary_entries_html is never None for guides
            if not summary_entries_html:
                summary_entries_html = '<tr><td colspan="6" class="border border-gray-300 px-2 py-3 text-center text-gray-500">No guides added</td></tr>'
            print(f"[SAVE SERVICE] Guide summary_entries_html length: {len(summary_entries_html) if summary_entries_html else 0}")
            print(f"[SAVE SERVICE] Guide summary_entries_html preview (first 500 chars): {summary_entries_html[:500] if summary_entries_html else 'None'}")
            # Count <td> tags in first row to verify column count
            if summary_entries_html and '<td' in summary_entries_html:
                first_row_start = summary_entries_html.find('<tr')
                first_row_end = summary_entries_html.find('</tr>', first_row_start)
                if first_row_start != -1 and first_row_end != -1:
                    first_row = summary_entries_html[first_row_start:first_row_end]
                    td_count = first_row.count('<td')
                    print(f"[SAVE SERVICE] Guide first row has {td_count} <td> tags (should be 6)")
                    if td_count != 6:
                        print(f"[SAVE SERVICE] ERROR: Guide row has wrong column count! First row HTML: {first_row}")

        elif service_type == 'meal':
            # Always use direct query to ensure fresh data (same pattern as hotels)
            meals_list = InboundMeal.query.filter_by(request_id=request_id).order_by(InboundMeal.date).all()
            print(f"[SAVE SERVICE] Direct query found {len(meals_list)} meals")
            for m in meals_list:
                print(f"[SAVE SERVICE] Meal id={m.id}, meal_note={repr(m.meal_note)}")
            service_entries_html = render_template(
                'components/meal_entries.html',
                meals=meals_list,
                view_only=False
            )
            # Consolidate meals by type+restaurant for summary
            meal_groups = {}
            for m in meals_list:
                key = f"{m.meal_type or ''}-{m.restaurant or ''}"
                if key not in meal_groups:
                    meal_groups[key] = {
                        'meal_type': m.meal_type,
                        'restaurant': m.restaurant,
                        'location': getattr(m, 'location', None),
                        'total_cost': m.total_cost or 0,
                        'meal_note': m.meal_note or '',  # Get notes from first meal in group
                        'first_meal_id': m.id,  # Store first meal ID for edit/remove actions
                        'dates': []
                    }
                if m.date:
                    meal_groups[key]['dates'].append(m.date)
                # If this meal has notes and the group doesn't, use this meal's notes
                if m.meal_note and not meal_groups[key]['meal_note']:
                    meal_groups[key]['meal_note'] = m.meal_note

            # Format date ranges for summary
            for key in meal_groups:
                dates = meal_groups[key]['dates']
                if dates:
                    dates.sort()
                    if len(dates) > 1:
                        meal_groups[key]['date_range'] = f"{dates[0].strftime('%d %b %Y')} - {dates[-1].strftime('%d %b %Y')}"
                    else:
                        meal_groups[key]['date_range'] = dates[0].strftime('%d %b %Y')
                else:
                    meal_groups[key]['date_range'] = '-'

            # Use Trip Summary format for summary_entries_html (same as hotel pattern)
            # Use individual meals instead of grouped meals for consistency
            req_status = fresh_request.status
            meal_mapped_status = 'CONFIRMED' if req_status in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS'] else ('INVOICED' if req_status in ['INVOICE', 'COMPLETED', 'INVOICED'] else req_status)
            try:
                summary_entries_html = render_template(
                    'components/meal_summary_entries.html',
                    meals=meals_list,
                    request=fresh_request,
                    view_only=False,
                    mapped_status=meal_mapped_status
                )
            except Exception as template_error:
                print(f"[SAVE SERVICE] Error rendering meal summary template: {str(template_error)}")
                # Fallback to empty table row if template rendering fails
                summary_entries_html = '<tr><td colspan="8" class="border border-gray-300 px-2 py-3 text-center text-gray-500">Error loading meals</td></tr>'

        elif service_type in ('arrival', 'departure'):
            from app.models.inbound import ArrivalBatch, DepartureBatch

            arrivals = ArrivalBatch.query.filter_by(request_id=request_id).order_by(ArrivalBatch.arrival_date).all()
            departures = DepartureBatch.query.filter_by(request_id=request_id).order_by(DepartureBatch.departure_date).all()

            # For flights, service_entries_html and summary_entries_html are the same
            # (flights are unique entries, not consolidated by date range)
            flights_html = render_template_string('''
                {% for arr in arrivals %}
                <tr class="hover:bg-gray-50" data-service-type="arrival" data-record-id="{{ arr.id }}" data-visa-status="{{ arr.visa_status or '' }}">
                    <td class="border border-gray-300 px-2 py-1.5"><span class="px-2 py-0.5 rounded-full text-[10px] bg-green-100 text-green-800"><i class="fas fa-plane-arrival mr-1"></i>Arrival</span></td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ arr.arrival_date.strftime('%d %b %Y') if arr.arrival_date else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ arr.arrival_time.strftime('%H:%M') if arr.arrival_time else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ arr.arrival_point or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ arr.flight_number or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ arr.pax_count or 0 }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ arr.representative_name if arr.meet_assist else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ arr.notes or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center whitespace-nowrap print-hide-actions">
                        <div class="flex items-center justify-center gap-2">
                            <button onclick="handleEditServiceRow('arrival', {{ arr.id }})" class="p-1.5 bg-gray-100 hover:bg-gray-200 rounded text-gray-600 hover:text-gray-800 transition-colors" title="Edit"><i class="fas fa-edit text-sm"></i></button>
                            <button onclick="handleRemoveServiceRow('arrival', {{ arr.id }})" class="p-1.5 bg-gray-700 hover:bg-gray-800 rounded text-white transition-colors" title="Remove"><i class="fas fa-trash text-sm"></i></button>
                        </div>
                    </td>
                </tr>
                {% endfor %}
                {% for dep in departures %}
                <tr class="hover:bg-gray-50" data-service-type="departure" data-record-id="{{ dep.id }}" data-departure-tax="{{ dep.departure_tax or '' }}">
                    <td class="border border-gray-300 px-2 py-1.5"><span class="px-2 py-0.5 rounded-full text-[10px] bg-orange-100 text-orange-800"><i class="fas fa-plane-departure mr-1"></i>Departure</span></td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ dep.departure_date.strftime('%d %b %Y') if dep.departure_date else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ dep.departure_time.strftime('%H:%M') if dep.departure_time else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ dep.departure_point or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ dep.flight_number or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center">{{ dep.pax_count or 0 }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ dep.representative_name if dep.meet_assist else '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5">{{ dep.notes or '-' }}</td>
                    <td class="border border-gray-300 px-2 py-1.5 text-center whitespace-nowrap print-hide-actions">
                        <div class="flex items-center justify-center gap-2">
                            <button onclick="handleEditServiceRow('departure', {{ dep.id }})" class="p-1.5 bg-gray-100 hover:bg-gray-200 rounded text-gray-600 hover:text-gray-800 transition-colors" title="Edit"><i class="fas fa-edit text-sm"></i></button>
                            <button onclick="handleRemoveServiceRow('departure', {{ dep.id }})" class="p-1.5 bg-gray-700 hover:bg-gray-800 rounded text-white transition-colors" title="Remove"><i class="fas fa-trash text-sm"></i></button>
                        </div>
                    </td>
                </tr>
                {% endfor %}
                {% if not arrivals and not departures %}
                <tr><td colspan="9" class="border border-gray-300 px-2 py-3 text-center text-gray-500">No flights added</td></tr>
                {% endif %}
            ''', arrivals=arrivals, departures=departures)

            # Use same HTML for both (flights are unique entries)
            service_entries_html = flights_html
            summary_entries_html = flights_html

        response_data = {
            'success': True, 
            'message': f'{service_type.capitalize()} data saved',
            'itinerary_html': itinerary_html,
            'service_entries_html': service_entries_html,
            'summary_entries_html': summary_entries_html,
            'service_type': service_type
        }
        if service_type == 'guide':
            response_data['itinerary_guide_slot_saved'] = _itinerary_guide_slot_saved_map(fresh_request)
            if guide_record_id_for_json:
                response_data['guide_id'] = guide_record_id_for_json

        if service_type in ('arrival', 'departure'):
            transports_list = InboundTransport.query.filter_by(request_id=request_id).order_by(InboundTransport.date).all()
            response_data['transport_summary_entries_html'] = render_template(
                'components/transport_summary_entries.html',
                transports=_trip_summary_transports(transports_list),
                view_only=False
            )

        # For hotel: include hotel_id and rooms data for room distribution modal
        if service_type == 'hotel' and 'hotel' in locals():
            hotel_obj = hotel  # Reference the hotel object created/updated above
            if hotel_obj and hotel_obj.id:
                rooms = HotelRoom.query.filter_by(hotel_id=hotel_obj.id).all()
                response_data['hotel_id'] = hotel_obj.id
                response_data['rooms'] = [{
                    'id': r.id,
                    'room_type': r.room_type,
                    'room_category': r.room_category or '',
                    'room_option': r.room_option or '',
                    'board_basis': r.board_basis or 'BB',
                    'dietary_requirements': r.dietary_requirements or '',
                    'adults': r.adults or 1,
                    'children': r.children or 0,
                    'guest_names': r.guest_names or ''
                } for r in rooms]

        # For transport: include transport_id for file upload
        if service_type == 'transport' and 'transport' in locals():
            if transport and transport.id:
                response_data['transport_id'] = transport.id

        # For arrival: include arrival_id for Attach Visa file upload
        if service_type == 'arrival' and 'arrival' in locals():
            if arrival and arrival.id:
                response_data['arrival_id'] = arrival.id

        # For meal: include meal_id for file upload
        print(f"[SAVE SERVICE] Building response for service_type={service_type}")
        print(f"[SAVE SERVICE] 'meal' in locals(): {'meal' in locals()}")

        if service_type == 'meal':
            if 'meal' in locals():
                print(f"[SAVE SERVICE] meal variable exists: {meal}")
                print(f"[SAVE SERVICE] meal.id = {meal.id if meal else 'None'}")
                if meal and meal.id:
                    response_data['meal_id'] = meal.id
                    print(f"[SAVE SERVICE] ADDED meal_id={meal.id} to response_data")
                else:
                    print(f"[SAVE SERVICE] DID NOT ADD meal_id - meal is {meal}, or meal.id is falsy")
            else:
                print(f"[SAVE SERVICE] ERROR: meal not in locals for meal service!")
                # Try to find it anyway
                try:
                    if meal and meal.id:
                        response_data['meal_id'] = meal.id
                        print(f"[SAVE SERVICE] Found meal in outer scope, added meal_id={meal.id}")
                except NameError:
                    print(f"[SAVE SERVICE] meal variable not found at all!")

        print(f"[SAVE SERVICE] Final response_data keys: {list(response_data.keys())}")
        print(f"[SAVE SERVICE] Final response_data: {response_data}")
        return jsonify(response_data)

    except OSError as os_err:
        db.session.rollback()
        # OSError [Errno 22] Invalid argument can occur on Windows with date/strftime operations
        return jsonify({'success': False, 'error': 'An unexpected error occurred while saving. Please try again.'}), 500
    except Exception as e:
        db.session.rollback()
        # Sanitize internal errors for user display
        err_msg = str(e)
        if 'OSError' in err_msg or 'Errno 22' in err_msg or 'Werkzeug' in err_msg or 'Invalid argument' in err_msg:
            err_msg = 'An unexpected error occurred while saving. Please try again.'
        return jsonify({'success': False, 'error': err_msg}), 500

@inbound_bp.route('/api/itinerary-row/<int:row_id>/update-field', methods=['POST'])
@csrf.exempt
def api_update_itinerary_row_field(row_id):
    """Update a single field on an itinerary row"""
    try:
        row = ItineraryRow.query.get_or_404(row_id)
        data = request.get_json()

        field = data.get('field')
        value = data.get('value')

        # Only allow specific fields to be updated
        allowed_fields = ['description', 'restaurant', 'meal_type', 'comment', 'cash_expense', 'restaurant_supplier_id', 'note', 'current_pax', 'itinerary_guide_supplier_ids']

        if field not in allowed_fields:
            return jsonify({'success': False, 'error': f'Field {field} not allowed'}), 400

        # Update the field
        if field == 'itinerary_guide_supplier_ids':
            if isinstance(value, list):
                row.set_itinerary_guide_supplier_id_list(value)
            elif isinstance(value, str):
                try:
                    parsed = json.loads(value)
                    row.set_itinerary_guide_supplier_id_list(parsed if isinstance(parsed, list) else [])
                except (json.JSONDecodeError, TypeError):
                    row.set_itinerary_guide_supplier_id_list([])
            else:
                row.set_itinerary_guide_supplier_id_list([])
            db.session.commit()
            return jsonify({'success': True})
        if field == 'cash_expense':
            try:
                value = float(value) if value else 0.0
            except ValueError:
                value = 0.0
        elif field == 'restaurant_supplier_id':
            try:
                value = int(value) if value else None
            except ValueError:
                value = None
        elif field in ('note', 'current_pax'):
            # Store in comment JSON
            import json
            comment_data = {}
            if row.comment:
                try:
                    comment_data = json.loads(row.comment)
                except (json.JSONDecodeError, TypeError):
                    if row.comment and str(row.comment).strip():
                        comment_data['_text'] = row.comment
            if field == 'note':
                comment_data['note'] = (value or '').strip() if isinstance(value, str) else ''
            else:
                try:
                    comment_data['current_pax'] = int(value) if value else None
                except (ValueError, TypeError):
                    comment_data['current_pax'] = None
            row.comment = json.dumps(comment_data)
        else:
            setattr(row, field, value)
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/itinerary-row', methods=['POST'])
@csrf.exempt
def api_save_itinerary_row(request_id):
    """Add or update an itinerary row"""
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        # Authorization check
        if request_obj.user_id != 1:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        data = request.get_json()

        row_id = data.get('row_id')
        itinerary_date = data.get('itinerary_date')
        description = data.get('description', '')
        meal_type = data.get('meal_type', '')
        restaurant_supplier_id = data.get('restaurant_supplier_id')
        current_pax = data.get('current_pax')
        parent_row_id = data.get('parent_row_id')  # For child rows
        note = data.get('note').strip() if isinstance(data.get('note'), str) else ''

        # Parse date (use from_date as default for new rows)
        date_obj = None
        if itinerary_date:
            date_obj = datetime.strptime(itinerary_date, '%Y-%m-%d').date()
        elif request_obj.from_date:
            date_obj = request_obj.from_date
        else:
            date_obj = datetime.now().date()

        # Parse restaurant supplier ID
        if restaurant_supplier_id:
            try:
                restaurant_supplier_id = int(restaurant_supplier_id)
            except ValueError:
                restaurant_supplier_id = None
        else:
            restaurant_supplier_id = None

        if row_id:
            # Update existing row
            row = ItineraryRow.query.filter_by(id=row_id, request_id=request_id).first()
            if not row:
                return jsonify({'success': False, 'error': 'Row not found'}), 404

            row.date = date_obj
            row.description = description
            row.meal_type = meal_type
            row.restaurant_supplier_id = restaurant_supplier_id
            
            # Handle current_pax and note - store in comment field as JSON (keep each row isolated)
            import json
            comment_data = {}
            if row.comment:
                try:
                    comment_data = json.loads(row.comment)
                except (json.JSONDecodeError, TypeError):
                    if row.comment and str(row.comment).strip():
                        comment_data['_text'] = row.comment
            if current_pax is not None:
                try:
                    comment_data['current_pax'] = int(current_pax)
                except (ValueError, TypeError):
                    pass
            if 'note' in data:
                comment_data['note'] = note
            if comment_data:
                row.comment = json.dumps(comment_data)
        else:
            # Create new row
            # If parent_row_id exists, allow empty description (for repeated/duplicated rows)
            # Otherwise, use default description for new standalone rows
            default_description = '' if parent_row_id else 'New day'
            row = ItineraryRow(
                request_id=request_id,
                date=date_obj,
                description=description if description is not None else default_description,
                meal_type=meal_type,
                restaurant_supplier_id=restaurant_supplier_id
            )
            
            # Handle current_pax and parent_row_id for new row - store in comment as JSON
            import json
            comment_data = {}
            if current_pax is not None:
                try:
                    comment_data['current_pax'] = int(current_pax)
                except (ValueError, TypeError):
                    pass
            if parent_row_id:
                try:
                    comment_data['parent_row_id'] = int(parent_row_id)
                except (ValueError, TypeError):
                    pass
            
            if comment_data:
                row.comment = json.dumps(comment_data)
            
            db.session.add(row)

        db.session.commit()

        return jsonify({'success': True, 'row_id': row.id})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/itinerary-row/<int:row_id>/delete', methods=['POST'])
@csrf.exempt
def api_delete_itinerary_row_by_id(row_id):
    """Delete an itinerary row by ID"""
    try:
        row = ItineraryRow.query.get_or_404(row_id)

        # Authorization check
        request_obj = InboundRequest.query.get_or_404(row.request_id)
        if request_obj.user_id != 1:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        db.session.delete(row)
        db.session.commit()

        return jsonify({'success': True})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

def sort_itinerary_rows_with_children(rows):
    """Sort itinerary rows ensuring child rows appear directly below their parent"""
    all_rows = list(rows)
    
    # Build parent-child mapping from comment field
    parent_child_map = {}  # parent_id -> list of child rows
    row_by_id = {}  # id -> row for quick lookup
    orphan_rows = []  # Child rows whose parent doesn't exist
    root_rows = []  # Rows without parent_row_id
    
    # First pass: build row_by_id map
    for row in all_rows:
        row_by_id[row.id] = row
    
    # Second pass: classify rows
    for row in all_rows:
        parent_id = None
        if row.comment:
            try:
                comment_data = json.loads(row.comment)
                parent_id = comment_data.get('parent_row_id')
            except (json.JSONDecodeError, TypeError):
                pass
        
        if parent_id:
            # Check if parent exists
            if parent_id in row_by_id:
                # Parent exists - add to parent_child_map
                if parent_id not in parent_child_map:
                    parent_child_map[parent_id] = []
                parent_child_map[parent_id].append(row)
            else:
                # Parent doesn't exist - treat as orphan (add to root)
                orphan_rows.append(row)
        else:
            # No parent - this is a root row
            root_rows.append(row)
    
    # Sort root rows by date, then by ID
    root_rows.sort(key=lambda r: (r.date or date(1900, 1, 1), r.id))
    
    # Sort orphan rows by date, then by ID (treat as root rows)
    orphan_rows.sort(key=lambda r: (r.date or date(1900, 1, 1), r.id))
    
    # Sort children within each parent group by ID (creation order)
    for parent_id in parent_child_map:
        parent_child_map[parent_id].sort(key=lambda r: r.id)
    
    # Rebuild sorted list maintaining parent-child relationships
    sorted_rows = []
    processed_ids = set()
    
    def add_row_and_children(row):
        """Add a row and all its children recursively"""
        if row.id in processed_ids:
            return
        
        # Add the row
        sorted_rows.append(row)
        processed_ids.add(row.id)
        
        # Add all children of this row (they will appear directly below)
        if row.id in parent_child_map:
            for child in parent_child_map[row.id]:
                add_row_and_children(child)
    
    # Process all root rows (rows without parents) in date order
    for row in root_rows:
        add_row_and_children(row)
    
    # Add orphan rows at the end
    for row in orphan_rows:
        add_row_and_children(row)
    
    return sorted_rows

@inbound_bp.route('/api/<int:request_id>/itinerary-rows-html', methods=['GET'])
def api_get_itinerary_rows_html(request_id):
    """Get the HTML for the itinerary rows table"""
    from app.models.supplier import Supplier
    try:
        request_obj = InboundRequest.query.options(
            selectinload(InboundRequest.itinerary_rows),
            selectinload(InboundRequest.inbound_guides),
            selectinload(InboundRequest.arrival_batches),
            selectinload(InboundRequest.departure_batches)
        ).get_or_404(request_id)
        # Custom sorting to ensure child rows appear directly below their parent
        rows = sort_itinerary_rows_with_children(request_obj.itinerary_rows)

        # Get restaurant suppliers for the dropdown
        restaurant_suppliers = Supplier.query.filter_by(supplier_type='RESTAURANT', is_active=True).order_by(Supplier.name).all()

        return render_template('components/itinerary_rows.html',
            rows=rows,
            request=request_obj,
            inbound_request=request_obj,  # Explicit variable to avoid Flask request confusion
            restaurant_suppliers=restaurant_suppliers,
            guide_suppliers=_guide_suppliers_for_itinerary_ui(),
            itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(request_obj),
            view_only=False
        )

    except Exception as e:
        return f'<tr><td colspan="7" class="text-center text-red-500">Error loading itinerary: {str(e)}</td></tr>', 500

@inbound_bp.route('/api/<int:request_id>/itinerary-row-html/<int:row_id>', methods=['GET'])
def api_get_itinerary_row_html(request_id, row_id):
    """Get HTML for a single itinerary row (for inserting without full refresh)"""
    from app.models.supplier import Supplier
    try:
        request_obj = InboundRequest.query.options(
            selectinload(InboundRequest.itinerary_rows),
            selectinload(InboundRequest.inbound_hotels),
            selectinload(InboundRequest.inbound_transports),
            selectinload(InboundRequest.inbound_guides),
            selectinload(InboundRequest.inbound_meals),
        ).get_or_404(request_id)
        row = ItineraryRow.query.filter_by(id=row_id, request_id=request_id).first()
        if not row:
            return '', 404
        hotel_map = {h.source_itinerary_id: h for h in (request_obj.inbound_hotels or [])}
        transport_map = {t.source_itinerary_id: t for t in (request_obj.inbound_transports or [])}
        guide_map = {g.source_itinerary_id: g for g in (request_obj.inbound_guides or [])}
        meal_map = {m.source_itinerary_id: m for m in (request_obj.inbound_meals or [])}
        global_hotel = hotel_map.get(None)
        global_transport = transport_map.get(None)
        global_guide = guide_map.get(None)
        global_meal = meal_map.get(None)
        restaurant_suppliers = Supplier.query.filter_by(supplier_type='RESTAURANT', is_active=True).order_by(Supplier.name).all()
        return render_template('components/itinerary_rows.html',
            rows=[row],
            request=request_obj,
            inbound_request=request_obj,
            hotel_map=hotel_map,
            transport_map=transport_map,
            guide_map=guide_map,
            meal_map=meal_map,
            global_hotel=global_hotel,
            global_transport=global_transport,
            global_guide=global_guide,
            global_meal=global_meal,
            restaurant_suppliers=restaurant_suppliers,
            guide_suppliers=_guide_suppliers_for_itinerary_ui(),
            itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(request_obj),
            view_only=False
        )
    except Exception as e:
        return '', 500

@inbound_bp.route('/api/<int:request_id>/itinerary-summary-html', methods=['GET'])
def api_get_itinerary_summary_html(request_id):
    """Get the HTML for the itinerary summary table in Trip Summary"""
    from markupsafe import escape
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)
        rows = request_obj.itinerary_rows

        # Build HTML rows for the summary table - show ALL itinerary rows
        html_rows = []
        if rows:
            # Sort rows with parent-child relationships maintained
            sorted_rows = sort_itinerary_rows_with_children(rows)

            for row in sorted_rows:
                # Extract PAX from comment field if it exists
                pax_value = ''
                if row.comment:
                    try:
                        comment_data = json.loads(row.comment)
                        if comment_data.get('current_pax') is not None:
                            pax_value = str(comment_data.get('current_pax'))
                    except (json.JSONDecodeError, TypeError):
                        pass
                
                date_str = row.date.strftime('%d %b %Y') if row.date else '-'
                description = escape(row.description) if row.description else '-'
                note_value = ''
                if row.comment:
                    try:
                        comment_data = json.loads(row.comment)
                        note_value = escape(comment_data.get('note', '') or '')
                    except (json.JSONDecodeError, TypeError):
                        if row.comment.strip() and '{' not in (row.comment or ''):
                            note_value = escape(row.comment) or ''

                description_block = f'''
                    <div class="flex flex-col gap-1">
                        <div>{description}</div>
                    </div>
                '''
                if note_value:
                    description_block = f'''
                        <div class="flex flex-col gap-1">
                            <div>{description}</div>
                            <div class="text-slate-600 border-t border-slate-200 pt-1" style="border-top: 1px solid #e2e8f0; padding-top: 0.25rem;">
                                <i class="fas fa-sticky-note mr-1 text-amber-500"></i>{note_value}
                            </div>
                        </div>
                    '''

                html_rows.append(f'''
                    <tr class="hover:bg-gray-50" data-service-type="itinerary" data-record-id="{row.id}">
                        <td class="text-center summary-date-col">{date_str}</td>
                        <td>{description_block}</td>
                        <td class="text-center summary-compact-col whitespace-nowrap print-hide-actions">
                            <div class="flex items-center justify-center gap-1">
                                <button onclick="handleEditServiceRow('itinerary', {row.id})" class="text-blue-600 hover:text-blue-800" title="Edit"><i class="fas fa-edit"></i></button>
                                <button onclick="handleRemoveServiceRow('itinerary', {row.id})" class="text-red-600 hover:text-red-800" title="Remove"><i class="fas fa-trash"></i></button>
                            </div>
                        </td>
                    </tr>
                ''')
            if html_rows:
                return ''.join(html_rows)
            else:
                return '<tr><td colspan="3" class="px-4 py-3 text-center text-gray-500">No itinerary rows added</td></tr>'
        else:
            return '<tr><td colspan="3" class="px-4 py-3 text-center text-gray-500">No itinerary rows added</td></tr>'

    except Exception as e:
        return f'<tr><td colspan="3" class="text-center text-red-500">Error loading itinerary: {escape(str(e))}</td></tr>', 500

@inbound_bp.route('/api/<int:request_id>/itinerary-rows-bulk', methods=['POST'])
def api_save_itinerary_rows_bulk(request_id):
    """Save all itinerary rows in bulk"""
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        # Authorization check
        if request_obj.user_id != 1:
            return jsonify({'success': False, 'error': 'Unauthorized'}), 403

        data = request.get_json()
        updates = data.get('updates', [])

        for update in updates:
            row_id = update.get('row_id')
            if not row_id:
                continue

            row = ItineraryRow.query.filter_by(id=row_id, request_id=request_id).first()
            if not row:
                continue

            # Update fields
            row.description = update.get('description', '')
            row.meal_type = update.get('meal_type', '')

            # Handle restaurant supplier ID
            restaurant_id = update.get('restaurant_supplier_id')
            if restaurant_id:
                try:
                    row.restaurant_supplier_id = int(restaurant_id)
                except ValueError:
                    row.restaurant_supplier_id = None
            else:
                row.restaurant_supplier_id = None
            
            # Handle current_pax, status and note - store in comment field as JSON
            current_pax = update.get('current_pax')
            note_val = (update.get('note', '') or '').strip()
            status_val = update.get('status', 'REQUEST')
            if current_pax is not None or 'note' in update or 'status' in update:
                try:
                    comment_data = json.loads(row.comment) if row.comment else {}
                except (json.JSONDecodeError, TypeError):
                    comment_data = {}
                try:
                    if current_pax is not None:
                        comment_data['current_pax'] = int(current_pax)
                    comment_data['note'] = note_val
                    comment_data['status'] = status_val
                    row.comment = json.dumps(comment_data)
                except (ValueError, TypeError):
                    fallback = {}
                    if current_pax is not None:
                        fallback['current_pax'] = int(current_pax)
                    fallback['note'] = note_val
                    fallback['status'] = status_val
                    row.comment = json.dumps(fallback)

            raw_guides = update.get('itinerary_guide_supplier_ids')
            if raw_guides is not None:
                if isinstance(raw_guides, list):
                    row.set_itinerary_guide_supplier_id_list(raw_guides)
                elif isinstance(raw_guides, str):
                    try:
                        parsed = json.loads(raw_guides)
                        row.set_itinerary_guide_supplier_id_list(parsed if isinstance(parsed, list) else [])
                    except (json.JSONDecodeError, TypeError):
                        row.set_itinerary_guide_supplier_id_list([])
                else:
                    row.set_itinerary_guide_supplier_id_list([])

        db.session.commit()

        fresh_req = InboundRequest.query.filter_by(id=request_id).options(
            selectinload(InboundRequest.inbound_guides),
            selectinload(InboundRequest.itinerary_rows),
        ).first()
        slot_saved = _itinerary_guide_slot_saved_map(fresh_req) if fresh_req else {}

        return jsonify({
            'success': True,
            'message': 'Itinerary saved successfully',
            'itinerary_guide_slot_saved': slot_saved,
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/add-hotel', methods=['POST'])
@csrf.exempt
def api_add_hotel():
    """Add a new hotel to the suppliers list"""
    from app.models.supplier import Supplier

    def _request_payload():
        return request.get_json(silent=True) or request.form

    def _save_contract_upload():
        file = request.files.get('contract_file')
        if not file or not file.filename:
            return None
        safe_name = secure_filename(file.filename)
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'supplier_contracts')
        os.makedirs(upload_dir, exist_ok=True)
        stamped = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_name}"
        file.save(os.path.join(upload_dir, stamped))
        return stamped

    try:
        data = _request_payload()
        hotel_name = data.get('name', '').strip()
        hotel_city = data.get('city', '').strip() or 'Amman'  # Default to Amman if empty

        if not hotel_name:
            return jsonify({'success': False, 'error': 'Hotel name is required'}), 400

        phone_val = data.get('phone', '').strip()
        if phone_val and not is_valid_phone(phone_val):
            return jsonify({'success': False, 'error': f'Phone: {PHONE_ERROR}'}), 400

        # Check if hotel already exists (with timeout protection)
        try:
            existing = Supplier.query.filter(
                Supplier.name == hotel_name,
                Supplier.supplier_type == 'HOTEL'
            ).first()
            if existing:
                return jsonify({'success': False, 'error': 'Hotel already exists'}), 400
        except Exception as e:
            print(f"[ERROR] Failed to check existing hotel: {e}")
            # Continue anyway - worst case is duplicate

        # Generate unique code (with timeout protection)
        import re
        city_code = re.sub(r'[^A-Z]', '', hotel_city.upper()[:3]) or 'OTH'
        try:
            # Check for existing codes with this pattern (limit query to prevent timeout)
            count = Supplier.query.filter(Supplier.code.like(f'HTL-{city_code}-%')).limit(1000).count()
            new_code = f'HTL-{city_code}-{count + 1:03d}'

            # Ensure code is unique (in case of race condition) - limit iterations
            max_iterations = 100
            iteration = 0
            while Supplier.query.filter_by(code=new_code).first() and iteration < max_iterations:
                count += 1
                new_code = f'HTL-{city_code}-{count + 1:03d}'
                iteration += 1
            
            if iteration >= max_iterations:
                # Fallback to timestamp-based code
                import time
                new_code = f'HTL-{city_code}-{int(time.time())}'
        except Exception as e:
            print(f"[ERROR] Failed to generate code: {e}")
            # Fallback to timestamp-based code
            import time
            new_code = f'HTL-{city_code}-{int(time.time())}'

        payment_terms = data.get('payment_terms', '').strip() or None
        payment_method = data.get('payment_method', '').strip()
        bank_name_val, bank_account_val = _resolve_supplier_bank_fields(data)
        contract_file = _save_contract_upload()

        # Create new supplier with all provided fields
        # Use provided city or None if empty (don't use default 'Amman' for storage)
        city_value = data.get('city', '').strip() or None
        # Store category and room_category in notes field as JSON if provided
        category = data.get('category', '').strip()
        room_category = data.get('room_category', '').strip()
        notes_value = data.get('notes', '').strip() or None
        if category or room_category:
            import json
            notes_dict = {}
            if notes_value:
                notes_dict['original_notes'] = notes_value
            if category:
                notes_dict['category'] = category
            if room_category:
                notes_dict['room_category'] = room_category
                # Also store in room_categories list for consistency with Room List fetching
                default_categories = ['Standard Rooms', 'Junior Suites', 'Executive Suites', 'Presidential Suites']
                if room_category not in default_categories:
                    notes_dict['room_categories'] = [room_category]
            if payment_method:
                notes_dict['payment_method'] = payment_method
            if contract_file:
                notes_dict['contract_file'] = contract_file
            notes_value = json.dumps(notes_dict)
        elif payment_method or contract_file:
            import json
            notes_dict = {}
            if notes_value:
                notes_dict['original_notes'] = notes_value
            if payment_method:
                notes_dict['payment_method'] = payment_method
            if contract_file:
                notes_dict['contract_file'] = contract_file
            notes_value = json.dumps(notes_dict)
        
        new_hotel = Supplier(
            name=hotel_name,
            code=new_code,
            supplier_type='HOTEL',
            city=city_value,
            country=data.get('country', '').strip() or None,
            contact_person=data.get('contact_person', '').strip() or None,
            email=data.get('email', '').strip() or None,
            phone=data.get('phone', '').strip() or None,
            website=data.get('website', '').strip() or None,
            payment_terms=payment_terms,
            default_currency=data.get('default_currency', 'USD') or 'USD',
            address=data.get('address', '').strip() or None,
            bank_name=bank_name_val,
            bank_account=bank_account_val,
            tax_number=data.get('tax_number', '').strip() or None,
            notes=notes_value,
            is_active=True
        )
        db.session.add(new_hotel)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()

        return jsonify({
            'success': True,
            'hotel': {
                'id': new_hotel.id,
                'name': new_hotel.name,
                'city': new_hotel.city,
                'category': category,  # Return category for immediate use
                'room_category': room_category
            }
        })

    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:supplier_id>/room-categories', methods=['GET'])
@csrf.exempt
def api_get_hotel_room_categories(supplier_id):
    """Get room categories for a hotel supplier"""
    from app.models.supplier import Supplier
    import json
    try:
        supplier = Supplier.query.get_or_404(supplier_id)
        default_categories = ['Standard Rooms', 'Junior Suites', 'Executive Suites', 'Presidential Suites']
        custom_categories = []
        
        # Extract custom categories from supplier notes JSON
        if supplier.notes:
            try:
                notes_dict = json.loads(supplier.notes)
                # Read from room_categories (list) - added via + button
                custom_categories = notes_dict.get('room_categories', [])
                # Also read from room_category (singular string) - added via Add Hotel modal
                single_cat = notes_dict.get('room_category', '')
                if single_cat and single_cat not in custom_categories and single_cat not in default_categories:
                    custom_categories.append(single_cat)
            except (json.JSONDecodeError, TypeError):
                pass
        
        # Also scan existing hotel rooms for any categories in use (catches categories
        # that were saved to rooms but might not be in the supplier notes)
        try:
            hotels = InboundHotel.query.filter_by(hotel_name=supplier.name).all()
            for hotel in hotels:
                for room in hotel.rooms:
                    if room.room_category and room.room_category not in custom_categories and room.room_category not in default_categories:
                        custom_categories.append(room.room_category)
        except Exception:
            pass  # Non-critical; don't fail the whole request
        
        # Merge defaults + custom, preserving order and uniqueness
        all_categories = list(default_categories)
        for cat in custom_categories:
            if cat not in all_categories:
                all_categories.append(cat)
        
        return jsonify({'success': True, 'categories': all_categories})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:supplier_id>/room-categories', methods=['POST'])
@csrf.exempt
def api_add_hotel_room_category(supplier_id):
    """Add a custom room category for a hotel supplier"""
    from app.models.supplier import Supplier
    import json
    try:
        supplier = Supplier.query.get_or_404(supplier_id)
        data = request.get_json()
        new_category = data.get('category', '').strip()
        
        if not new_category:
            return jsonify({'success': False, 'error': 'Category name is required'}), 400
        
        # Parse existing notes
        notes_dict = {}
        if supplier.notes:
            try:
                notes_dict = json.loads(supplier.notes)
            except (json.JSONDecodeError, TypeError):
                notes_dict = {'original_notes': supplier.notes}
        
        # Add to custom room_categories list
        room_categories = notes_dict.get('room_categories', [])
        default_categories = ['Standard Rooms', 'Junior Suites', 'Executive Suites', 'Presidential Suites']
        
        if new_category in room_categories or new_category in default_categories:
            return jsonify({'success': False, 'error': 'Category already exists'}), 400
        
        room_categories.append(new_category)
        notes_dict['room_categories'] = room_categories
        supplier.notes = json.dumps(notes_dict)
        db.session.commit()
        
        # Return full list
        all_categories = list(default_categories)
        for cat in room_categories:
            if cat not in all_categories:
                all_categories.append(cat)
        
        return jsonify({'success': True, 'categories': all_categories})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:supplier_id>/bed-types', methods=['GET'])
@csrf.exempt
def api_get_hotel_bed_types(supplier_id):
    """Get bed types for a hotel supplier"""
    from app.models.supplier import Supplier
    import json
    try:
        supplier = Supplier.query.get_or_404(supplier_id)
        default_bed_types = ['King', 'Queen', 'Twin', 'Triple']
        custom_bed_types = []

        # Extract custom bed types from supplier notes JSON
        if supplier.notes:
            try:
                notes_dict = json.loads(supplier.notes)
                # Read from bed_types (list) - added via + button
                custom_bed_types = notes_dict.get('bed_types', [])
            except (json.JSONDecodeError, TypeError):
                pass

        # Also scan existing hotel rooms for any bed types in use (catches values
        # that were saved to rooms but might not be in the supplier notes)
        try:
            hotels = InboundHotel.query.filter_by(hotel_name=supplier.name).all()
            for hotel in hotels:
                for room in hotel.rooms:
                    if room.room_option and room.room_option not in custom_bed_types and room.room_option not in default_bed_types:
                        custom_bed_types.append(room.room_option)
        except Exception:
            pass  # Non-critical; don't fail the whole request

        # Merge defaults + custom, preserving order and uniqueness
        all_bed_types = list(default_bed_types)
        for bt in custom_bed_types:
            if bt not in all_bed_types:
                all_bed_types.append(bt)

        return jsonify({'success': True, 'bed_types': all_bed_types})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:supplier_id>/bed-types', methods=['POST'])
@csrf.exempt
def api_add_hotel_bed_type(supplier_id):
    """Add a custom bed type for a hotel supplier"""
    from app.models.supplier import Supplier
    import json
    try:
        supplier = Supplier.query.get_or_404(supplier_id)
        data = request.get_json()
        new_bed_type = data.get('bed_type', '').strip()

        if not new_bed_type:
            return jsonify({'success': False, 'error': 'Bed type name is required'}), 400

        # Parse existing notes
        notes_dict = {}
        if supplier.notes:
            try:
                notes_dict = json.loads(supplier.notes)
            except (json.JSONDecodeError, TypeError):
                notes_dict = {'original_notes': supplier.notes}

        # Add to custom bed_types list
        bed_types = notes_dict.get('bed_types', [])
        default_bed_types = ['King', 'Queen', 'Twin', 'Triple']

        if new_bed_type in bed_types or new_bed_type in default_bed_types:
            return jsonify({'success': False, 'error': 'Bed type already exists'}), 400

        bed_types.append(new_bed_type)
        notes_dict['bed_types'] = bed_types
        supplier.notes = json.dumps(notes_dict)
        db.session.commit()

        # Return full list
        all_bed_types = list(default_bed_types)
        for bt in bed_types:
            if bt not in all_bed_types:
                all_bed_types.append(bt)

        return jsonify({'success': True, 'bed_types': all_bed_types})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/room-categories/global', methods=['GET'])
@csrf.exempt
def api_get_global_hotel_room_categories():
    """Get globally shared room categories across browsers/devices."""
    default_categories = ['Standard Rooms', 'Junior Suites', 'Executive Suites', 'Presidential Suites']
    try:
        global_path = os.path.join(current_app.instance_path, 'global_room_categories.json')
        categories = []
        if os.path.exists(global_path):
            try:
                with open(global_path, 'r', encoding='utf-8') as f:
                    parsed = json.load(f)
                if isinstance(parsed, list):
                    categories = [str(v).strip() for v in parsed if str(v).strip()]
            except Exception:
                categories = []
        merged = list(default_categories)
        for cat in categories:
            if cat not in merged:
                merged.append(cat)
        return jsonify({'success': True, 'categories': merged})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/room-categories/global', methods=['POST'])
@csrf.exempt
def api_add_global_hotel_room_category():
    """Persist a global room category shared by all browsers/devices."""
    default_categories = ['Standard Rooms', 'Junior Suites', 'Executive Suites', 'Presidential Suites']
    try:
        data = request.get_json(silent=True) or {}
        new_category = str(data.get('category', '')).strip()
        if not new_category:
            return jsonify({'success': False, 'error': 'Category name is required'}), 400

        global_path = os.path.join(current_app.instance_path, 'global_room_categories.json')
        os.makedirs(current_app.instance_path, exist_ok=True)

        categories = []
        if os.path.exists(global_path):
            try:
                with open(global_path, 'r', encoding='utf-8') as f:
                    parsed = json.load(f)
                if isinstance(parsed, list):
                    categories = [str(v).strip() for v in parsed if str(v).strip()]
            except Exception:
                categories = []

        all_existing = default_categories + categories
        if any(c.lower() == new_category.lower() for c in all_existing):
            merged_existing = list(default_categories)
            for cat in categories:
                if cat not in merged_existing:
                    merged_existing.append(cat)
            return jsonify({'success': True, 'categories': merged_existing})

        categories.append(new_category)
        unique = []
        for cat in categories:
            if not any(x.lower() == cat.lower() for x in unique):
                unique.append(cat)

        with open(global_path, 'w', encoding='utf-8') as f:
            json.dump(unique, f, ensure_ascii=False)

        merged = list(default_categories)
        for cat in unique:
            if cat not in merged:
                merged.append(cat)
        return jsonify({'success': True, 'categories': merged})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/supplier-option-values/global', methods=['GET'])
@csrf.exempt
def api_get_global_supplier_option_values():
    """Return globally shared custom select values by field key."""
    try:
        key = str(request.args.get('key', '')).strip()
        if not key:
            return jsonify({'success': False, 'error': 'key is required'}), 400

        global_path = os.path.join(current_app.instance_path, 'global_supplier_option_values.json')
        values_map = {}
        if os.path.exists(global_path):
            try:
                with open(global_path, 'r', encoding='utf-8') as f:
                    parsed = json.load(f)
                if isinstance(parsed, dict):
                    values_map = parsed
            except Exception:
                values_map = {}

        values = values_map.get(key, [])
        if not isinstance(values, list):
            values = []
        cleaned = [str(v).strip() for v in values if str(v).strip()]
        return jsonify({'success': True, 'values': cleaned})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/supplier-option-values/global', methods=['POST'])
@csrf.exempt
def api_add_global_supplier_option_value():
    """Persist a globally shared custom select value by field key."""
    try:
        payload = request.get_json(silent=True) or {}
        key = str(payload.get('key', '')).strip()
        value = str(payload.get('value', '')).strip()
        if not key or not value:
            return jsonify({'success': False, 'error': 'key and value are required'}), 400

        global_path = os.path.join(current_app.instance_path, 'global_supplier_option_values.json')
        os.makedirs(current_app.instance_path, exist_ok=True)

        values_map = {}
        if os.path.exists(global_path):
            try:
                with open(global_path, 'r', encoding='utf-8') as f:
                    parsed = json.load(f)
                if isinstance(parsed, dict):
                    values_map = parsed
            except Exception:
                values_map = {}

        existing = values_map.get(key, [])
        if not isinstance(existing, list):
            existing = []
        existing_clean = [str(v).strip() for v in existing if str(v).strip()]

        if not any(v.lower() == value.lower() for v in existing_clean):
            existing_clean.append(value)
        values_map[key] = existing_clean

        with open(global_path, 'w', encoding='utf-8') as f:
            json.dump(values_map, f, ensure_ascii=False)

        return jsonify({'success': True, 'values': existing_clean})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/supplier/<int:supplier_id>', methods=['GET'])
@csrf.exempt
def api_get_supplier(supplier_id):
    """Get supplier details by ID"""
    from app.models.supplier import Supplier
    try:
        supplier = Supplier.query.get_or_404(supplier_id)
        return jsonify({
            'success': True,
            'supplier': {
                'id': supplier.id,
                'name': supplier.name,
                'notes': supplier.notes
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/add-guide', methods=['POST'])
@csrf.exempt
def api_add_guide():
    """Add a new guide to the suppliers list"""
    from app.models.supplier import Supplier

    def _request_payload():
        return request.get_json(silent=True) or request.form

    def _save_contract_upload():
        file = request.files.get('contract_file')
        if not file or not file.filename:
            return None
        safe_name = secure_filename(file.filename)
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'supplier_contracts')
        os.makedirs(upload_dir, exist_ok=True)
        stamped = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_name}"
        file.save(os.path.join(upload_dir, stamped))
        return stamped

    try:
        data = _request_payload()
        guide_name = data.get('name', '').strip()

        if not guide_name:
            return jsonify({'success': False, 'error': 'Guide name is required'}), 400

        phone_val = data.get('phone', '').strip()
        if phone_val and not is_valid_phone(phone_val):
            return jsonify({'success': False, 'error': f'Phone: {PHONE_ERROR}'}), 400

        # Check if guide already exists
        existing = Supplier.query.filter_by(name=guide_name, supplier_type='GUIDE').first()
        if existing:
            return jsonify({'success': False, 'error': 'Guide already exists'}), 400

        # Generate unique code
        count = Supplier.query.filter(Supplier.code.like('GDE-%')).count()
        new_code = f'GDE-{count + 1:03d}'

        # Validate languages (mandatory for guides)
        languages = data.get('languages', '').strip()
        if not languages:
            return jsonify({'success': False, 'error': 'At least one language is required'}), 400
        
        payment_terms = data.get('payment_terms', '').strip() or None
        payment_method = data.get('payment_method', '').strip()
        bank_name_val, bank_account_val = _resolve_supplier_bank_fields(data)
        contract_file = _save_contract_upload()
        notes_value = data.get('notes', '').strip() or ''
        if contract_file:
            notes_value = f"Contract file: {contract_file}\n{notes_value}".strip()
        notes_value = _merge_payment_method_into_notes(notes_value, payment_method) or None

        # Create new supplier with all provided fields
        new_guide = Supplier(
            name=guide_name,
            code=new_code,
            supplier_type='GUIDE',
            languages=languages,  # Store comma-separated languages
            phone=data.get('phone', '').strip() or None,
            email=data.get('email', '').strip() or None,
            city=data.get('city', '').strip() or None,
            country=data.get('country', '').strip() or None,
            payment_terms=payment_terms,
            default_currency=data.get('default_currency', 'USD') or 'USD',
            address=data.get('address', '').strip() or None,
            bank_name=bank_name_val,
            bank_account=bank_account_val,
            tax_number=data.get('tax_number', '').strip() or None,
            notes=notes_value or None,
            is_active=True
        )
        db.session.add(new_guide)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()

        return jsonify({
            'success': True,
            'supplier_id': new_guide.id,
            'guide': {
                'id': new_guide.id,
                'name': new_guide.name,
                'languages': new_guide.languages,
                'phone': new_guide.phone,
                'email': new_guide.email,
                'contact_person': new_guide.contact_person,
                'website': new_guide.website,
                'address': new_guide.address,
                'city': new_guide.city,
                'country': new_guide.country,
                'payment_terms': new_guide.payment_terms,
                'default_currency': new_guide.default_currency,
                'bank_name': new_guide.bank_name,
                'bank_account': new_guide.bank_account,
                'tax_number': new_guide.tax_number,
                'notes': new_guide.notes,
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/add-restaurant', methods=['POST'])
@csrf.exempt
def api_add_restaurant():
    """Add a new restaurant to the suppliers list"""
    from app.models.supplier import Supplier

    def _request_payload():
        return request.get_json(silent=True) or request.form

    def _save_contract_upload():
        file = request.files.get('contract_file')
        if not file or not file.filename:
            return None
        safe_name = secure_filename(file.filename)
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'supplier_contracts')
        os.makedirs(upload_dir, exist_ok=True)
        stamped = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_name}"
        file.save(os.path.join(upload_dir, stamped))
        return stamped

    try:
        data = _request_payload()
        restaurant_name = data.get('name', '').strip()

        if not restaurant_name:
            return jsonify({'success': False, 'error': 'Restaurant name is required'}), 400

        phone_val = data.get('phone', '').strip()
        if phone_val and not is_valid_phone(phone_val):
            return jsonify({'success': False, 'error': f'Phone: {PHONE_ERROR}'}), 400

        # Check if restaurant already exists
        existing = Supplier.query.filter_by(name=restaurant_name, supplier_type='RESTAURANT').first()
        if existing:
            return jsonify({'success': False, 'error': 'Restaurant already exists'}), 400

        # Generate unique code
        count = Supplier.query.filter(Supplier.code.like('RST-%')).count()
        new_code = f'RST-{count + 1:03d}'

        payment_terms = data.get('payment_terms', '').strip() or None
        payment_method = data.get('payment_method', '').strip()
        bank_name_val, bank_account_val = _resolve_supplier_bank_fields(data)
        contract_file = _save_contract_upload()
        notes_value = data.get('notes', '').strip() or ''
        if contract_file:
            notes_value = f"Contract file: {contract_file}\n{notes_value}".strip()
        notes_value = _merge_payment_method_into_notes(notes_value, payment_method) or None

        # Create new supplier with all provided fields
        new_restaurant = Supplier(
            name=restaurant_name,
            code=new_code,
            supplier_type='RESTAURANT',
            address=data.get('address', '').strip() or data.get('location', '').strip() or None,
            city=data.get('city', '').strip() or None,
            country=data.get('country', '').strip() or None,
            contact_person=data.get('contact_person', '').strip() or None,
            email=data.get('email', '').strip() or None,
            phone=data.get('phone', '').strip() or None,
            website=data.get('website', '').strip() or None,
            payment_terms=payment_terms,
            default_currency=data.get('default_currency', 'USD') or 'USD',
            bank_name=bank_name_val,
            bank_account=bank_account_val,
            tax_number=data.get('tax_number', '').strip() or None,
            notes=notes_value or None,
            is_active=True
        )
        db.session.add(new_restaurant)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()

        return jsonify({
            'success': True,
            'supplier_id': new_restaurant.id,
            'restaurant': {
                'id': new_restaurant.id,
                'name': new_restaurant.name,
                'location': new_restaurant.address
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/add-transport', methods=['POST'])
@csrf.exempt
def api_add_transport():
    """Add a new transport supplier to the suppliers list"""
    from app.models.supplier import Supplier

    def _request_payload():
        return request.get_json(silent=True) or request.form

    def _save_contract_upload():
        file = request.files.get('contract_file')
        if not file or not file.filename:
            return None
        safe_name = secure_filename(file.filename)
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'supplier_contracts')
        os.makedirs(upload_dir, exist_ok=True)
        stamped = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_name}"
        file.save(os.path.join(upload_dir, stamped))
        return stamped

    try:
        data = _request_payload()
        transport_name = data.get('name', '').strip()

        if not transport_name:
            return jsonify({'success': False, 'error': 'Supplier name is required'}), 400

        phone_val = data.get('phone', '').strip()
        if phone_val and not is_valid_phone(phone_val):
            return jsonify({'success': False, 'error': f'Phone: {PHONE_ERROR}'}), 400

        # Check if transport supplier already exists
        existing = Supplier.query.filter_by(name=transport_name, supplier_type='TRANSPORT').first()
        if existing:
            return jsonify({'success': False, 'error': 'Transport supplier already exists'}), 400

        # Generate unique code
        count = Supplier.query.filter(Supplier.code.like('TRN-%')).count()
        new_code = f'TRN-{count + 1:03d}'

        entity_type = data.get('entity_type', 'COMPANY').strip() or 'COMPANY'
        payment_terms = data.get('payment_terms', '').strip() or None
        payment_method = data.get('payment_method', '').strip()
        bank_name_val, bank_account_val = _resolve_supplier_bank_fields(data)
        contract_file = _save_contract_upload()
        notes_value = data.get('notes', '').strip() or ''
        if contract_file:
            notes_value = f"Contract file: {contract_file}\n{notes_value}".strip()
        notes_value = _merge_payment_method_into_notes(notes_value, payment_method) or None

        # Create new supplier with all provided fields
        new_transport = Supplier(
            name=transport_name,
            code=new_code,
            supplier_type='TRANSPORT',
            entity_type=entity_type,
            phone=data.get('phone', '').strip() or None,
            contact_person=data.get('contact_person', '').strip() or None,
            email=data.get('email', '').strip() or None,
            website=data.get('website', '').strip() or None,
            city=data.get('city', '').strip() or None,
            country=data.get('country', '').strip() or None,
            payment_terms=payment_terms,
            payment_method=payment_method,
            default_currency=data.get('default_currency', 'USD') or 'USD',
            address=data.get('address', '').strip() or None,
            bank_name=bank_name_val,
            bank_account=bank_account_val,
            tax_number=data.get('tax_number', '').strip() or None,
            notes=notes_value or None,
            is_active=True
        )
        db.session.add(new_transport)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()

        return jsonify({
            'success': True,
            'supplier_id': new_transport.id,
            'transport': {
                'id': new_transport.id,
                'name': new_transport.name,
                'phone': new_transport.phone
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/add-ground-handler', methods=['POST'])
@csrf.exempt
def api_add_ground_handler():
    """Add a new ground handler supplier to the suppliers list"""
    from app.models.supplier import Supplier

    def _request_payload():
        return request.get_json(silent=True) or request.form

    def _save_contract_upload():
        file = request.files.get('contract_file')
        if not file or not file.filename:
            return None
        safe_name = secure_filename(file.filename)
        upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'supplier_contracts')
        os.makedirs(upload_dir, exist_ok=True)
        stamped = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_name}"
        file.save(os.path.join(upload_dir, stamped))
        return stamped

    try:
        data = _request_payload()
        supplier_name = data.get('name', '').strip()

        if not supplier_name:
            return jsonify({'success': False, 'error': 'Supplier name is required'}), 400

        phone_val = data.get('phone', '').strip()
        if phone_val and not is_valid_phone(phone_val):
            return jsonify({'success': False, 'error': f'Phone: {PHONE_ERROR}'}), 400

        contract_file = _save_contract_upload()
        notes_value = data.get('notes', '').strip() or ''
        if contract_file:
            notes_value = f"Contract file: {contract_file}\n{notes_value}".strip()
        supplier_payload = dict(data)
        if notes_value:
            supplier_payload['notes'] = notes_value

        existing = _find_ground_handler_supplier_by_name(supplier_name)
        if existing:
            _ensure_inbound_representative(supplier_name)
            db.session.commit()
            _invalidate_supplier_dropdown_cache()
            return jsonify({
                'success': True,
                'supplier_id': existing.id,
                'supplier': {
                    'id': existing.id,
                    'name': existing.name,
                    'phone': existing.phone,
                },
            })

        _rep, new_supplier = _sync_meet_assist_representative_pair(supplier_name, supplier_payload)
        db.session.commit()
        _invalidate_supplier_dropdown_cache()

        return jsonify({
            'success': True,
            'supplier_id': new_supplier.id,
            'supplier': {
                'id': new_supplier.id,
                'name': new_supplier.name,
                'phone': new_supplier.phone
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:hotel_id>/rooms', methods=['POST'])
@csrf.exempt
def api_save_hotel_rooms(hotel_id):
    """Save room distribution details for a hotel"""
    try:
        hotel = InboundHotel.query.get_or_404(hotel_id)
        data = request.get_json()
        rooms_data = data.get('rooms', [])

        # Validate - don't allow completely empty room list if hotel already has rooms
        existing_count = HotelRoom.query.filter_by(hotel_id=hotel_id).count()
        if len(rooms_data) == 0 and existing_count > 0:
            return jsonify({'success': False, 'error': 'Cannot save empty room list. Add at least one room or cancel.'}), 400

        # Update existing rooms or delete and recreate
        # Get existing room IDs for matching
        existing_rooms = {r.id: r for r in HotelRoom.query.filter_by(hotel_id=hotel_id).all()}
        updated_ids = set()

        for room_data in rooms_data:
            room_id = room_data.get('id')
            if room_id and room_id in existing_rooms:
                # Update existing room
                room = existing_rooms[room_id]
                room.room_type = room_data.get('room_type', room.room_type)
                room.room_category = room_data.get('room_category', '')
                room.room_option = room_data.get('room_option', '')
                room.board_basis = room_data.get('board_basis', 'BB')
                room.dietary_requirements = room_data.get('dietary_requirements', '')
                room.adults = room_data.get('adults', 1)
                room.children = room_data.get('children', 0)
                room.guest_names = room_data.get('guest_names', '')
                updated_ids.add(room_id)
            else:
                # Create new room
                room = HotelRoom(
                    hotel_id=hotel_id,
                    room_type=room_data.get('room_type', 'Single'),
                    room_count=1,
                    room_category=room_data.get('room_category', ''),
                    room_option=room_data.get('room_option', ''),
                    board_basis=room_data.get('board_basis', 'BB'),
                    dietary_requirements=room_data.get('dietary_requirements', ''),
                    adults=room_data.get('adults', 1),
                    children=room_data.get('children', 0),
                    guest_names=room_data.get('guest_names', '')
                )
                db.session.add(room)

        # Delete rooms that were removed
        for room_id, room in existing_rooms.items():
            if room_id not in updated_ids:
                db.session.delete(room)

        # Update hotel room counts
        hotel.single_rooms = sum(1 for r in rooms_data if r.get('room_type') == 'Single')
        hotel.double_rooms = sum(1 for r in rooms_data if r.get('room_type') in ['Double', 'Twin', 'King'])
        hotel.triple_rooms = sum(1 for r in rooms_data if r.get('room_type') in ['Triple', 'Suite'])

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'{len(rooms_data)} rooms saved'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/delete-service', methods=['POST'])
@csrf.exempt
def api_delete_service(request_id):
    """Delete a service (hotel, transport, guide, meal)"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        data = request.get_json()
        service_type = data.get('service_type')
        service_id = data.get('service_id')

        if not service_type or not service_id:
            return jsonify({'success': False, 'error': 'Missing service_type or service_id'}), 400

        # Delete based on service type
        if service_type == 'hotel':
            service = InboundHotel.query.filter_by(id=service_id, request_id=request_id).first()
        elif service_type == 'transport':
            service = InboundTransport.query.filter_by(id=service_id, request_id=request_id).first()
            if service:
                _detach_individual_transport_from_batch(service)
        elif service_type == 'guide':
            service = InboundGuide.query.filter_by(id=service_id, request_id=request_id).first()
        elif service_type == 'meal':
            service = InboundMeal.query.filter_by(id=service_id, request_id=request_id).first()
        else:
            return jsonify({'success': False, 'error': f'Unknown service type: {service_type}'}), 400

        if service:
            if service_type == 'guide':
                g = cast(InboundGuide, service)
                if g.source_itinerary_id and g.supplier_id:
                    it_row = ItineraryRow.query.filter_by(
                        id=g.source_itinerary_id, request_id=request_id
                    ).first()
                    if it_row:
                        cur = itinerary_row_guide_supplier_id_list(it_row)
                        if g.supplier_id in cur:
                            it_row.set_itinerary_guide_supplier_id_list(
                                [i for i in cur if i != g.supplier_id]
                            )
            db.session.delete(service)
            db.session.commit()

            # Expire cached data and re-query with eager loading
            db.session.expire_all()
            fresh_request = InboundRequest.query.options(
                selectinload(InboundRequest.inbound_hotels),
                selectinload(InboundRequest.inbound_transports),
                selectinload(InboundRequest.inbound_guides),
                selectinload(InboundRequest.inbound_meals),
                selectinload(InboundRequest.itinerary_rows),
                selectinload(InboundRequest.arrival_batches),
                selectinload(InboundRequest.departure_batches)
            ).get(request_id)

            # Build service lookup maps for template
            hotel_map = {h.source_itinerary_id: h for h in fresh_request.inbound_hotels}
            transport_map = {t.source_itinerary_id: t for t in fresh_request.inbound_transports}
            guide_map = {g.source_itinerary_id: g for g in fresh_request.inbound_guides}
            meal_map = {m.source_itinerary_id: m for m in fresh_request.inbound_meals}

            # Render updated HTML partials for instant DOM update
            sorted_itin = sort_itinerary_rows_with_children(fresh_request.itinerary_rows)
            itinerary_html = render_template(
                'components/itinerary_rows.html',
                rows=sorted_itin,
                request=fresh_request,
                inbound_request=fresh_request,
                guide_suppliers=_guide_suppliers_for_itinerary_ui(),
                itinerary_guide_slot_saved=_itinerary_guide_slot_saved_map(fresh_request),
                view_only=False,
                hotel_map=hotel_map,
                transport_map=transport_map,
                guide_map=guide_map,
                meal_map=meal_map,
                global_hotel=hotel_map.get(None),
                global_transport=transport_map.get(None),
                global_guide=guide_map.get(None),
                global_meal=meal_map.get(None)
            )

            return jsonify({
                'success': True, 
                'message': f'{service_type.capitalize()} deleted',
                'itinerary_html': itinerary_html,
                'deleted_id': service_id,
                'deleted_type': service_type
            })
        else:
            return jsonify({'success': False, 'error': 'Service not found'}), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<int:request_id>/service/<service_type>/<int:record_id>', methods=['GET'])
def api_get_service_record(request_id, service_type, record_id):
    """Get a service record for editing in the summary table"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        record_data = None

        if service_type == 'arrival':
            record = ArrivalBatch.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                nt = getattr(record, 'needs_transport', None)
                record_data = {
                    'id': record.id,
                    'arrival_date': record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else '',
                    'arrival_time': record.arrival_time.strftime('%H:%M') if record.arrival_time else '',
                    'arrival_point': record.arrival_point or '',
                    'flight_number': record.flight_number or '',
                    'pax_count': record.pax_count or 0,
                    'driver_name': record.driver_name or '',
                    'visa_status': record.visa_status or '',
                    'meet_assist': record.meet_assist or False,
                    'representative_name': record.representative_name or '',
                    'status': record.status or 'REQUESTED',
                    'notes': getattr(record, 'notes', '') or '',
                    'supplier_id': record.supplier_id,
                    'needs_transport': True if nt is None else bool(nt),
                    'confirmation_email_filename': record.confirmation_email_filename or '',
                }

        elif service_type == 'departure':
            record = DepartureBatch.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                nt = getattr(record, 'needs_transport', None)
                record_data = {
                    'id': record.id,
                    'departure_date': record.departure_date.strftime('%Y-%m-%d') if record.departure_date else '',
                    'departure_time': record.departure_time.strftime('%H:%M') if record.departure_time else '',
                    'departure_point': record.departure_point or '',
                    'flight_number': record.flight_number or '',
                    'pax_count': record.pax_count or 0,
                    'departure_tax': record.departure_tax or '',
                    'meet_assist': record.meet_assist or False,
                    'representative_name': record.representative_name or '',
                    'status': record.status or 'REQUESTED',
                    'notes': getattr(record, 'notes', '') or '',
                    'supplier_id': record.supplier_id,
                    'needs_transport': True if nt is None else bool(nt),
                }

        elif service_type == 'hotel':
            record = InboundHotel.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                confirmation = ''
                if record.rooms:
                    confirmation = (record.rooms[0].confirmation or '').strip()
                record_data = {
                    'id': record.id,
                    'hotel_name': record.hotel_name or '',
                    'hotel_category': record.hotel_category or '',
                    'check_in_date': record.check_in_date.strftime('%Y-%m-%d') if record.check_in_date else '',
                    'check_out_date': record.check_out_date.strftime('%Y-%m-%d') if record.check_out_date else '',
                    'cut_off_date': record.cut_off_date.strftime('%Y-%m-%d') if record.cut_off_date else '',
                    'nights': record.nights or 0,
                    'status': record.status or 'REQUESTED',
                    'meal_plan': record.meal_plan or 'BB',
                    'single_rooms': record.single_rooms or 0,
                    'double_rooms': record.double_rooms or 0,
                    'triple_rooms': record.triple_rooms or 0,
                    'notes': record.notes or '',
                    'hotel_confirmation_number': confirmation,
                    'confirmation_email_filename': record.confirmation_email_filename or '',
                    'rooming_list_filename': record.rooming_list_filename or ''
                }

        elif service_type == 'transport':
            record = InboundTransport.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                # Use end_date for to_date if it exists, otherwise use date
                to_date = record.end_date if record.end_date else record.date
                pending_fill = _is_individual_transport_from_flight(record) and not _transport_flight_stub_complete(record)
                record_data = {
                    'id': record.id,
                    'service_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'from_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'to_date': to_date.strftime('%Y-%m-%d') if to_date else '',
                    'cut_off_date': record.cut_off_date.strftime('%Y-%m-%d') if record.cut_off_date else '',
                    'vehicle_type': record.vehicle_type or '',
                    'pickup_point': record.pickup_location or '',
                    'drop_off_point': record.dropoff_location or '',
                    'driver_name': record.driver_name or '',
                    'driver_phone': record.driver_phone or '',
                    'license_number': record.license_number or '',
                    'transport_notes': record.note or '',
                    'status': record.status or 'REQUESTED',
                    'supplier_id': record.supplier_id,
                    'cost': float(record.cost) if record.cost is not None else 0,
                    'currency': record.currency or 'JOD',
                    'pending_transport_fill': pending_fill,
                    'source_arrival_batch_id': record.source_arrival_batch_id,
                    'source_departure_batch_id': record.source_departure_batch_id,
                    'confirmation_email_filename': record.confirmation_email_filename or ''
                }

        elif service_type == 'guide':
            record = InboundGuide.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                # Use end_date for to_date if it exists, otherwise use date
                to_date = record.end_date if record.end_date else record.date
                record_data = {
                    'id': record.id,
                    'service_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'from_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'to_date': to_date.strftime('%Y-%m-%d') if to_date else '',
                    'guide_name': record.guide_name or '',
                    'language': record.language or '',
                    'telephone': record.telephone_number or '',
                    'guide_notes': record.additional_comments or '',
                    'supplier_id': record.supplier_id,
                    'status': record.status or 'REQUESTED',
                    'source_itinerary_id': record.source_itinerary_id,
                    'guide_cost': float(record.cost) if record.cost is not None else 0,
                    'guide_cost_currency': record.currency or 'JOD',
                }

        elif service_type == 'meal':
            record = InboundMeal.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                record_data = {
                    'id': record.id,
                    'service_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'from_date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'cut_off_date': record.cut_off_date.strftime('%Y-%m-%d') if record.cut_off_date else '',
                    'meal_type': record.meal_type or '',
                    'meal_status': record.status or 'REQUESTED',
                    'meal_cost': record.total_cost,
                    'meal_cost_currency': record.currency or 'JOD',
                    'restaurant_name': (record.supplier_ref.name if record.supplier_ref else None) or record.restaurant or '',
                    'location': record.location or '',
                    'meal_notes': record.meal_note or '',
                    'pax_count': getattr(record, 'pax_count', None),
                    'supplier_id': record.supplier_id,
                    'confirmation_email_filename': record.confirmation_email_filename or ''
                }

        elif service_type == 'itinerary':
            record = ItineraryRow.query.filter_by(id=record_id, request_id=request_id).first()
            if record:
                record_data = {
                    'id': record.id,
                    'date': record.date.strftime('%Y-%m-%d') if record.date else '',
                    'description': record.description or '',
                    'meal_type': record.meal_type or '',
                    'restaurant_supplier_id': record.restaurant_supplier_id,
                    'restaurant_name': record.restaurant_name or ''
                }

        else:
            return jsonify({'success': False, 'error': f'Unknown service type: {service_type}'}), 400

        if record_data:
            return jsonify({'success': True, 'record': record_data})
        else:
            return jsonify({'success': False, 'error': 'Record not found'}), 404

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<int:request_id>/service/<service_type>/<int:record_id>', methods=['DELETE'])
@csrf.exempt
def api_delete_service_record(request_id, service_type, record_id):
    """Delete a specific service record from the summary table"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        service = None

        if service_type == 'arrival':
            InboundTransport.query.filter_by(
                request_id=request_id, source_arrival_batch_id=record_id
            ).delete(synchronize_session=False)
            service = ArrivalBatch.query.filter_by(id=record_id, request_id=request_id).first()
        elif service_type == 'departure':
            InboundTransport.query.filter_by(
                request_id=request_id, source_departure_batch_id=record_id
            ).delete(synchronize_session=False)
            service = DepartureBatch.query.filter_by(id=record_id, request_id=request_id).first()
        elif service_type == 'hotel':
            service = InboundHotel.query.filter_by(id=record_id, request_id=request_id).first()
            if service:
                # Also delete associated rooms
                HotelRoom.query.filter_by(hotel_id=record_id).delete()
        elif service_type == 'transport':
            service = InboundTransport.query.filter_by(id=record_id, request_id=request_id).first()
            if service:
                _detach_individual_transport_from_batch(service)
        elif service_type == 'guide':
            service = InboundGuide.query.filter_by(id=record_id, request_id=request_id).first()
        elif service_type == 'meal':
            service = InboundMeal.query.filter_by(id=record_id, request_id=request_id).first()
        elif service_type == 'itinerary':
            service = ItineraryRow.query.filter_by(id=record_id, request_id=request_id).first()
        else:
            return jsonify({'success': False, 'error': f'Unknown service type: {service_type}'}), 400

        if service:
            if service_type == 'guide':
                g = cast(InboundGuide, service)
                if g.source_itinerary_id and g.supplier_id:
                    it_row = ItineraryRow.query.filter_by(
                        id=g.source_itinerary_id, request_id=request_id
                    ).first()
                    if it_row:
                        cur = itinerary_row_guide_supplier_id_list(it_row)
                        if g.supplier_id in cur:
                            it_row.set_itinerary_guide_supplier_id_list(
                                [i for i in cur if i != g.supplier_id]
                            )
            db.session.delete(service)
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'{service_type.capitalize()} record deleted successfully',
                'deleted_id': record_id,
                'deleted_type': service_type
            })
        else:
            return jsonify({'success': False, 'error': 'Record not found'}), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<int:request_id>/create-default-itinerary', methods=['POST'])
@csrf.exempt
def api_create_default_itinerary(request_id):
    """Create default itinerary rows for a new request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    # Check if itinerary already exists
    if request_obj.itinerary_rows and len(request_obj.itinerary_rows) > 0:
        return jsonify({'success': True, 'message': 'Itinerary already exists'})

    try:
        data = request.get_json() or {}

        # Get dates from request or use defaults
        from_date_str = data.get('from_date')
        to_date_str = data.get('to_date')

        if from_date_str:
            request_obj.from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date()
        if to_date_str:
            request_obj.to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date()

        # If no dates set, use today + 3 days
        if not request_obj.from_date:
            request_obj.from_date = date.today()
        if not request_obj.to_date:
            request_obj.to_date = request_obj.from_date + timedelta(days=3)

        # Calculate days
        request_obj.calculate_days()

        # Create itinerary rows for each day
        current_date = request_obj.from_date
        day_counter = 1

        while current_date <= request_obj.to_date:
            # Generate description based on day
            if day_counter == 1:
                description = "Arrival Day"
            elif current_date == request_obj.to_date:
                description = "Departure Day"
            else:
                description = f"Day {day_counter}"

            row = ItineraryRow(
                request_id=request_obj.id,
                date=current_date,
                description=description,
                flag_hotel=(day_counter != (request_obj.to_date - request_obj.from_date).days + 1),  # Hotel except last day
                flag_transport=True,
                flag_guide=(day_counter > 1 and day_counter < (request_obj.to_date - request_obj.from_date).days + 1),  # Guide for middle days
                flag_meal=(day_counter > 1 and day_counter < (request_obj.to_date - request_obj.from_date).days + 1),  # Meals for middle days
                flag_airport=(day_counter == 1 or current_date == request_obj.to_date)  # Airport on first and last day
            )
            db.session.add(row)

            current_date += timedelta(days=1)
            day_counter += 1

        db.session.commit()

        return jsonify({'success': True, 'message': 'Default itinerary created'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/update-status', methods=['POST'])
@csrf.exempt

def api_update_status(request_id):
    """Update request status"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    new_status = data.get('status')

    # Accept all workflow statuses used in the sidebar
    valid_statuses = [
        'REQUEST', 'SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 
        'INVOICE', 'PROCESSING', 'COMPLETED',
        STATUS_REQUEST, STATUS_BOOKED, STATUS_IN_PROGRESS, STATUS_CONFIRMED
    ]

    if new_status not in valid_statuses:
        return jsonify({'error': 'Invalid status'}), 400

    request_obj.status = new_status
    db.session.commit()

    return jsonify({'success': True, 'status': new_status})

@inbound_bp.route('/api/<int:request_id>/cancel-request', methods=['POST'])
@csrf.exempt
def api_cancel_request(request_id):
    """Cancel a request - sets status to CANCELLED"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    try:
        request_obj.status = 'CANCELLED'
        db.session.commit()
        return jsonify({'success': True, 'status': 'CANCELLED'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/submit-request', methods=['POST'])
@csrf.exempt
def api_submit_request(request_id):
    """Submit request - transitions REQUEST → CONFIRMED or CONFIRMED → INVOICED"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    target_status = data.get('target_status', '').upper()

    # Map current status to determine valid transitions
    current_status = str(request_obj.status).upper()

    # Determine target status based on current status if not provided
    if not target_status:
        if current_status == 'REQUEST':
            target_status = 'CONFIRMED'
        elif current_status in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS']:
            target_status = 'INVOICED'
        else:
            return jsonify({'success': False, 'message': 'Invalid status for submission'}), 400

    # Validate transition
    if current_status == 'REQUEST' and target_status != 'CONFIRMED':
        return jsonify({'success': False, 'message': 'Request status can only transition to CONFIRMED'}), 400

    if current_status in ['SUPPLIER_CONFIRMED', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS'] and target_status != 'INVOICED':
        return jsonify({'success': False, 'message': 'Confirmed status can only transition to INVOICED'}), 400

    try:
        # Map to actual status values used in database
        if target_status == 'CONFIRMED':
            request_obj.status = 'CONFIRMED'
        elif target_status == 'INVOICED':
            request_obj.status = 'INVOICED'
            request_obj.pending_invoice_queue = False
            request_obj.deleted_reason = None

            # Automatically create invoice record when status changes to INVOICED
            from app.models.invoice import Invoice

            # Check if invoice already exists for this request (prevent duplicates)
            existing_invoice = Invoice.query.filter_by(inbound_request_id=request_obj.id).first()

            if not existing_invoice:
                # Calculate total amount from itinerary rows
                try:
                    total = request_obj.calculate_total()
                except Exception as calc_error:
                    # Fallback to request's total_amount if calculation fails
                    total = request_obj.total_amount or 0.0
                    print(f"Error calculating total from itinerary, using request total_amount: {calc_error}", file=sys.stderr)

                # Generate invoice number
                # Use document_sequence if available, otherwise use request_number
                invoice_number = request_obj.document_sequence or request_obj.request_number

                # Ensure invoice number is unique by appending suffix if needed
                base_invoice_number = invoice_number
                counter = 1
                while Invoice.query.filter_by(invoice_number=invoice_number).first():
                    invoice_number = f"{base_invoice_number}-INV-{counter:03d}"
                    counter += 1

                # Create invoice record
                invoice = Invoice(
                    inbound_request_id=request_obj.id,
                    booking_id=request_obj.booking_id,  # May be None
                    invoice_number=invoice_number,
                    invoice_date=datetime.utcnow(),
                    total_amount=total,
                    notes=request_obj.special_note
                )
                db.session.add(invoice)
        else:
            return jsonify({'success': False, 'message': 'Invalid target status'}), 400

        db.session.commit()

        # Prepare response with redirect URL if status changed to INVOICED
        response_data = {
            'success': True, 
            'status': request_obj.status, 
            'message': f'Request status updated to {request_obj.status}'
        }

        # Add redirect URL for INVOICED status
        if target_status == 'INVOICED':
            response_data['redirect_url'] = url_for('inbound.generate_invoice', request_id=request_obj.id)

        return jsonify(response_data)
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/generate-services', methods=['POST'])
@csrf.exempt

def api_generate_services(request_id):
    """Generate services and create normal booking"""
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        if request_obj.user_id != 1:
            return jsonify({'error': 'Access denied'}), 403

        # Import the necessary models
        from app.models import Booking, ServiceItem, Customer
        from app.models import SERVICE_HOTEL, SERVICE_TRANSPORT, SERVICE_RESTAURANT, SERVICE_GUIDE

        # Create or get booking record
        if request_obj.booking_id:
            booking = Booking.query.get(request_obj.booking_id)
            if not booking:
                booking = None
        else:
            booking = None

        if not booking:
            # Use the customer from the inbound request if available
            # For existing requests without customer_id, try to find by contact name
            customer_id = getattr(request_obj, 'customer_id', None)
            if customer_id:
                customer = Customer.query.get(customer_id)
            else:
                # Fallback: find or create customer by contact name
                customer = Customer.query.filter_by(first_name=request_obj.contact_name).first()
                if not customer:
                    customer = Customer()
                    customer.first_name = request_obj.contact_name
                    customer.last_name = ""
                    customer.phone = "TBD"
                    customer.email = "tbd@example.com"
                    customer.nationality = request_obj.nationality
                    db.session.add(customer)
                    db.session.flush()

            # Create new booking
            booking = Booking()
            booking.reference_number = request_obj.request_number
            booking.user_id = request_obj.user_id
            booking.customer_id = customer.id
            booking.status = request_obj.status
            booking.total_amount = request_obj.total_amount
            db.session.add(booking)
            db.session.flush()

            # Link booking to inbound request and update status to BOOKED
            request_obj.booking_id = booking.id
            request_obj.status = 'BOOKED'

        # Clear existing service items
        ServiceItem.query.filter_by(booking_id=booking.id).delete()

        services_created = 0

        # Generate ServiceItem records based on itinerary flags
        for row in request_obj.itinerary_rows:
            row_cost = row.calculate_row_cost(request_obj.pax)

            # Hotel service - include room distribution data
            if row.flag_hotel:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_HOTEL
                service_item.start_date = row.date
                service_item.end_date = row.date + timedelta(days=1)
                service_item.description = f"Hotel accommodation - {row.description}"
                service_item.amount = row_cost
                service_item.status = STATUS_REQUEST
                # Store room distribution data in the description for now
                room_summary = f"S:{row.hotel_single_rooms or 0} D:{row.hotel_double_rooms or 0} T:{row.hotel_triple_rooms or 0} O:{row.hotel_other_rooms or 0}"
                service_item.description = f"Hotel accommodation - {row.description} | Rooms: {room_summary}"
                db.session.add(service_item)
                services_created += 1

            # Transport service
            if row.flag_transport:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_TRANSPORT
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Transport service - {row.description}"
                service_item.amount = row_cost
                service_item.status = STATUS_REQUEST
                db.session.add(service_item)
                services_created += 1

            # Restaurant/Meal service
            if row.flag_meal:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_RESTAURANT
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Restaurant meal - {row.description}"
                service_item.amount = row_cost
                service_item.status = STATUS_REQUEST
                db.session.add(service_item)
                services_created += 1

            # Guide service
            if row.flag_guide:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_GUIDE
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Tour guide service - {row.description}"
                service_item.amount = row_cost
                service_item.status = STATUS_REQUEST
                db.session.add(service_item)
                services_created += 1

        # Update booking total
        booking.calculate_total()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': f"Generated {services_created} services",
            'booking_id': booking.id,
            'redirect_url': url_for('booking.details', booking_id=booking.id)
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# Removed view_services route - no longer needed since we redirect to normal booking page

@inbound_bp.route('/<int:request_id>/invoice')
@login_required
def generate_invoice(request_id):
    """Generate invoice for the request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    if request_obj.status == STATUS_REQUEST:
        abort(400, 'Cannot generate invoice for request status')

    try:
        saved_admin = json.loads(request_obj.admin_invoice_data) if request_obj.admin_invoice_data else None
    except (TypeError, ValueError):
        saved_admin = None
    return render_template('inbound/invoice.html', request=request_obj, saved_admin_invoice=saved_admin)

@inbound_bp.route('/<int:request_id>/customer-invoice')
@login_required
def customer_invoice(request_id):
    """Generate customer-facing invoice (Windows of Jordan format)"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        abort(403)
    if request_obj.status == STATUS_REQUEST:
        abort(400, 'Cannot generate invoice for request status')
    # Build rooms summary: e.g. "1 Dbl & 1 Single on BB"
    rooms_parts = []
    board = "BB"
    for h in request_obj.inbound_hotels:
        if h.meal_plan:
            board = h.meal_plan
        for r in h.rooms:
            if r.room_type == "SINGLE" and r.room_count:
                rooms_parts.append(f"{r.room_count} Single")
            elif r.room_type == "DOUBLE" and r.room_count:
                rooms_parts.append(f"{r.room_count} Dbl")
            elif r.room_type == "TRIPLE" and r.room_count:
                rooms_parts.append(f"{r.room_count} Triple")
        if not h.rooms and (h.single_rooms or h.double_rooms or h.triple_rooms):
            if h.single_rooms:
                rooms_parts.append(f"{h.single_rooms} Single")
            if h.double_rooms:
                rooms_parts.append(f"{h.double_rooms} Dbl")
            if h.triple_rooms:
                rooms_parts.append(f"{h.triple_rooms} Triple")
    rooms_display = " & ".join(rooms_parts) + f" on {board}" if rooms_parts else "TBA"
    # Tour ref display: first itinerary description if short, else "X Days" or document_sequence
    tour_ref_display = None
    if request_obj.itinerary_rows:
        first_desc = request_obj.itinerary_rows[0].description or ''
        if first_desc and len(first_desc) <= 60:
            tour_ref_display = first_desc
    if not tour_ref_display and request_obj.no_of_days:
        tour_ref_display = f"{request_obj.no_of_days} Days"
    try:
        saved_customer = json.loads(request_obj.customer_invoice_data) if request_obj.customer_invoice_data else None
    except (TypeError, ValueError):
        saved_customer = None
    return render_template('inbound/customer_invoice.html', request=request_obj, rooms_display=rooms_display, tour_ref_display=tour_ref_display, saved_customer_invoice=saved_customer)

@inbound_bp.route('/<int:request_id>/save-admin-invoice', methods=['POST'])
@login_required
@csrf.exempt
def save_admin_invoice(request_id):
    """Save editable admin invoice content as JSON"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        abort(403)
    data = request.get_json(silent=True) or {}
    request_obj.admin_invoice_data = json.dumps(data) if data else None
    db.session.commit()
    return jsonify({'ok': True})

@inbound_bp.route('/<int:request_id>/save-customer-invoice', methods=['POST'])
@login_required
@csrf.exempt
def save_customer_invoice(request_id):
    """Save editable customer invoice content as JSON"""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        abort(403)
    data = request.get_json(silent=True) or {}
    request_obj.customer_invoice_data = json.dumps(data) if data else None
    db.session.commit()
    return jsonify({'ok': True})

@inbound_bp.route('/<int:id>/hotel-voucher')
@login_required
def generate_hotel_voucher(id):
    """Generate hotel services voucher for the request (print layout)"""
    request_obj = InboundRequest.query.get_or_404(id)
    if request_obj.user_id != 1:
        abort(403)
    # Show only when request status is confirmed (CONFIRMED, SUPPLIER_CONFIRMED, QUOTED, etc.)
    confirmed_statuses = ['CONFIRMED', 'SUPPLIER_CONFIRMED', 'QUOTED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS', 'INVOICE', 'COMPLETED', 'INVOICED']
    if request_obj.status not in confirmed_statuses:
        abort(400, 'Hotel voucher is only available when request is confirmed')
    return render_template('inbound/hotel_voucher.html', request=request_obj)

@inbound_bp.route('/<int:id>/restaurant-voucher')
@inbound_bp.route('/<int:id>/restaurant-voucher/<int:meal_id>')
@login_required
def generate_restaurant_voucher(id, meal_id=None):
    """Generate restaurant services voucher for the request (print layout)"""
    request_obj = InboundRequest.query.get_or_404(id)
    if request_obj.user_id != 1:
        abort(403)
    confirmed_statuses = ['CONFIRMED', 'SUPPLIER_CONFIRMED', 'QUOTED', 'PROCESSING', 'BOOKED', 'IN_PROGRESS', 'INVOICE', 'COMPLETED', 'INVOICED']
    if request_obj.status not in confirmed_statuses:
        abort(400, 'Restaurant voucher is only available when request is confirmed')

    target_meal = None
    if meal_id:
        target_meal = InboundMeal.query.filter_by(id=meal_id, request_id=id).first()

    return render_template('inbound/restaurant_voucher.html', request=request_obj, target_meal=target_meal)

def _record_document(request_id: int, doc_type: str, filename: str,
                     filepath: str, file_bytes: bytes, user_id: int = 1,
                     mime_type: str = 'application/pdf') -> None:
    """Record a generated document in inbound_document table (non-fatal)."""
    try:
        doc = InboundDocument(
            request_id=request_id,
            document_type=doc_type,
            filename=filename,
            original_filename=filename,
            filepath=filepath,
            file_size=len(file_bytes),
            mime_type=mime_type,
            uploaded_by=user_id,
        )
        db.session.add(doc)
        db.session.commit()
    except Exception as e:
        current_app.logger.warning(f'Document record failed (non-fatal): {e}')
        try:
            db.session.rollback()
        except Exception:
            pass


@inbound_bp.route('/<int:request_id>/voucher')
@login_required
def generate_voucher(request_id):
    """Generate visual timeline voucher for the request"""
    from datetime import datetime
    from weasyprint import HTML
    from flask import make_response
    import io

    request_obj = InboundRequest.query.get_or_404(request_id)

    # Temporarily disabled user validation for testing
    # if request_obj.user_id != 1:
    #     abort(403)

    # Allow voucher generation for testing/preview
    # if request_obj.status in [STATUS_REQUEST, STATUS_BOOKED]:
    #     abort(400, 'Cannot generate voucher until confirmed')

    # Get layout preference from query parameter (default to vertical)
    layout = request.args.get('layout', 'vertical')

    # Choose template based on layout
    if layout == 'horizontal':
        template = 'inbound/voucher_timeline_horizontal.html'
    else:
        template = 'inbound/voucher_timeline.html'

    # Render the timeline template
    html = render_template(template, 
                          request=request_obj,
                          now=datetime.now())

    # Try to generate PDF using WeasyPrint
    try:
        # Create PDF from HTML
        pdf_buffer = io.BytesIO()
        HTML(string=html).write_pdf(pdf_buffer)
        pdf = pdf_buffer.getvalue()
        pdf_buffer.close()

        # Record generated voucher in inbound_document (non-fatal)
        try:
            original_name = f'tour_itinerary_{request_obj.request_number}.pdf'
            rel_path, stored_name = document_storage.save(
                pdf, request_id=request_id, doc_type='VOUCHER',
                original_filename=original_name
            )
            _record_document(request_id, 'VOUCHER', stored_name, rel_path, pdf)
        except Exception as _rec_err:
            current_app.logger.warning(f'Voucher document record skipped: {_rec_err}')

        response = make_response(pdf)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename=tour_itinerary_{request_obj.request_number}.pdf'

        return response
    except Exception as e:
        # If PDF generation fails, return HTML version for debugging
        print(f"PDF generation failed: {e}")
        import traceback
        traceback.print_exc()

        # Return HTML with error info for debugging
        error_html = f"<h1>Voucher Generation Error</h1><p>Error: {str(e)}</p><hr>{html}"
        return error_html

@inbound_bp.route('/api/<int:request_id>/create-booking', methods=['POST'])
@csrf.exempt

def api_create_booking(request_id):
    """Create a booking from an inbound request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    # Check if booking already exists by looking for existing services
    if request_obj.booking_id:
        booking = Booking.query.get(request_obj.booking_id)
        if booking:
            return jsonify({
                'success': True,
                'message': 'Booking already exists',
                'booking_url': url_for('booking.details', booking_id=booking.id)
            })

    try:
        # Use the existing generate services function logic
        result = api_generate_services(request_id)
        result_data = result.get_json()

        if result_data.get('success'):
            return jsonify({
                'success': True,
                'message': 'Booking created successfully',
                'booking_url': result_data.get('redirect_url')
            })
        else:
            return jsonify({
                'success': False,
                'message': result_data.get('error', 'Failed to create booking')
            }), 500

    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error creating booking: {str(e)}'
        }), 500

@inbound_bp.route('/api/<int:request_id>/save-hotels', methods=['POST'])
@csrf.exempt  
def api_save_hotels(request_id):
    """Save hotel configuration data including rooms"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    hotels_data = data.get('hotels', [])

    from app.models.inbound import InboundHotel, HotelRoom

    try:
        # Delete existing hotels and their rooms for this request
        InboundHotel.query.filter_by(request_id=request_id).delete()

        # Create new hotel records
        for hotel_data in hotels_data:
            hotel = InboundHotel(
                request_id=request_id,
                hotel_name=hotel_data.get('hotel_name', ''),
                hotel_category=hotel_data.get('hotel_category', ''),
                status='REQUEST'
            )

            # Get check-in/out from hotel-level date inputs (NOT from rooms)
            if hotel_data.get('check_in_date'):
                hotel.check_in_date = datetime.strptime(hotel_data['check_in_date'], '%Y-%m-%d').date()
            if hotel_data.get('check_out_date'):
                hotel.check_out_date = datetime.strptime(hotel_data['check_out_date'], '%Y-%m-%d').date()

            # If no hotel-level dates provided, auto-inherit from request dates
            if not hotel.check_in_date:
                hotel.check_in_date = request_obj.from_date
            if not hotel.check_out_date:
                hotel.check_out_date = request_obj.to_date

            # Calculate nights based on hotel-level dates
            if hotel.check_in_date and hotel.check_out_date:
                hotel.nights = (hotel.check_out_date - hotel.check_in_date).days

            rooms = hotel_data.get('rooms', [])

            db.session.add(hotel)
            db.session.flush()  # Get the hotel ID before creating rooms

            # Create HotelRoom records for each room with detailed data
            for room_data in rooms:
                # Determine room type
                room_category = room_data.get('room_category', 'Single Room')
                if 'Single' in room_category:
                    room_type = 'SINGLE'
                elif 'Double' in room_category:
                    room_type = 'DOUBLE'
                elif 'Triple' in room_category:
                    room_type = 'TRIPLE'
                else:
                    room_type = 'OTHER'

                # Store additional room details as JSON in notes field
                # Note: check_in/check_out are at HOTEL level only, rooms inherit via @property
                room_details = {
                    'hotel_room_option': room_data.get('hotel_room_option', ''),
                    'board_basis': room_data.get('board_basis', 'BB'),
                    'dietary_requirements': room_data.get('dietary_requirements', ''),
                    'lead_passenger': room_data.get('lead_passenger', ''),
                    'adults': room_data.get('adults', 1),
                    'children': room_data.get('children', 0)
                }

                hotel_room = HotelRoom(  # type: ignore[call-arg]
                    hotel_id=hotel.id,
                    room_type=room_type,
                    room_count=1,
                    status='REQUEST',
                    notes=json.dumps(room_details)  # Store detailed data as JSON
                )
                db.session.add(hotel_room)

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Saved {len(hotels_data)} hotel(s) successfully'
        })

    except Exception as e:
        db.session.rollback()
        print(f"Error saving hotels: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/arrivals', methods=['GET'])
@csrf.exempt
def api_get_arrivals(request_id):
    """Get all arrival batches for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import ArrivalBatch

    batches = ArrivalBatch.query.filter_by(request_id=request_id).order_by(ArrivalBatch.arrival_date).all()

    batches_data = []
    for batch in batches:
        batches_data.append({
            'id': batch.id,
            'batch_name': batch.batch_name or '',
            'arrival_date': batch.arrival_date.strftime('%Y-%m-%d') if batch.arrival_date else '',
            'arrival_point': batch.arrival_point or '',
            'arrival_time': batch.arrival_time.strftime('%H:%M') if batch.arrival_time else '',
            'driver_name': batch.driver_name or '',
            'vehicle_details': batch.vehicle_details or '',
            'pax_count': batch.pax_count or 0,
            'flight_number': batch.flight_number or '',
            'visa_status': getattr(batch, 'visa_status', 'NOT_INCLUDED'),
            'meet_assist': getattr(batch, 'meet_assist', False),
            'representative_name': getattr(batch, 'representative_name', '')
        })

    return jsonify({'success': True, 'batches': batches_data})

@inbound_bp.route('/api/<int:request_id>/departures', methods=['GET'])
@csrf.exempt
def api_get_departures(request_id):
    """Get all departure batches for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import DepartureBatch

    batches = DepartureBatch.query.filter_by(request_id=request_id).order_by(DepartureBatch.departure_date).all()

    batches_data = []
    for batch in batches:
        batches_data.append({
            'id': batch.id,
            'batch_name': batch.batch_name or '',
            'departure_date': batch.departure_date.strftime('%Y-%m-%d') if batch.departure_date else '',
            'departure_point': batch.departure_point or '',
            'departure_time': batch.departure_time.strftime('%H:%M') if batch.departure_time else '',
            'driver_name': batch.driver_name or '',
            'vehicle_details': batch.vehicle_details or '',
            'pax_count': batch.pax_count or 0,
            'flight_number': batch.flight_number or '',
            'meet_greet': batch.meet_greet if hasattr(batch, 'meet_greet') else False,
            'meet_assist': getattr(batch, 'meet_assist', False),
            'representative_name': getattr(batch, 'representative_name', ''),
            'departure_tax': getattr(batch, 'departure_tax', 'NOT_INCLUDED'),
            'notes': getattr(batch, 'notes', '')
        })

    return jsonify({'success': True, 'batches': batches_data})

@inbound_bp.route('/api/<int:request_id>/get-flights-data', methods=['GET'])
@csrf.exempt
def api_get_flights_data(request_id):
    """Get combined arrivals and departures for summary table"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import ArrivalBatch, DepartureBatch

    # Expire any cached data to ensure fresh query
    db.session.expire_all()
    
    # Query fresh data from database
    arrivals = ArrivalBatch.query.filter_by(request_id=request_id).order_by(ArrivalBatch.arrival_date).all()
    departures = DepartureBatch.query.filter_by(request_id=request_id).order_by(DepartureBatch.departure_date).all()

    arrivals_data = []
    for arr in arrivals:
        arrivals_data.append({
            'id': arr.id,
            'arrival_date': arr.arrival_date.strftime('%Y-%m-%d') if arr.arrival_date else '',
            'arrival_time': arr.arrival_time.strftime('%H:%M') if arr.arrival_time else '',
            'arrival_point': arr.arrival_point or '',
            'flight_number': arr.flight_number or '',
            'pax_count': arr.pax_count or 0,
            'driver_name': arr.driver_name or '',
            'meet_assist': bool(getattr(arr, 'meet_assist', False)),
            'representative_name': getattr(arr, 'representative_name', '') or '',
            'notes': arr.notes or '',
            'visa_status': getattr(arr, 'visa_status', '')
        })
        
        # Debug logging for each arrival
        print(f"[API GET FLIGHTS] Arrival ID {arr.id}: arrival_point={repr(arr.arrival_point)}, notes={repr(arr.notes)}, notes_in_response={repr(arr.notes or '')}")

    departures_data = []
    for dep in departures:
        departures_data.append({
            'id': dep.id,
            'departure_date': dep.departure_date.strftime('%Y-%m-%d') if dep.departure_date else '',
            'departure_time': dep.departure_time.strftime('%H:%M') if dep.departure_time else '',
            'departure_point': dep.departure_point or '',
            'flight_number': dep.flight_number or '',
            'pax_count': dep.pax_count or 0,
            'driver_name': dep.driver_name or '',
            'meet_assist': bool(getattr(dep, 'meet_assist', False)),
            'representative_name': getattr(dep, 'representative_name', '') or '',
            'notes': getattr(dep, 'notes', '') or ''
        })

    return jsonify({
        'success': True,
        'arrivals': arrivals_data,
        'departures': departures_data
    })


@inbound_bp.route('/api/<int:request_id>/transport-flight-chips', methods=['GET'])
@csrf.exempt
def api_transport_flight_chips(request_id):
    """Chips for Transport tab: one per flight-linked InboundTransport stub (arrival/departure)."""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403
    try:
        _reconcile_flight_linked_transports(request_id)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('transport-flight-chips reconcile: %s', exc)
    from sqlalchemy import or_
    rows = InboundTransport.query.filter(
        InboundTransport.request_id == request_id,
        or_(
            InboundTransport.source_arrival_batch_id.isnot(None),
            InboundTransport.source_departure_batch_id.isnot(None),
        ),
    ).order_by(InboundTransport.date, InboundTransport.id).all()
    chips = []
    for t in rows:
        d = t.date
        tm = t.pickup_time
        pax = t.pax or 0
        date_s = d.strftime('%m/%d/%Y') if d else '-'
        time_s = tm.strftime('%H:%M') if tm else '--:--'
        kind = 'arrival' if t.source_arrival_batch_id else 'departure'
        src_id = t.source_arrival_batch_id or t.source_departure_batch_id
        chips.append({
            'transport_id': t.id,
            'kind': kind,
            'source_batch_id': src_id,
            'label': f'{date_s} – {time_s} – {pax} Pax',
            'complete': _transport_flight_stub_complete(t),
            'pending_fill': not _transport_flight_stub_complete(t),
        })
    return jsonify({'success': True, 'chips': chips})


@inbound_bp.route('/api/<int:request_id>/transport-summary-html', methods=['GET'])
@csrf.exempt
def api_transport_summary_html(request_id):
    """Refresh Trip Summary transport tbody HTML."""
    request_obj = InboundRequest.query.get_or_404(request_id)
    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403
    try:
        _reconcile_flight_linked_transports(request_id)
        db.session.commit()
    except Exception as exc:
        db.session.rollback()
        current_app.logger.exception('transport-summary-html reconcile: %s', exc)
    transports_list = InboundTransport.query.filter_by(request_id=request_id).order_by(InboundTransport.date).all()
    html = render_template(
        'components/transport_summary_entries.html',
        transports=_trip_summary_transports(transports_list),
        view_only=False
    )
    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@inbound_bp.route('/api/<int:request_id>/arrivals/<int:arrival_id>', methods=['DELETE'])
@csrf.exempt
def api_delete_arrival(request_id, arrival_id):
    """Delete a specific arrival batch"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import ArrivalBatch, ItineraryRow

    try:
        # Find and delete the specific arrival
        arrival = ArrivalBatch.query.filter_by(id=arrival_id, request_id=request_id).first()
        if arrival:
            InboundTransport.query.filter_by(
                request_id=request_id, source_arrival_batch_id=arrival_id
            ).delete(synchronize_session=False)
            db.session.delete(arrival)
            db.session.commit()

            # Re-calculate flags for remaining arrivals
            ItineraryRow.query.filter_by(request_id=request_id).update({'flag_airport': False})

            all_arrivals = ArrivalBatch.query.filter_by(request_id=request_id).all()
            for arr in all_arrivals:
                if arr.arrival_date:
                    arrival_rows = ItineraryRow.query.filter_by(
                        request_id=request_id,
                        date=arr.arrival_date
                    ).all()
                    for row in arrival_rows:
                        row.flag_airport = True

            db.session.commit()

            return jsonify({'success': True, 'message': 'Arrival deleted successfully'})
        else:
            return jsonify({'error': 'Arrival not found'}), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/departures/<int:departure_id>', methods=['DELETE'])
@csrf.exempt
def api_delete_departure(request_id, departure_id):
    """Delete a specific departure batch"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import DepartureBatch, ItineraryRow

    try:
        # Find and delete the specific departure
        departure = DepartureBatch.query.filter_by(id=departure_id, request_id=request_id).first()
        if departure:
            InboundTransport.query.filter_by(
                request_id=request_id, source_departure_batch_id=departure_id
            ).delete(synchronize_session=False)
            db.session.delete(departure)
            db.session.commit()

            # Re-calculate flags for remaining departures
            ItineraryRow.query.filter_by(request_id=request_id).update({'flag_drive': False})

            all_departures = DepartureBatch.query.filter_by(request_id=request_id).all()
            for dep in all_departures:
                if dep.departure_date:
                    departure_rows = ItineraryRow.query.filter_by(
                        request_id=request_id,
                        date=dep.departure_date
                    ).all()
                    for row in departure_rows:
                        row.flag_drive = True

            db.session.commit()

            return jsonify({'success': True, 'message': 'Departure deleted successfully'})
        else:
            return jsonify({'error': 'Departure not found'}), 404

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/departures', methods=['POST'])
@csrf.exempt
def api_save_departures(request_id):
    """Save departure batches for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    batches_data = data.get('batches', [])

    from app.models.inbound import DepartureBatch, ItineraryRow

    try:
        # Don't delete existing batches - just add new ones
        # This allows multiple departures to be saved

        # Create new batches
        for batch_data in batches_data:
            # Parse date
            departure_date = None
            if batch_data.get('departure_date'):
                try:
                    departure_date = datetime.strptime(batch_data['departure_date'], '%Y-%m-%d').date()
                except:
                    continue

            if not departure_date:
                continue

            # Parse time
            departure_time = None
            if batch_data.get('departure_time'):
                try:
                    departure_time = datetime.strptime(batch_data['departure_time'], '%H:%M').time()
                except:
                    pass

            # Parse pax count
            pax_count = 0
            if batch_data.get('pax_count'):
                try:
                    pax_count = int(batch_data['pax_count'])
                except:
                    pax_count = 0

            # Parse meet_greet boolean
            meet_greet = False
            if batch_data.get('meet_greet'):
                meet_greet = batch_data['meet_greet'] in ['true', 'True', True, 1, '1']

            # Parse meet_assist boolean
            meet_assist = False
            if batch_data.get('meet_assist'):
                meet_assist = batch_data['meet_assist'] in ['true', 'True', True, 1, '1']

            # Handle supplier_id
            supplier_id = batch_data.get('supplier_id')
            supplier_id = int(supplier_id) if supplier_id else None

            batch = DepartureBatch(  # type: ignore[call-arg]
                request_id=request_id,
                batch_name=batch_data.get('batch_name') or None,
                departure_date=departure_date,
                departure_point=batch_data.get('departure_point') or None,
                departure_time=departure_time,
                driver_name=batch_data.get('driver_name') or None,
                vehicle_details=batch_data.get('vehicle_details') or None,
                pax_count=pax_count,
                flight_number=batch_data.get('flight_number') or None,
                meet_greet=meet_greet,
                meet_assist=meet_assist,
                representative_name=batch_data.get('representative_name') or None,
                departure_tax=batch_data.get('departure_tax', 'NOT_INCLUDED'),
                supplier_id=supplier_id
            )
            db.session.add(batch)

        db.session.commit()

        # Auto-flag itinerary rows with flag_drive for ALL departure dates
        # First, clear all drive flags for this request
        ItineraryRow.query.filter_by(request_id=request_id).update({'flag_drive': False})

        # Then, set flags for ALL saved departures (not just current batch)
        all_departures = DepartureBatch.query.filter_by(request_id=request_id).all()
        for departure in all_departures:
            if departure.departure_date:
                departure_rows = ItineraryRow.query.filter_by(
                    request_id=request_id,
                    date=departure.departure_date
                ).all()
                for row in departure_rows:
                    row.flag_drive = True

        db.session.commit()

        return jsonify({'success': True, 'message': 'Departures saved successfully'})

    except Exception as e:
        db.session.rollback()
        print(f"Error saving departures: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/arrivals', methods=['POST'])
@csrf.exempt
def api_save_arrivals(request_id):
    """Save arrival batches for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    batches_data = data.get('batches', [])

    from app.models.inbound import ArrivalBatch, ItineraryRow

    try:
        # Don't delete existing batches - just add new ones
        # This allows multiple arrivals to be saved

        # Create new batches
        for batch_data in batches_data:
            # Parse date
            arrival_date = None
            if batch_data.get('arrival_date'):
                try:
                    arrival_date = datetime.strptime(batch_data['arrival_date'], '%Y-%m-%d').date()
                except:
                    continue

            if not arrival_date:
                continue

            # Parse time
            arrival_time = None
            if batch_data.get('arrival_time'):
                try:
                    arrival_time = datetime.strptime(batch_data['arrival_time'], '%H:%M').time()
                except:
                    pass

            # Parse pax count
            pax_count = 0
            if batch_data.get('pax_count'):
                try:
                    pax_count = int(batch_data['pax_count'])
                except:
                    pax_count = 0

            # Parse meet_assist boolean
            meet_assist = False
            if batch_data.get('meet_assist'):
                meet_assist = batch_data['meet_assist'] in ['true', 'True', True, 1, '1', 'on']

            # Handle supplier_id
            supplier_id = batch_data.get('supplier_id')
            supplier_id = int(supplier_id) if supplier_id else None

            batch = ArrivalBatch(  # type: ignore[call-arg]
                request_id=request_id,
                batch_name=batch_data.get('batch_name') or None,
                arrival_date=arrival_date,
                arrival_point=batch_data.get('arrival_point') or None,
                arrival_time=arrival_time,
                driver_name=batch_data.get('driver_name') or None,
                vehicle_details=batch_data.get('vehicle_details') or None,
                pax_count=pax_count,
                flight_number=batch_data.get('flight_number') or None,
                visa_status=batch_data.get('visa_status', 'NOT_INCLUDED'),
                meet_assist=meet_assist,
                representative_name=batch_data.get('representative_name') or None,
                supplier_id=supplier_id
            )
            db.session.add(batch)

        db.session.commit()

        # Auto-flag itinerary rows with flag_airport for ALL arrival dates
        # First, clear all airport flags for this request
        ItineraryRow.query.filter_by(request_id=request_id).update({'flag_airport': False})

        # Then, set flags for ALL saved arrivals (not just current batch)
        all_arrivals = ArrivalBatch.query.filter_by(request_id=request_id).all()
        for arrival in all_arrivals:
            if arrival.arrival_date:
                arrival_rows = ItineraryRow.query.filter_by(
                    request_id=request_id,
                    date=arrival.arrival_date
                ).all()
                for row in arrival_rows:
                    row.flag_airport = True

        db.session.commit()

        return jsonify({'success': True, 'message': 'Arrivals saved successfully'})

    except Exception as e:
        db.session.rollback()
        print(f"Error saving arrivals: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/arrival-departures', methods=['GET'])
def api_get_arrival_departures(request_id):
    """Get all arrival/departure records for a request (combined model)"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    from app.models.inbound import ArrivalDeparture

    records = ArrivalDeparture.query.filter_by(request_id=request_id).order_by(ArrivalDeparture.id).all()

    records_list = []
    for record in records:
        records_list.append({
            'id': record.id,
            'batch_name': record.batch_name,
            'pax_count': record.pax_count,
            'arrival_date': record.arrival_date.strftime('%Y-%m-%d') if record.arrival_date else None,
            'arrival_time': record.arrival_time.strftime('%H:%M') if record.arrival_time else None,
            'arrival_point': record.arrival_point,
            'visa_status': record.visa_status,
            'arrival_driver_name': record.arrival_driver_name,
            'flight_number': record.flight_number,
            'departure_date': record.departure_date.strftime('%Y-%m-%d') if record.departure_date else None,
            'departure_time': record.departure_time.strftime('%H:%M') if record.departure_time else None,
            'departure_point': record.departure_point,
            'meet_assist': record.meet_assist,
            'representative_name': record.representative_name,
            'departure_tax': record.departure_tax
        })

    return jsonify({'success': True, 'records': records_list})

@inbound_bp.route('/api/<int:request_id>/arrival-departures', methods=['POST'])
@csrf.exempt
def api_save_arrival_departures(request_id):
    """Save arrival/departure records for a request (combined model)"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    data = request.get_json()
    records_data = data.get('records', [])

    from app.models.inbound import ArrivalDeparture

    try:
        for record_data in records_data:
            # Parse dates
            arrival_date = None
            if record_data.get('arrival_date'):
                try:
                    arrival_date = datetime.strptime(record_data['arrival_date'], '%Y-%m-%d').date()
                except:
                    pass

            departure_date = None
            if record_data.get('departure_date'):
                try:
                    departure_date = datetime.strptime(record_data['departure_date'], '%Y-%m-%d').date()
                except:
                    pass

            # Parse times
            arrival_time = None
            if record_data.get('arrival_time'):
                try:
                    arrival_time = datetime.strptime(record_data['arrival_time'], '%H:%M').time()
                except:
                    pass

            departure_time = None
            if record_data.get('departure_time'):
                try:
                    departure_time = datetime.strptime(record_data['departure_time'], '%H:%M').time()
                except:
                    pass

            # Parse pax count
            pax_count = 1
            if record_data.get('pax_count'):
                try:
                    pax_count = int(record_data['pax_count'])
                except:
                    pax_count = 1

            # Parse meet_assist boolean
            meet_assist = False
            if record_data.get('meet_assist'):
                meet_assist = record_data['meet_assist'] in ['true', 'True', True, 1, '1', 'on']

            record = ArrivalDeparture(
                request_id=request_id,
                batch_name=record_data.get('batch_name') or None,
                pax_count=pax_count,
                arrival_date=arrival_date,
                arrival_time=arrival_time,
                arrival_point=record_data.get('arrival_point') or None,
                visa_status=record_data.get('visa_status', 'NOT_INCLUDED'),
                arrival_driver_name=record_data.get('arrival_driver_name') or None,
                flight_number=record_data.get('flight_number') or None,
                departure_date=departure_date,
                departure_time=departure_time,
                departure_point=record_data.get('departure_point') or None,
                meet_assist=meet_assist,
                representative_name=record_data.get('representative_name') or None,
                departure_tax=record_data.get('departure_tax', 'NOT_INCLUDED')
            )
            db.session.add(record)

        db.session.commit()
        return jsonify({'success': True, 'message': 'Arrival/Departure records saved successfully'})

    except Exception as e:
        db.session.rollback()
        print(f"Error saving arrival/departures: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/arrival-departures/<int:record_id>', methods=['DELETE'])
@csrf.exempt
def api_delete_arrival_departure(request_id, record_id):
    """Delete an arrival/departure record"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models.inbound import ArrivalDeparture

    record = ArrivalDeparture.query.filter_by(id=record_id, request_id=request_id).first()

    if not record:
        return jsonify({'error': 'Record not found'}), 404

    try:
        db.session.delete(record)
        db.session.commit()
        return jsonify({'success': True, 'message': 'Record deleted successfully'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/generate-quote', methods=['POST'])
@csrf.exempt

def api_generate_quote(request_id):
    """Generate a quote from an inbound request (creates booking with QUOTED status)"""
    # Import the necessary models
    from app.models import Booking, ServiceItem, Customer
    from app.models import SERVICE_HOTEL, SERVICE_TRANSPORT, SERVICE_RESTAURANT, SERVICE_GUIDE

    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    # Check if quote already exists
    if request_obj.booking_id:
        booking = Booking.query.get(request_obj.booking_id)
        if booking:
            return jsonify({
                'success': True,
                'message': 'Quote already exists',
                'booking_id': booking.id
            })

    try:

        # Create or get customer
        customer_id = getattr(request_obj, 'customer_id', None)
        if customer_id:
            customer = Customer.query.get(customer_id)
        else:
            # Fallback: find or create customer by contact name
            customer = Customer.query.filter_by(first_name=request_obj.contact_name).first()
            if not customer:
                customer = Customer()
                customer.first_name = request_obj.contact_name
                customer.last_name = ""
                customer.phone = "TBD"
                customer.email = "tbd@example.com"
                customer.nationality = request_obj.nationality
                db.session.add(customer)
                db.session.flush()

        # Create new booking with QUOTED status
        booking = Booking()
        booking.reference_number = request_obj.request_number
        booking.user_id = request_obj.user_id
        booking.customer_id = customer.id
        booking.status = 'QUOTED'  # Set as quoted instead of booked
        booking.total_amount = request_obj.total_amount
        db.session.add(booking)
        db.session.flush()

        # Link booking to inbound request and update status to QUOTED
        request_obj.booking_id = booking.id
        request_obj.status = 'QUOTED'

        # Clear existing service items and create new ones
        ServiceItem.query.filter_by(booking_id=booking.id).delete()

        services_created = 0

        # Create service items based on itinerary flags
        for row in request_obj.itinerary_rows:
            if row.flag_hotel:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_HOTEL
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Hotel service for {row.city}"
                service_item.amount = row.hotel_cost or 0
                service_item.status = 'QUOTED'
                db.session.add(service_item)
                services_created += 1

            if row.flag_transport:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_TRANSPORT
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Transport service for {row.city}"
                service_item.amount = row.transport_cost or 0
                service_item.status = 'QUOTED'
                db.session.add(service_item)
                services_created += 1

            if row.flag_meal:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_RESTAURANT
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Meal service for {row.city}"
                service_item.amount = row.meal_cost or 0
                service_item.status = 'QUOTED'
                db.session.add(service_item)
                services_created += 1

            if row.flag_guide:
                service_item = ServiceItem()
                service_item.booking_id = booking.id
                service_item.service_type = SERVICE_GUIDE
                service_item.start_date = row.date
                service_item.end_date = row.date
                service_item.description = f"Guide service for {row.city}"
                service_item.amount = row.guide_cost or 0
                service_item.status = 'QUOTED'
                db.session.add(service_item)
                services_created += 1

        # Recalculate booking total
        booking.calculate_total()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': f'Quote generated successfully with {services_created} services',
            'booking_id': booking.id,
            'services_count': services_created
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error generating quote: {str(e)}'
        }), 500

@inbound_bp.route('/api/<int:request_id>/confirm-all-suppliers', methods=['POST'])
@csrf.exempt
def api_confirm_all_suppliers(request_id):
    """Confirm all services with suppliers - changes all to RESERVED and updates request status"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models import STATUS_RESERVED

    try:
        # Update all hotels to RESERVED (individual services)
        for hotel in request_obj.inbound_hotels:
            hotel.status = STATUS_RESERVED

        # Update all transports to RESERVED (individual services)
        for transport in request_obj.inbound_transports:
            transport.status = STATUS_RESERVED

        # Update all meals to RESERVED (individual services)
        for meal in request_obj.inbound_meals:
            meal.status = STATUS_RESERVED

        # Update all guides to RESERVED (individual services)
        for guide in request_obj.inbound_guides:
            guide.status = STATUS_RESERVED

        # Update parent request status to QUOTED (after all suppliers are confirmed)
        request_obj.status = STATUS_QUOTED

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'All services confirmed with suppliers. Status updated to QUOTED.',
            'new_status': STATUS_QUOTED
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/service/<int:service_id>/confirm-supplier', methods=['POST'])
@csrf.exempt
def api_confirm_supplier(request_id, service_id):
    """Confirm a service with supplier - changes status to RESERVED"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    from app.models import STATUS_RESERVED
    from app.models.inbound import InboundHotel, InboundTransport

    service_type = request.json.get('service_type')

    try:
        # Find and update the service
        if service_type == 'hotel':
            service = InboundHotel.query.filter_by(id=service_id, request_id=request_id).first()
        elif service_type == 'transport':
            service = InboundTransport.query.filter_by(id=service_id, request_id=request_id).first()
        else:
            return jsonify({'error': 'Invalid service type'}), 400

        if not service:
            return jsonify({'error': 'Service not found'}), 404

        # Update status to RESERVED (supplier confirmed)
        service.status = STATUS_RESERVED
        db.session.commit()

        # Check if all services are confirmed
        all_confirmed = True
        for hotel in request_obj.hotels:
            if hotel.status != STATUS_RESERVED and hotel.status != STATUS_QUOTED:
                all_confirmed = False
                break

        for transport in request_obj.transports:
            if transport.status != STATUS_RESERVED and transport.status != STATUS_QUOTED:
                all_confirmed = False
                break

        return jsonify({
            'success': True,
            'message': 'Supplier confirmation recorded',
            'new_status': STATUS_RESERVED,
            'all_confirmed': all_confirmed
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/generate-proforma', methods=['POST'])
@csrf.exempt

def api_generate_proforma(request_id):
    """Generate a proforma invoice for a confirmed booking - changes status to QUOTED"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403


    # Check if status is QUOTED (suppliers confirmed)
    if request_obj.status != STATUS_QUOTED:
        return jsonify({
            'success': False,
            'message': 'Please confirm all services with suppliers first. Status must be QUOTED to generate proforma invoice.'
        }), 400

    if not request_obj.booking_id:
        return jsonify({
            'success': False,
            'message': 'No booking found. Please create a booking first.'
        }), 400

    try:
        booking = Booking.query.get(request_obj.booking_id)
        if not booking:
            return jsonify({
                'success': False,
                'message': 'Booking not found'
            }), 404

        # Generate proforma invoice number if not exists
        if not booking.invoice_number:
            booking.generate_invoice_number()

        # Status remains QUOTED (already set during supplier confirmation)
        booking.status = STATUS_QUOTED
        # request_obj.status already STATUS_QUOTED from supplier confirmation

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Proforma invoice generated successfully',
            'invoice_number': booking.invoice_number,
            'booking_id': booking.id,
            'redirect_url': f'/booking/{booking.id}/proforma-invoice'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error generating proforma invoice: {str(e)}'
        }), 500

@inbound_bp.route('/<int:request_id>/preview-proforma', methods=['GET'])
@login_required
def preview_proforma(request_id):
    """Preview proforma invoice on a web page before exporting to Word"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    # CREATE BOOKING IF IT DOESN'T EXIST
    if not request_obj.booking_id:
        # Get or create customer
        if request_obj.customer_id:
            customer = Customer.query.get(request_obj.customer_id)
        else:
            customer = Customer.query.filter_by(first_name=request_obj.contact_name).first()
            if not customer:
                customer = Customer()
                customer.first_name = request_obj.contact_name
                customer.last_name = ""
                customer.phone = "TBD"
                customer.email = "tbd@example.com"
                customer.nationality = request_obj.nationality
                db.session.add(customer)
                db.session.flush()

        # Create booking with QUOTED status
        booking = Booking()
        booking.reference_number = request_obj.request_number
        booking.user_id = request_obj.user_id
        booking.customer_id = customer.id
        booking.status = 'QUOTED'
        booking.total_amount = request_obj.total_amount or 0
        db.session.add(booking)
        db.session.flush()

        # Generate invoice number AFTER flush (so booking.id exists)
        booking.generate_invoice_number()

        # Link booking to request
        request_obj.booking_id = booking.id
        request_obj.status = 'QUOTED'
        db.session.commit()

    # Collect customer information
    customer_data = {}
    if request_obj.customer_id:
        customer = Customer.query.get(request_obj.customer_id)
        if customer:
            customer_data = {
                'name': customer.name,
                'company_name': customer.company_name,
                'email': customer.email,
                'phone': customer.phone,
                'nationality': customer.nationality
            }
    else:
        # Use contact name from request if no customer linked
        customer_data = {
            'name': request_obj.contact_name,
            'nationality': request_obj.nationality
        }

    # Collect tour information
    tour_data = {
        'from_date': request_obj.from_date.strftime('%d %b %Y') if request_obj.from_date else '',
        'to_date': request_obj.to_date.strftime('%d %b %Y') if request_obj.to_date else '',
        'pax': request_obj.pax,
        'nationality': request_obj.nationality
    }

    # Collect all service items with date ranges
    service_items = []

    # Add hotels
    for hotel in request_obj.inbound_hotels:
        service_items.append({
            'type': 'Hotel',
            'description': f"Hotel: {hotel.hotel_name or 'TBD'} - {hotel.location or ''} ({hotel.room_type or 'Standard'}, {hotel.meal_plan or 'BB'})",
            'date_from': hotel.check_in_date,
            'date_to': hotel.check_out_date,
            'pax': request_obj.pax,
            'unit_price': hotel.total_cost or 0,
            'total': hotel.total_cost or 0
        })

    # Add transport
    for transport in request_obj.inbound_transports:
        service_items.append({
            'type': 'Transport',
            'description': f"Transport: {transport.vehicle_type or 'Vehicle'} - {transport.pickup_location or ''} to {transport.dropoff_location or ''}",
            'date_from': transport.date,
            'date_to': transport.end_date if transport.end_date else transport.date,
            'pax': request_obj.pax,
            'unit_price': transport.cost or 0,
            'total': transport.cost or 0
        })

    # Add meals
    for meal in request_obj.inbound_meals:
        service_items.append({
            'type': 'Meal',
            'description': f"Meal: {meal.meal_type or 'Meal'} at {meal.restaurant or 'Restaurant'} - {meal.location or ''}",
            'date_from': meal.date,
            'date_to': meal.end_date if meal.end_date else meal.date,
            'pax': request_obj.pax,
            'unit_price': meal.cost_per_person or 0,
            'total': meal.total_cost or 0
        })

    # Add guides
    for guide in request_obj.inbound_guides:
        service_items.append({
            'type': 'Guide',
            'description': f"Guide: {guide.service_type or 'Guide Service'} - {guide.guide_name or 'TBD'} ({guide.language or 'English'})",
            'date_from': guide.date,
            'date_to': guide.end_date if guide.end_date else guide.date,
            'pax': request_obj.pax,
            'unit_price': guide.cost or 0,
            'total': guide.cost or 0
        })

    # Sort service items by date
    service_items.sort(key=lambda x: x['date_from'] if x['date_from'] else datetime.max.date())

    # Calculate total
    grand_total = sum(item['total'] for item in service_items)

    # Update status to QUOTED when generating proforma invoice preview
    if request_obj.status not in ['QUOTED', 'CONFIRMED']:
        request_obj.status = 'QUOTED'

        # Also update booking status if it exists
        if request_obj.booking_id:
            booking = Booking.query.get(request_obj.booking_id)
            if booking:
                booking.status = 'QUOTED'

        db.session.commit()

    # Prepare invoice data for template
    invoice_data = {
        'invoice_number': request_obj.request_number,
        'invoice_date': datetime.now().strftime('%d %b %Y'),
        'company_name': 'Windows of Jordan',
        'company_address': 'Amman, Jordan',
        'customer': customer_data,
        'tour': tour_data,
        'service_items': service_items,
        'grand_total': grand_total
    }

    return render_template('inbound/preview_proforma.html', 
                         request=request_obj,
                         invoice=invoice_data)

@inbound_bp.route('/api/<int:request_id>/update-proforma-prices', methods=['POST'])
@csrf.exempt

def update_proforma_prices(request_id):
    """Update pricing for proforma invoice service items"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        items = data.get('items', [])

        # Collect all service models in order
        all_services = []
        for hotel in request_obj.inbound_hotels:
            all_services.append(('hotel', hotel))
        for transport in request_obj.inbound_transports:
            all_services.append(('transport', transport))
        for meal in request_obj.inbound_meals:
            all_services.append(('meal', meal))
        for guide in request_obj.inbound_guides:
            all_services.append(('guide', guide))

        # Update each service based on index
        for item in items:
            index = item['index']
            if index < len(all_services):
                service_type, service = all_services[index]

                # Parse dates if provided
                from datetime import datetime as dt
                date_from = None
                date_to = None
                if item.get('date_from'):
                    try:
                        date_from = dt.strptime(item['date_from'], '%Y-%m-%d').date()
                    except:
                        pass
                if item.get('date_to'):
                    try:
                        date_to = dt.strptime(item['date_to'], '%Y-%m-%d').date()
                    except:
                        pass

                if service_type == 'hotel':
                    service.total_cost = item['total']
                    if item.get('description'):
                        service.hotel_name = item['description']
                    # Note: check_in_date and check_out_date are managed at InboundHotel level only
                    # Rooms cascade dates from parent hotel via @property methods (read-only)
                elif service_type == 'transport':
                    service.cost = item['unit_price']
                    if item.get('description'):
                        service.notes = item['description']
                    if date_from:
                        service.date = date_from
                    if date_to:
                        service.end_date = date_to
                elif service_type == 'meal':
                    service.cost_per_person = item['unit_price']
                    service.total_cost = item['total']
                    if item.get('description'):
                        service.restaurant = item['description']
                    if date_from:
                        service.date = date_from
                    if date_to:
                        service.end_date = date_to
                elif service_type == 'guide':
                    service.cost_per_day = item['unit_price']
                    service.total_cost = item['total']
                    if item.get('description'):
                        service.guide_name = item['description']
                    if date_from:
                        service.date = date_from
                    if date_to:
                        service.end_date = date_to

        # Recalculate total
        total = sum(item['total'] for item in items)
        request_obj.total_amount = total

        # Update booking total if exists
        if request_obj.booking_id:
            booking = Booking.query.get(request_obj.booking_id)
            if booking:
                booking.total_amount = total

        db.session.commit()
        return jsonify({'success': True, 'message': 'Prices updated successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/update-pricing-mode', methods=['POST'])
@csrf.exempt

def update_pricing_mode(request_id):
    """Update pricing mode for proforma invoice (ITEMIZED or LUMPSUM)"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        pricing_mode = data.get('pricing_mode', 'ITEMIZED')

        if pricing_mode not in ['ITEMIZED', 'LUMPSUM']:
            return jsonify({'success': False, 'message': 'Invalid pricing mode'}), 400

        request_obj.pricing_mode = pricing_mode
        db.session.commit()

        return jsonify({'success': True, 'message': f'Pricing mode updated to {pricing_mode}'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/export-proforma-doc', methods=['GET'])
@csrf.exempt

def api_export_proforma_doc(request_id):
    """Export proforma invoice as Word document with service line items"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    try:
        # Collect customer information
        customer_data = {}
        if request_obj.customer_id:
            customer = Customer.query.get(request_obj.customer_id)
            if customer:
                customer_data = {
                    'name': customer.name,
                    'company_name': customer.company_name,
                    'email': customer.email,
                    'phone': customer.phone,
                    'nationality': customer.nationality
                }
        else:
            # Use contact name from request if no customer linked
            customer_data = {
                'name': request_obj.contact_name,
                'nationality': request_obj.nationality
            }

        # Collect tour information
        tour_data = {
            'from_date': request_obj.from_date.strftime('%d %b %Y') if request_obj.from_date else '',
            'to_date': request_obj.to_date.strftime('%d %b %Y') if request_obj.to_date else '',
            'pax': request_obj.pax,
            'nationality': request_obj.nationality
        }

        # Collect all service items with date ranges
        service_items = []

        # Add hotels
        for hotel in request_obj.inbound_hotels:
            service_items.append({
                'description': f"Hotel: {hotel.hotel_name or 'TBD'} - {hotel.location or ''} ({hotel.room_type or 'Standard'}, {hotel.meal_plan or 'BB'})",
                'date_from': hotel.check_in_date,
                'date_to': hotel.check_out_date,
                'pax': request_obj.pax,
                'unit_price': hotel.total_cost,
                'total': hotel.total_cost
            })

        # Add transport (exclude unsaved individual-transport stubs from flights)
        for transport in request_obj.inbound_transports:
            if not _include_transport_in_trip_summary(transport):
                continue
            service_items.append({
                'description': f"Transport: {transport.vehicle_type or 'Vehicle'} - {transport.pickup_location or ''} to {transport.dropoff_location or ''}",
                'date_from': transport.date,
                'date_to': transport.end_date if transport.end_date else transport.date,
                'pax': request_obj.pax,
                'unit_price': transport.cost,
                'total': transport.cost
            })

        # Add meals
        for meal in request_obj.inbound_meals:
            service_items.append({
                'description': f"Meal: {meal.meal_type or 'Meal'} at {meal.restaurant or 'Restaurant'} - {meal.location or ''}",
                'date_from': meal.date,
                'date_to': meal.end_date if meal.end_date else meal.date,
                'pax': request_obj.pax,
                'unit_price': meal.cost_per_person,
                'total': meal.total_cost
            })

        # Add guides
        for guide in request_obj.inbound_guides:
            service_items.append({
                'description': f"Guide: {guide.service_type or 'Guide Service'} - {guide.guide_name or 'TBD'} ({guide.language or 'English'})",
                'date_from': guide.date,
                'date_to': guide.end_date if guide.end_date else guide.date,
                'pax': request_obj.pax,
                'unit_price': guide.cost,
                'total': guide.cost
            })

        # Sort service items by date
        service_items.sort(key=lambda x: x['date_from'] if x['date_from'] else datetime.max.date())

        # Update booking status to QUOTED when exporting proforma
        if request_obj.booking_id:
            booking = Booking.query.get(request_obj.booking_id)
            if booking and booking.status != 'QUOTED':
                booking.status = 'QUOTED'
                request_obj.status = 'QUOTED'
                db.session.commit()

        # Prepare invoice data
        invoice_data = {
            'invoice_number': request_obj.request_number,
            'invoice_date': datetime.now().strftime('%d %b %Y'),
            'company_name': 'Windows of Jordan',
            'company_address': 'Amman, Jordan',
            'customer': customer_data,
            'tour': tour_data,
            'service_items': service_items
        }

        # Generate Word document
        generator = ProformaDocGenerator()
        output_path = generator.generate_proforma(invoice_data)

        # Record generated proforma in inbound_document (non-fatal)
        try:
            original_name = f'Proforma_{request_obj.request_number}.docx'
            with open(output_path, 'rb') as _f:
                doc_bytes = _f.read()
            rel_path, stored_name = document_storage.save(
                doc_bytes, request_id=request_id, doc_type='PROFORMA',
                original_filename=original_name
            )
            _record_document(request_id, 'PROFORMA', stored_name, rel_path, doc_bytes,
                             mime_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        except Exception as _rec_err:
            current_app.logger.warning(f'Proforma document record skipped: {_rec_err}')

        # Send file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'Proforma_{request_obj.request_number}.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

    except Exception as e:
        flash(f'Error generating proforma document: {str(e)}', 'error')
        return redirect(url_for('inbound.view_request', request_id=request_id))

@inbound_bp.route('/api/<int:request_id>/export-voucher-doc', methods=['GET'])
@csrf.exempt

def api_export_voucher_doc(request_id):
    """Export trip voucher as Word document with full itinerary"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    try:
        # Collect tour information
        customer_name = request_obj.contact_name
        if request_obj.customer_id:
            customer = Customer.query.get(request_obj.customer_id)
            if customer:
                customer_name = customer.name

        tour_data = {
            'guest_name': customer_name,
            'nationality': request_obj.nationality,
            'pax': request_obj.pax,
            'agent_ref': request_obj.agent_ref or '',
            'notes': request_obj.special_note or '',
            'tour_file': request_obj.request_number,
            'from_date': request_obj.from_date.strftime('%d-%b-%y') if request_obj.from_date else '',
            'to_date': request_obj.to_date.strftime('%d-%b-%y') if request_obj.to_date else ''
        }

        # Collect arrivals/departures data from flagged transport services
        arrivals_data = []
        arrival_departure_transports = [
            t for t in request_obj.inbound_transports 
            if (t.is_arrival or t.is_departure) and _include_transport_in_trip_summary(t)
        ]

        if arrival_departure_transports:
            for transport in sorted(arrival_departure_transports, key=lambda x: x.date):
                border = 'Airport'
                if transport.pickup_location and 'border' in transport.pickup_location.lower():
                    border = 'Border'

                drop_point = transport.dropoff_location or 'TBA'
                if transport.is_departure:
                    drop_point = transport.pickup_location or 'TBA'

                time_str = transport.pickup_time.strftime('%H:%M') if transport.pickup_time else ''

                arrivals_data.append({
                    'date': transport.date.strftime('%d-%b-%y'),
                    'border': border,
                    'drop_point': drop_point,
                    'pax': request_obj.pax,
                    'carrier': '',
                    'flight': '',
                    'time': time_str,
                    'note': f"{transport.vehicle_type}" if transport.vehicle_type else ''
                })
        else:
            # Add default arrival/departure if no flagged transfers
            arrivals_data.append({
                'date': request_obj.from_date.strftime('%d-%b-%y') if request_obj.from_date else '',
                'border': 'Airport',
                'drop_point': 'TBA',
                'pax': request_obj.pax,
                'carrier': '',
                'flight': '',
                'time': '',
                'note': ''
            })
            arrivals_data.append({
                'date': request_obj.to_date.strftime('%d-%b-%y') if request_obj.to_date else '',
                'border': 'Airport',
                'drop_point': 'TBA',
                'pax': request_obj.pax,
                'carrier': '',
                'flight': '',
                'time': '',
                'note': ''
            })

        # Collect hotel details
        hotels_data = []
        for hotel in request_obj.inbound_hotels:
            # Get room data from itinerary rows for this hotel's date range
            single_rooms = 0
            double_rooms = 0
            twin_rooms = 0
            triple_rooms = 0
            other_rooms = 0

            # Find itinerary row with hotel flag for this hotel's check-in date
            for row in request_obj.itinerary_rows:
                if row.flag_hotel and row.date == hotel.check_in_date:
                    single_rooms = row.hotel_single_rooms or 0
                    double_rooms = row.hotel_double_rooms or 0
                    twin_rooms = 0  # We use double for DBL
                    triple_rooms = row.hotel_triple_rooms or 0
                    other_rooms = row.hotel_other_rooms or 0
                    break

            hotels_data.append({
                'check_in': hotel.check_in_date.strftime('%d-%b-%y') if hotel.check_in_date else 'TBA',
                'check_out': hotel.check_out_date.strftime('%d-%b-%y') if hotel.check_out_date else 'TBA',
                'name': hotel.hotel_name or 'Hotel TBA',
                'board_basis': hotel.meal_plan or 'BB',
                'note': '',
                'single_rooms': single_rooms,
                'double_rooms': double_rooms,
                'twin_rooms': twin_rooms,
                'triple_rooms': triple_rooms,
                'other_rooms': other_rooms
            })

        # Build itinerary organized by service type
        itinerary_days = []

        # Group all services by date
        services_by_date = {}

        # Add hotels
        for hotel in request_obj.inbound_hotels:
            current_date = hotel.check_in_date
            while current_date < hotel.check_out_date:
                date_key = current_date.strftime('%d-%b-%y')
                if date_key not in services_by_date:
                    services_by_date[date_key] = []

                services_by_date[date_key].append(f"Hotel: {hotel.hotel_name or 'TBA'}")
                current_date += timedelta(days=1)

        # Add transport
        for transport in request_obj.inbound_transports:
            if not _include_transport_in_trip_summary(transport):
                continue
            date_key = transport.date.strftime('%d-%b-%y')
            if date_key not in services_by_date:
                services_by_date[date_key] = []

            services_by_date[date_key].append(
                f"Transport: {transport.pickup_location or 'TBA'} → {transport.dropoff_location or 'TBA'}"
            )

        # Add meals
        for meal in request_obj.inbound_meals:
            current_date = meal.date
            end_date = meal.end_date if meal.end_date else meal.date
            while current_date <= end_date:
                date_key = current_date.strftime('%d-%b-%y')
                if date_key not in services_by_date:
                    services_by_date[date_key] = []

                services_by_date[date_key].append(
                    f"{meal.meal_type or 'Meal'}: {meal.restaurant or 'TBA'}"
                )
                current_date += timedelta(days=1)

        # Add guides
        for guide in request_obj.inbound_guides:
            current_date = guide.date
            end_date = guide.end_date if guide.end_date else guide.date
            while current_date <= end_date:
                date_key = current_date.strftime('%d-%b-%y')
                if date_key not in services_by_date:
                    services_by_date[date_key] = []

                services_by_date[date_key].append(
                    f"Guide: {guide.service_type or 'Guide Service'} ({guide.language or 'English'})"
                )
                current_date += timedelta(days=1)

        # Convert to list format
        for date_key in sorted(services_by_date.keys(), key=lambda x: datetime.strptime(x, '%d-%b-%y')):
            description = '\n'.join(services_by_date[date_key])
            itinerary_days.append({
                'date': date_key,
                'description': description
            })

        # Collect meals data
        meals_data = []
        for meal in request_obj.inbound_meals:
            start_date = meal.date
            end_date = meal.end_date if meal.end_date else meal.date

            # Generate entry for each day in range
            current_date = start_date
            while current_date <= end_date:
                meals_data.append({
                    'date': current_date.strftime('%d-%b-%y'),
                    'restaurant': meal.restaurant or 'Restaurant',
                    'meal_type': meal.meal_type or 'Lunch',
                    'pax': request_obj.pax,
                    'note': ''
                })
                current_date += timedelta(days=1)

        # Collect transport data
        transport_data = []
        for transport in request_obj.inbound_transports:
            if not _include_transport_in_trip_summary(transport):
                continue
            start_date = transport.date
            end_date = transport.end_date if transport.end_date else transport.date

            # Generate entry for date range
            current_date = start_date
            while current_date <= end_date:
                transport_data.append({
                    'time': f"{current_date.strftime('%d-%b-%y')} - {end_date.strftime('%d-%b-%y')}",
                    'name': transport.vehicle_type or 'Vehicle',
                    'note': f"{transport.pickup_location or ''} to {transport.dropoff_location or ''}",
                    'driver': ''
                })
                break  # Only add once for range

        # Collect guides data
        guides_data = []
        for guide in request_obj.inbound_guides:
            guides_data.append({
                'from_date': guide.date.strftime('%d-%b-%y') if guide.date else '',
                'to_date': guide.end_date.strftime('%d-%b-%y') if guide.end_date else guide.date.strftime('%d-%b-%y'),
                'name': guide.guide_name or 'TBA',
                'language': guide.language or 'English',
                'note': guide.service_type or ''
            })

        # Collect cash expenses data
        cash_expenses_data = []
        for expense in request_obj.inbound_cash_expenses:
            start_date = expense.date
            end_date = expense.end_date if expense.end_date else expense.date

            # Generate entry for each day in range
            current_date = start_date
            while current_date <= end_date:
                amount_display = f"{expense.currency} {expense.amount:.2f}"
                if expense.is_per_person:
                    amount_display += " pp"

                cash_expenses_data.append({
                    'date': current_date.strftime('%d-%b-%y'),
                    'category': expense.category or 'Expense',
                    'description': expense.description,
                    'amount': amount_display,
                    'driver_name': expense.driver_name or '',
                    'note': expense.location or ''
                })
                current_date += timedelta(days=1)

        # Prepare voucher data
        voucher_data = {
            'tour_file': request_obj.request_number,
            'company_name': 'Windows of Jordan',
            'tour': tour_data,
            'arrivals': arrivals_data,
            'hotels': hotels_data,
            'itinerary_days': itinerary_days,
            'meals': meals_data,
            'transport': transport_data,
            'guides': guides_data,
            'cash_expenses': cash_expenses_data
        }

        # Generate Word document
        generator = VoucherTripPlanGenerator()
        output_path = generator.generate_voucher(voucher_data)

        # Record generated voucher doc in inbound_document (non-fatal)
        try:
            original_name = f'Voucher_{request_obj.request_number}.docx'
            with open(output_path, 'rb') as _f:
                doc_bytes = _f.read()
            rel_path, stored_name = document_storage.save(
                doc_bytes, request_id=request_id, doc_type='VOUCHER',
                original_filename=original_name
            )
            _record_document(request_id, 'VOUCHER', stored_name, rel_path, doc_bytes,
                             mime_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')
        except Exception as _rec_err:
            current_app.logger.warning(f'Voucher doc record skipped: {_rec_err}')

        # Send file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'Voucher_{request_obj.request_number}.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )

    except Exception as e:
        flash(f'Error generating voucher document: {str(e)}', 'error')
        return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/api/<int:request_id>/confirm-booking', methods=['POST'])
@csrf.exempt

def api_confirm_booking(request_id):
    """Confirm a booking after proforma invoice is generated"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    if not request_obj.booking_id:
        return jsonify({
            'success': False,
            'message': 'No booking found. Please generate a quote first.'
        }), 400

    try:
        booking = Booking.query.get(request_obj.booking_id)
        if not booking:
            return jsonify({
                'success': False,
                'message': 'Booking not found'
            }), 404

        if booking.status not in ['QUOTED', 'PROFORMA_GENERATED']:
            return jsonify({
                'success': False,
                'message': 'Booking must have proforma invoice before confirmation'
            }), 400

        # Confirm the booking and move to CONFIRMED status
        booking.status = 'CONFIRMED'
        request_obj.status = 'CONFIRMED'

        # Update all service items to CONFIRMED status
        for service_item in booking.service_items:
            service_item.status = 'CONFIRMED'

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Booking confirmed successfully',
            'booking_id': booking.id
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error confirming booking: {str(e)}'
        }), 500

@inbound_bp.route('/api/<int:request_id>/start-processing', methods=['POST'])
@csrf.exempt

def api_start_processing(request_id):
    """Start processing an itinerary - change status from CONFIRMED to PROCESSING"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'error': 'Access denied'}), 403

    if request_obj.status != 'CONFIRMED':
        return jsonify({
            'success': False,
            'message': 'Itinerary must be CONFIRMED before processing can start'
        }), 400

    try:
        # Change status to PROCESSING (operations active)
        request_obj.status = 'PROCESSING'

        if request_obj.booking_id:
            booking = Booking.query.get(request_obj.booking_id)
            if booking:
                booking.status = 'PROCESSING'

        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'Processing started successfully. Operations are now active.',
            'new_status': 'PROCESSING'
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'message': f'Error starting processing: {str(e)}'
        }), 500

# ============================================================
# RUN-DOWN PLAN DASHBOARD
# ============================================================

RUN_DOWN_SERVICES = [
    {'key': 'HOTEL', 'label': 'Accommodation', 'icon': 'fa-hotel', 'accent': '#1d4ed8'},
    {'key': 'TRANSPORT', 'label': 'Transportation', 'icon': 'fa-bus', 'accent': '#c2410c'},
    {'key': 'GUIDE', 'label': 'Guides', 'icon': 'fa-user-tie', 'accent': '#6d28d9'},
    {'key': 'MEAL', 'label': 'Restaurant', 'icon': 'fa-utensils', 'accent': '#9333ea'},
    {'key': 'GROUND_HANDLER', 'label': 'Meet & Assist', 'icon': 'fa-handshake', 'accent': '#b45309'},
    {'key': 'OPTIONAL', 'label': 'Optional', 'icon': 'fa-star', 'accent': '#15803d'},
]

_RUN_DOWN_SUPPLIER_PATTERNS = {
    'HOTEL': ['HOTEL', 'ACCOMMODATION'],
    'TRANSPORT': ['TRANSPORT', 'TRANSPORTATION', 'TRANSFER'],
    'GUIDE': ['GUIDE'],
    'MEAL': ['RESTAURANT', 'MEAL', 'FOOD'],
    'GROUND_HANDLER': ['GROUND_HANDLER', 'MEET', 'ASSIST'],
}


def _parse_run_down_dates(default_to_today=True):
    """Parse date_from/date_to query params for run-down endpoints."""
    today = datetime.now().date()
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')
    try:
        date_from = (
            datetime.strptime(date_from_str, '%Y-%m-%d').date()
            if date_from_str else (today if default_to_today else None)
        )
        date_to = (
            datetime.strptime(date_to_str, '%Y-%m-%d').date()
            if date_to_str else (today if default_to_today else None)
        )
        if date_from and date_to and date_from > date_to:
            date_from, date_to = date_to, date_from
        return date_from, date_to
    except ValueError:
        return None, None


def _run_down_row(request_obj, service_date, description, status, pax, service_type, record_id=None, meal_type=None, meal_note=None, voucher_notes=None, check_out_date=None, room_type=None, meal_plan=None, hotel_obj=None, end_date=None, language=None, guide_notes=None, transport_notes=None, pickup_location=None, dropoff_location=None, meet_assist_notes=None, supplier_obj=None, service_time=None, flight_number=None):
    """Build a normalized run-down request row for API responses."""
    base_row = {
        'record_id': record_id,
        'request_id': request_obj.id,
        'request_number': request_obj.request_number,
        'date': service_date.strftime('%Y-%m-%d'),
        'date_display': service_date.strftime('%d %b %Y'),
        'description': description,
        'status': status or 'REQUEST',
        'file_status': request_obj.status or 'REQUEST',
        'pax': pax or request_obj.pax or 0,
        'service_type': service_type,
        'view_url': url_for('inbound.view_request', id=request_obj.id),
    }

    # Add extended fields for specific services
    if service_type == 'MEAL':
        group_name = request_obj.agent_ref or ''

        base_row.update({
            'day_of_week': service_date.strftime('%a'),
            'group_name': group_name,
            'nationality': request_obj.nationality or '—',
            'meal': meal_type if meal_type else 'Meal',
            'notes': request_obj.special_note or '—',
            'restaurant_note': meal_note or '—',
        })
    elif service_type == 'HOTEL':
        group_name = request_obj.agent_ref or ''

        # Calculate room counts from HotelRoom records and collect all room categories
        sgl = dbl = twn = trpl = other = 0
        room_categories = set()
        if hotel_obj and hasattr(hotel_obj, 'rooms'):
            for room in hotel_obj.rooms:
                room_type_upper = (room.room_type or '').upper()
                if room_type_upper == 'SINGLE':
                    sgl += room.room_count or 1
                elif room_type_upper == 'DOUBLE':
                    dbl += room.room_count or 1
                elif room_type_upper == 'TWIN':
                    twn += room.room_count or 1
                elif room_type_upper == 'TRIPLE':
                    trpl += room.room_count or 1
                elif room_type_upper == 'OTHER':
                    other += room.room_count or 1
                # Collect all unique room categories
                if room.room_category:
                    room_categories.add(room.room_category)

        total = sgl + dbl + twn + trpl + other
        room_category = ', '.join(sorted(room_categories)) if room_categories else '—'

        nights_calc = 0
        if hotel_obj and hotel_obj.check_in_date and hotel_obj.check_out_date:
            nights_calc = (hotel_obj.check_out_date - hotel_obj.check_in_date).days

        base_row.update({
            'check_out_date': check_out_date.strftime('%d %b %Y') if check_out_date else '—',
            'group_name': group_name,
            'nationality': request_obj.nationality or '—',
            'meal_plan': meal_plan or '—',
            'nights': nights_calc,
            'room_category': room_category,
            'sgl': sgl,
            'dbl': dbl,
            'twn': twn,
            'trpl': trpl,
            'other': other,
            'total': total,
        })
    elif service_type == 'GUIDE':
        group_name = request_obj.agent_ref or ''

        base_row.update({
            'date_from': service_date.strftime('%d %b %Y'),
            'date_to': end_date.strftime('%d %b %Y') if end_date else service_date.strftime('%d %b %Y'),
            'group_name': group_name,
            'nationality': request_obj.nationality or '—',
            'language': language or '—',
            'notes': request_obj.special_note or '—',
            'guide_note': guide_notes or '—',
        })
    elif service_type == 'TRANSPORT':
        group_name = request_obj.agent_ref or ''

        base_row.update({
            'date_from': service_date.strftime('%d %b %Y'),
            'date_to': end_date.strftime('%d %b %Y') if end_date else service_date.strftime('%d %b %Y'),
            'day_of_week': service_date.strftime('%a'),
            'group_name': group_name,
            'nationality': request_obj.nationality or '—',
            'notes': request_obj.special_note or '—',
            'transport_note': transport_notes or '—',
            'pickup_location': pickup_location or '—',
            'dropoff_location': dropoff_location or '—',
        })
    elif service_type == 'GROUND_HANDLER':
        group_name = request_obj.agent_ref or request_obj.contact_name or ''
        supplier_languages = '—'
        if supplier_obj and supplier_obj.languages:
            supplier_languages = supplier_obj.languages

        time_display = '—'
        if service_time and hasattr(service_time, 'strftime'):
            time_display = service_time.strftime('%H:%M')

        base_row.update({
            'group_name': group_name,
            'nationality': request_obj.nationality or '—',
            'language': supplier_languages,
            'notes': request_obj.special_note or '—',
            'ma_notes': meet_assist_notes or '—',
            'time': time_display,
            'flight_number': flight_number or '—',
        })
    else:
        # Original fields for other services
        base_row['contact_name'] = request_obj.contact_name or '—'

    return base_row


def _fetch_run_down_supplier_requests(service_key, supplier, date_from, date_to):
    """Return inbound service rows for a supplier within a date range."""
    from app.models.inbound import InboundOptional
    from sqlalchemy import or_ as _or

    supplier_id = supplier.id
    supplier_name = supplier.name
    rows = []

    if service_key == 'HOTEL':
        hotels = (
            InboundHotel.query.join(InboundRequest)
            .filter(
                InboundHotel.check_in_date >= date_from,
                InboundHotel.check_in_date <= date_to,
                InboundHotel.hotel_name.ilike(supplier_name),
            )
            .all()
        )
        for hotel in hotels:
            req = hotel.request
            rows.append(_run_down_row(
                req,
                hotel.check_in_date,
                f"{hotel.hotel_name or supplier_name} – {hotel.location or ''} ({hotel.nights}n)".strip(' –'),
                hotel.status,
                req.pax,
                'HOTEL',
                hotel.id,
                check_out_date=hotel.check_out_date,
                room_type=hotel.room_type,
                meal_plan=hotel.meal_plan,
                hotel_obj=hotel,
            ))

    elif service_key == 'TRANSPORT':
        transports = (
            InboundTransport.query.join(InboundRequest)
            .filter(
                InboundTransport.date >= date_from,
                InboundTransport.date <= date_to,
                _or(
                    InboundTransport.supplier_id == supplier_id,
                    InboundTransport.supplier.ilike(supplier_name),
                ),
            )
            .all()
        )
        for transport in transports:
            req = transport.request
            # Ensure pickup and dropoff values are retrieved
            pickup = transport.pickup_location or transport.pickup_point or ''
            dropoff = transport.dropoff_location or transport.drop_off_point or ''
            rows.append(_run_down_row(
                req,
                transport.date,
                f"{transport.vehicle_type or 'Transport'} – {pickup or 'TBA'} → {dropoff or 'TBA'}",
                transport.status,
                transport.pax or req.pax,
                'TRANSPORT',
                transport.id,
                end_date=transport.end_date,
                transport_notes=transport.note,
                pickup_location=pickup,
                dropoff_location=dropoff,
            ))

    elif service_key == 'GUIDE':
        guides = (
            InboundGuide.query.join(InboundRequest)
            .filter(
                InboundGuide.date >= date_from,
                InboundGuide.date <= date_to,
                InboundGuide.is_cancelled == False,
                InboundGuide.guide_name.ilike(supplier_name),
            )
            .all()
        )
        for guide in guides:
            req = guide.request
            rows.append(_run_down_row(
                req,
                guide.date,
                f"{guide.guide_name or supplier_name} – {guide.language or 'N/A'}",
                guide.status,
                req.pax,
                'GUIDE',
                guide.id,
                end_date=guide.end_date,
                language=guide.language,
                guide_notes=guide.additional_comments,
            ))

    elif service_key == 'MEAL':
        meals = (
            InboundMeal.query.join(InboundRequest)
            .filter(
                InboundMeal.date >= date_from,
                InboundMeal.date <= date_to,
                _or(
                    InboundMeal.supplier_id == supplier_id,
                    InboundMeal.restaurant.ilike(supplier_name),
                ),
            )
            .all()
        )
        for meal in meals:
            req = meal.request
            rows.append(_run_down_row(
                req,
                meal.date,
                f"{meal.meal_type or 'Meal'} at {meal.supplier_name or supplier_name}",
                meal.status,
                req.pax,
                'MEAL',
                meal.id,
                meal_type=meal.meal_type,
                meal_note=meal.meal_note,
                voucher_notes=meal.voucher_notes,
            ))

    elif service_key == 'GROUND_HANDLER':
        arrivals = (
            ArrivalBatch.query.join(InboundRequest)
            .filter(
                ArrivalBatch.arrival_date >= date_from,
                ArrivalBatch.arrival_date <= date_to,
                ArrivalBatch.supplier_id == supplier_id,
            )
            .all()
        )
        for batch in arrivals:
            req = batch.request
            description = f"Arrival – {batch.arrival_point or 'TBA'}"
            if batch.batch_name and batch.batch_name != 'Batch':
                description += f" ({batch.batch_name})"
            rows.append(_run_down_row(
                req,
                batch.arrival_date,
                description,
                req.status or 'REQUEST',
                batch.pax_count or req.pax,
                'GROUND_HANDLER',
                batch.id,
                meet_assist_notes=batch.notes,
                supplier_obj=batch.supplier_ref,
                service_time=batch.arrival_time,
                flight_number=batch.flight_number,
            ))
        departures = (
            DepartureBatch.query.join(InboundRequest)
            .filter(
                DepartureBatch.departure_date >= date_from,
                DepartureBatch.departure_date <= date_to,
                DepartureBatch.supplier_id == supplier_id,
            )
            .all()
        )
        for batch in departures:
            req = batch.request
            description = f"Departure – {batch.departure_point or 'TBA'}"
            if batch.batch_name and batch.batch_name != 'Batch':
                description += f" ({batch.batch_name})"
            rows.append(_run_down_row(
                req,
                batch.departure_date,
                description,
                req.status or 'REQUEST',
                batch.pax_count or req.pax,
                'GROUND_HANDLER',
                batch.id,
                meet_assist_notes=batch.notes,
                supplier_obj=batch.supplier_ref,
                service_time=batch.departure_time,
                flight_number=batch.flight_number,
            ))

    elif service_key == 'OPTIONAL':
        optionals = (
            InboundOptional.query.join(InboundRequest)
            .filter(
                _or(
                    InboundOptional.date == None,
                    db.and_(
                        InboundOptional.date >= date_from,
                        InboundOptional.date <= date_to,
                    ),
                ),
                InboundOptional.supplier.ilike(supplier_name),
            )
            .all()
        )
        for optional in optionals:
            req = optional.request
            svc_date = optional.date or date_from
            rows.append(_run_down_row(
                req,
                svc_date,
                optional.service_name,
                optional.status,
                req.pax,
                'OPTIONAL',
                optional.id,
            ))

    rows.sort(key=lambda r: (r['date'], r['request_number']))
    return rows


def _run_down_supplier_matches(service_key, date_from, date_to):
    """Return (supplier_ids, supplier_names_lower) that have at least one
    request within the date range for the given service.

    Mirrors the exact date/name/id matching used by
    _fetch_run_down_supplier_requests so the dropdown list and the displayed
    results always agree on which suppliers fall inside the selected range.
    """
    from app.models.inbound import InboundOptional
    from sqlalchemy import or_ as _or

    ids = set()
    names = set()

    def _add_names(rows):
        for (val,) in rows:
            if val:
                names.add(val.lower())

    def _add_ids(rows):
        for (val,) in rows:
            if val:
                ids.add(val)

    if service_key == 'HOTEL':
        _add_names(
            db.session.query(InboundHotel.hotel_name)
            .join(InboundRequest)
            .filter(
                InboundHotel.check_in_date >= date_from,
                InboundHotel.check_in_date <= date_to,
            )
            .all()
        )

    elif service_key == 'TRANSPORT':
        rows = (
            db.session.query(InboundTransport.supplier_id, InboundTransport.supplier)
            .join(InboundRequest)
            .filter(
                InboundTransport.date >= date_from,
                InboundTransport.date <= date_to,
            )
            .all()
        )
        for sid, sname in rows:
            if sid:
                ids.add(sid)
            if sname:
                names.add(sname.lower())

    elif service_key == 'GUIDE':
        _add_names(
            db.session.query(InboundGuide.guide_name)
            .join(InboundRequest)
            .filter(
                InboundGuide.date >= date_from,
                InboundGuide.date <= date_to,
                InboundGuide.is_cancelled == False,
            )
            .all()
        )

    elif service_key == 'MEAL':
        rows = (
            db.session.query(InboundMeal.supplier_id, InboundMeal.restaurant)
            .join(InboundRequest)
            .filter(
                InboundMeal.date >= date_from,
                InboundMeal.date <= date_to,
            )
            .all()
        )
        for sid, sname in rows:
            if sid:
                ids.add(sid)
            if sname:
                names.add(sname.lower())

    elif service_key == 'GROUND_HANDLER':
        _add_ids(
            db.session.query(ArrivalBatch.supplier_id)
            .join(InboundRequest)
            .filter(
                ArrivalBatch.arrival_date >= date_from,
                ArrivalBatch.arrival_date <= date_to,
            )
            .all()
        )
        _add_ids(
            db.session.query(DepartureBatch.supplier_id)
            .join(InboundRequest)
            .filter(
                DepartureBatch.departure_date >= date_from,
                DepartureBatch.departure_date <= date_to,
            )
            .all()
        )

    elif service_key == 'OPTIONAL':
        # Option 1: undated optionals (date IS NULL) always appear, matching the
        # results behaviour in _fetch_run_down_supplier_requests.
        _add_names(
            db.session.query(InboundOptional.supplier)
            .join(InboundRequest)
            .filter(
                _or(
                    InboundOptional.date == None,
                    db.and_(
                        InboundOptional.date >= date_from,
                        InboundOptional.date <= date_to,
                    ),
                ),
            )
            .all()
        )

    return ids, names


def get_status_color(status):
    """Map booking status to color for visual coding"""
    status_colors = {
        'QUOTED': '#3b82f6',  # Blue
        'PROFORMA_GENERATED': '#8b5cf6',  # Purple
        'BOOKED': '#eab308',  # Yellow/Pending
        'CONFIRMED': '#22c55e',  # Green
        'COMPLETED': '#10b981',  # Green
        'CANCELLED': '#ef4444',  # Red
        'REQUEST': '#64748b',  # Gray
    }
    return status_colors.get(status, '#94a3b8')

@inbound_bp.route('/run-down')
@login_required
def run_down_plan():
    """Run-down page: date-filtered services with supplier lookup."""
    today = datetime.now().date()
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')
    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else today
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else today
        if date_from > date_to:
            date_from, date_to = date_to, date_from
    except ValueError:
        date_from = date_to = today

    return render_template(
        'inbound/run_down.html',
        services=RUN_DOWN_SERVICES,
        date_from=date_from,
        date_to=date_to,
    )


@inbound_bp.route('/run-down/suppliers')
@login_required
def run_down_suppliers():
    """JSON: searchable suppliers for a run-down service category."""
    from sqlalchemy import or_ as _or, func as _func

    service_key = request.args.get('service', '').upper()
    query = request.args.get('q', '').strip()
    date_from, date_to = _parse_run_down_dates()
    if date_from is None or date_to is None:
        return jsonify({'error': 'Invalid date format'}), 400

    patterns = _RUN_DOWN_SUPPLIER_PATTERNS.get(service_key)
    if patterns is None and service_key != 'OPTIONAL':
        return jsonify({'error': 'Unknown service type'}), 400

    # Restrict to suppliers that actually have requests within the selected
    # date range, using the same matching logic as the results modal so the
    # dropdown and the displayed requests always agree.
    match_ids, match_names = _run_down_supplier_matches(service_key, date_from, date_to)

    if not match_ids and not match_names:
        suppliers = []
    else:
        q = Supplier.query.filter(Supplier.is_active == True)
        if patterns:
            q = q.filter(_or(*[Supplier.supplier_type.ilike(f'%{p}%') for p in patterns]))
        if query:
            q = q.filter(Supplier.name.ilike(f'%{query}%'))

        range_conds = []
        if match_ids:
            range_conds.append(Supplier.id.in_(match_ids))
        if match_names:
            range_conds.append(_func.lower(Supplier.name).in_(match_names))
        q = q.filter(_or(*range_conds))

        q = q.order_by(Supplier.name).limit(500)

        suppliers = [
            {'id': s.id, 'name': s.name, 'city': s.city or '', 'country': s.country or ''}
            for s in q.all()
        ]
    return jsonify({
        'service': service_key,
        'suppliers': suppliers,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
    })


@inbound_bp.route('/run-down/supplier-requests')
@login_required
def run_down_supplier_requests():
    """JSON: inbound requests for a supplier within a service and date range."""
    service_key = request.args.get('service', '').upper()
    supplier_id = request.args.get('supplier_id', type=int)
    date_from, date_to = _parse_run_down_dates()
    if date_from is None or date_to is None:
        return jsonify({'error': 'Invalid date format'}), 400
    if not supplier_id:
        return jsonify({'error': 'supplier_id is required'}), 400
    if service_key not in {s['key'] for s in RUN_DOWN_SERVICES}:
        return jsonify({'error': 'Unknown service type'}), 400

    supplier = Supplier.query.get_or_404(supplier_id)
    rows = _fetch_run_down_supplier_requests(service_key, supplier, date_from, date_to)
    service_label = next((s['label'] for s in RUN_DOWN_SERVICES if s['key'] == service_key), service_key)

    return jsonify({
        'service': service_key,
        'service_label': service_label,
        'supplier': {'id': supplier.id, 'name': supplier.name},
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'date_from_display': date_from.strftime('%d %b %Y'),
        'date_to_display': date_to.strftime('%d %b %Y'),
        'requests': rows,
        'total': len(rows),
    })

@inbound_bp.route('/run-down/agents')
@login_required
def run_down_agents():
    """JSON: searchable agents (customers) for the Agent Run Down."""
    query = request.args.get('q', '').strip()
    date_from, date_to = _parse_run_down_dates()
    if date_from is None or date_to is None:
        return jsonify({'error': 'Invalid date format'}), 400

    # Only agents (customers) that have at least one request overlapping the
    # selected range — same overlap test used by run_down_agent_requests.
    overlap_customer_ids = (
        db.session.query(InboundRequest.customer_id)
        .filter(
            InboundRequest.from_date <= date_to,
            InboundRequest.to_date >= date_from,
        )
        .distinct()
    )

    q = Customer.query.filter(Customer.id.in_(overlap_customer_ids))

    if query:
        q = q.filter(
            (Customer.first_name.ilike(f'%{query}%')) |
            (Customer.last_name.ilike(f'%{query}%')) |
            (Customer.company_name.ilike(f'%{query}%'))
        )

    q = q.order_by(Customer.first_name, Customer.last_name).limit(500)
    customers = q.all()

    agents = []
    for c in customers:
        name = c.name if c.name else (c.company_name or 'Unknown')
        agents.append({'name': name})

    return jsonify({
        'agents': agents,
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
    })


@inbound_bp.route('/run-down/agent-requests')
@login_required
def run_down_agent_requests():
    """JSON: inbound requests for a customer (agent) within a date range."""
    from sqlalchemy import and_

    agent_name = request.args.get('agent', '').strip()
    date_from, date_to = _parse_run_down_dates()
    if date_from is None or date_to is None:
        return jsonify({'error': 'Invalid date format'}), 400
    if not agent_name:
        return jsonify({'error': 'agent is required'}), 400

    customer = None
    customers = Customer.query.all()
    for c in customers:
        customer_full_name = c.name if c.name else (c.company_name or '')
        if customer_full_name.lower() == agent_name.lower():
            customer = c
            break

    if not customer:
        return jsonify({
            'agent': agent_name,
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d'),
            'date_from_display': date_from.strftime('%d %b %Y'),
            'date_to_display': date_to.strftime('%d %b %Y'),
            'requests': [],
            'total': 0,
        })

    requests_query = (
        InboundRequest.query
        .filter(
            InboundRequest.customer_id == customer.id,
            and_(
                InboundRequest.from_date <= date_to,
                InboundRequest.to_date >= date_from,
            ),
        )
        .all()
    )

    rows = []
    for req in requests_query:
        customer_type = customer.customer_type or 'Direct'

        rows.append({
            'request_id': req.id,
            'request_number': req.request_number,
            'contact_name': req.contact_name or '—',
            'group_name': req.agent_ref or '—',
            'pax': req.pax or 0,
            'nationality': req.nationality or '—',
            'from_date': req.from_date.strftime('%d %b %Y') if req.from_date else '—',
            'to_date': req.to_date.strftime('%d %b %Y') if req.to_date else '—',
            'type': customer_type,
            'status': req.status or 'REQUEST',
            'view_url': url_for('inbound.view_request', id=req.id),
        })

    return jsonify({
        'agent': agent_name,
        'agent_type': customer.customer_type or 'Direct',
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d'),
        'date_from_display': date_from.strftime('%d %b %Y'),
        'date_to_display': date_to.strftime('%d %b %Y'),
        'requests': rows,
        'total': len(rows),
    })

@inbound_bp.route('/api/run-down-data')

def api_run_down_data():
    """API endpoint for run-down plan data"""
    from app.models.customer import Customer
    from sqlalchemy import and_

    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    status_filter = request.args.get('status', '')
    booking_filter = request.args.get('booking', '')

    # Parse dates
    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=7)

        if date_to_str:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date() + timedelta(days=30)
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400

    # Build query - join ServiceItem with Booking and Customer
    query = db.session.query(
        ServiceItem.start_date.label('service_date'),
        ServiceItem.service_type,
        ServiceItem.description,
        ServiceItem.status.label('service_status'),
        ServiceItem.amount,
        Booking.reference_number.label('booking_number'),
        Booking.status.label('booking_status'),
        Booking.id.label('booking_id'),
        Customer.first_name,
        Customer.last_name,
        Customer.company_name,
        InboundRequest.pax,
        InboundRequest.contact_name,
        InboundRequest.id.label('request_id')
    ).join(
        Booking, ServiceItem.booking_id == Booking.id
    ).outerjoin(
        Customer, Booking.customer_id == Customer.id
    ).outerjoin(
        InboundRequest, Booking.id == InboundRequest.booking_id
    ).filter(
        Booking.user_id == 1
    ).filter(
        and_(
            ServiceItem.start_date >= date_from,
            ServiceItem.start_date <= date_to
        )
    )

    # Apply status filter
    if status_filter:
        query = query.filter(Booking.status == status_filter)

    # Apply booking number filter
    if booking_filter:
        query = query.filter(Booking.reference_number.contains(booking_filter))

    # Order by date
    query = query.order_by(ServiceItem.start_date, Booking.reference_number)

    # Execute query
    results = query.all()

    # Format results by date
    run_down_data = {}
    for row in results:
        service_date = row.service_date.strftime('%Y-%m-%d')

        # Initialize date bucket if not exists
        if service_date not in run_down_data:
            run_down_data[service_date] = {
                'date': service_date,
                'date_formatted': row.service_date.strftime('%A, %B %d, %Y'),
                'services': []
            }

        # Build guest name
        if row.first_name and row.last_name:
            guest_name = f"{row.first_name} {row.last_name}"
        elif row.company_name:
            guest_name = row.company_name
        elif row.contact_name:
            guest_name = row.contact_name
        else:
            guest_name = "TBA"

        # Add service to date bucket
        service_data = {
            'booking_number': row.booking_number,
            'booking_id': row.booking_id,
            'request_id': row.request_id,
            'guest_name': guest_name,
            'pax': row.pax or 1,
            'service_type': row.service_type,
            'description': row.description or f"{row.service_type} Service",
            'amount': row.amount or 0,
            'status': row.booking_status,
            'status_color': get_status_color(row.booking_status),
            'service_status': row.service_status
        }

        run_down_data[service_date]['services'].append(service_data)

    # Convert to sorted list
    sorted_data = sorted(run_down_data.values(), key=lambda x: x['date'])

    return jsonify({
        'success': True,
        'data': sorted_data,
        'total_days': len(sorted_data),
        'date_from': date_from.strftime('%Y-%m-%d'),
        'date_to': date_to.strftime('%Y-%m-%d')
    })

@inbound_bp.route('/run-down-export-excel')
@login_required
def run_down_export_excel():
    """Export run-down plan to Excel"""
    from app.models.customer import Customer
    from sqlalchemy import and_
    import io
    from flask import send_file

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Excel export requires openpyxl package', 'error')
        return redirect(url_for('inbound.run_down_plan'))

    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    status_filter = request.args.get('status', '')
    booking_filter = request.args.get('booking', '')

    # Parse dates
    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=7)

        if date_to_str:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date() + timedelta(days=30)
    except ValueError:
        flash('Invalid date format', 'error')
        return redirect(url_for('inbound.run_down_plan'))

    # Build query
    query = db.session.query(
        ServiceItem.start_date.label('service_date'),
        ServiceItem.service_type,
        ServiceItem.description,
        ServiceItem.status.label('service_status'),
        ServiceItem.amount,
        Booking.reference_number.label('booking_number'),
        Booking.status.label('booking_status'),
        Customer.first_name,
        Customer.last_name,
        Customer.company_name,
        InboundRequest.pax,
        InboundRequest.contact_name
    ).join(
        Booking, ServiceItem.booking_id == Booking.id
    ).outerjoin(
        Customer, Booking.customer_id == Customer.id
    ).outerjoin(
        InboundRequest, Booking.id == InboundRequest.booking_id
    ).filter(
        Booking.user_id == 1
    ).filter(
        and_(
            ServiceItem.start_date >= date_from,
            ServiceItem.start_date <= date_to
        )
    )

    if status_filter:
        query = query.filter(Booking.status == status_filter)
    if booking_filter:
        query = query.filter(Booking.reference_number.contains(booking_filter))

    query = query.order_by(ServiceItem.start_date, Booking.reference_number)
    results = query.all()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = cast(Any, wb.active)
    ws.title = "Run-Down Plan"

    # Header styling
    header_fill = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
    header_font = Font(bold=True, color="000000", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = f"Run-Down Plan: {date_from.strftime('%B %d, %Y')} - {date_to.strftime('%B %d, %Y')}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")

    # Headers
    headers = ['Date', 'Booking #', 'Guest / Group', 'Pax', 'Service Type', 'Description', 'Amount', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    row = 4
    for result in results:
        # Guest name
        if result.first_name and result.last_name:
            guest_name = f"{result.first_name} {result.last_name}"
        elif result.company_name:
            guest_name = result.company_name
        elif result.contact_name:
            guest_name = result.contact_name
        else:
            guest_name = "TBA"

        ws.cell(row=row, column=1, value=result.service_date.strftime('%Y-%m-%d')).border = thin_border
        ws.cell(row=row, column=2, value=result.booking_number).border = thin_border
        ws.cell(row=row, column=3, value=guest_name).border = thin_border
        ws.cell(row=row, column=4, value=result.pax or 1).border = thin_border
        ws.cell(row=row, column=5, value=result.service_type).border = thin_border
        ws.cell(row=row, column=6, value=result.description or f"{result.service_type} Service").border = thin_border
        ws.cell(row=row, column=7, value=result.amount or 0).border = thin_border
        ws.cell(row=row, column=7, value=f"${result.amount or 0:.2f}").border = thin_border
        ws.cell(row=row, column=8, value=result.booking_status).border = thin_border

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"RunDown_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@inbound_bp.route('/run-down-export-pdf')
@login_required
def run_down_export_pdf():
    """Export run-down plan to PDF"""
    from app.models.customer import Customer
    from sqlalchemy import and_

    # Get filter parameters
    date_from_str = request.args.get('date_from')
    date_to_str = request.args.get('date_to')
    status_filter = request.args.get('status', '')
    booking_filter = request.args.get('booking', '')

    # Parse dates
    try:
        if date_from_str:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        else:
            date_from = datetime.now().date() - timedelta(days=7)

        if date_to_str:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
        else:
            date_to = datetime.now().date() + timedelta(days=30)
    except ValueError:
        flash('Invalid date format', 'error')
        return redirect(url_for('inbound.run_down_plan'))

    # Build query
    query = db.session.query(
        ServiceItem.start_date.label('service_date'),
        ServiceItem.service_type,
        ServiceItem.description,
        ServiceItem.amount,
        Booking.reference_number.label('booking_number'),
        Booking.status.label('booking_status'),
        Customer.first_name,
        Customer.last_name,
        Customer.company_name,
        InboundRequest.pax,
        InboundRequest.contact_name
    ).join(
        Booking, ServiceItem.booking_id == Booking.id
    ).outerjoin(
        Customer, Booking.customer_id == Customer.id
    ).outerjoin(
        InboundRequest, Booking.id == InboundRequest.booking_id
    ).filter(
        Booking.user_id == 1
    ).filter(
        and_(
            ServiceItem.start_date >= date_from,
            ServiceItem.start_date <= date_to
        )
    )

    if status_filter:
        query = query.filter(Booking.status == status_filter)
    if booking_filter:
        query = query.filter(Booking.reference_number.contains(booking_filter))

    query = query.order_by(ServiceItem.start_date, Booking.reference_number)
    results = query.all()

    # Group by date
    run_down_data = {}
    for row in results:
        service_date = row.service_date.strftime('%Y-%m-%d')
        if service_date not in run_down_data:
            run_down_data[service_date] = {
                'date': service_date,
                'date_formatted': row.service_date.strftime('%A, %B %d, %Y'),
                'services': []
            }

        if row.first_name and row.last_name:
            guest_name = f"{row.first_name} {row.last_name}"
        elif row.company_name:
            guest_name = row.company_name
        elif row.contact_name:
            guest_name = row.contact_name
        else:
            guest_name = "TBA"

        run_down_data[service_date]['services'].append({
            'booking_number': row.booking_number,
            'guest_name': guest_name,
            'pax': row.pax or 1,
            'service_type': row.service_type,
            'description': row.description or f"{row.service_type} Service",
            'amount': row.amount or 0,
            'status': row.booking_status,
            'status_color': get_status_color(row.booking_status)
        })

    sorted_data = sorted(run_down_data.values(), key=lambda x: x['date'])

    # Calculate total services
    total_services = sum(len(day['services']) for day in sorted_data)

    # Render PDF template
    return render_template('inbound/run_down_pdf.html',
                         data=sorted_data,
                         date_from=date_from,
                         date_to=date_to,
                         total_days=len(sorted_data),
                         total_services=total_services,
                         current_time=datetime.now())


@inbound_bp.route('/wizard/step1', methods=['GET', 'POST'])
@login_required
def wizard_step1():
    """Wizard Step 1: Arrival & Departure Batches"""
    from flask import session

    if request.method == 'POST':
        # Parse arrival batches
        arrivals = []
        departures = []

        # Parse arrivals[INDEX][FIELD] format
        arrival_indices = set()
        for key in request.form.keys():
            if key.startswith('arrivals['):
                index = key.split('[')[1].split(']')[0]
                arrival_indices.add(index)

        for index in arrival_indices:
            arrival_data = {
                'point': request.form.get(f'arrivals[{index}][point]'),
                'pax': int(request.form.get(f'arrivals[{index}][pax]', 0)),
                'date': request.form.get(f'arrivals[{index}][date]'),
                'time': request.form.get(f'arrivals[{index}][time]', ''),
                'reference': request.form.get(f'arrivals[{index}][reference]', ''),
                'driver': request.form.get(f'arrivals[{index}][driver]', ''),
                'vehicle': request.form.get(f'arrivals[{index}][vehicle]', '')
            }
            arrivals.append(arrival_data)

        # Parse departures[INDEX][FIELD] format
        departure_indices = set()
        for key in request.form.keys():
            if key.startswith('departures['):
                index = key.split('[')[1].split(']')[0]
                departure_indices.add(index)

        for index in departure_indices:
            departure_data = {
                'point': request.form.get(f'departures[{index}][point]'),
                'pax': int(request.form.get(f'departures[{index}][pax]', 0)),
                'date': request.form.get(f'departures[{index}][date]'),
                'time': request.form.get(f'departures[{index}][time]', ''),
                'reference': request.form.get(f'departures[{index}][reference]', ''),
                'driver': request.form.get(f'departures[{index}][driver]', ''),
                'vehicle': request.form.get(f'departures[{index}][vehicle]', '')
            }
            departures.append(departure_data)

        # Calculate date range from first arrival to last departure
        all_dates = []
        for arrival in arrivals:
            if arrival['date']:
                all_dates.append(datetime.strptime(arrival['date'], '%Y-%m-%d').date())
        for departure in departures:
            if departure['date']:
                all_dates.append(datetime.strptime(departure['date'], '%Y-%m-%d').date())

        if all_dates:
            from_date = min(all_dates)
            to_date = max(all_dates)
            no_of_days = (to_date - from_date).days + 1
        else:
            from_date = datetime.now().date()
            to_date = from_date + timedelta(days=1)
            no_of_days = 1

        # Get customer_id if selected, otherwise use contact_name
        customer_id = request.form.get('customer_id', '')

        # Store wizard data in session
        session['wizard_data'] = {
            # Arrival/Departure batches
            'arrivals': arrivals,
            'departures': departures,

            # Contact & Group info
            'customer_id': customer_id if customer_id else None,
            'contact_name': request.form.get('contact_name'),
            'agent_ref': request.form.get('agent_ref', ''),
            'customer_type': request.form.get('customer_type', 'AGENCY'),
            'nationality': request.form.get('nationality'),
            'pax': int(request.form.get('pax', 1)),
            'special_note': request.form.get('special_note', ''),

            # Calculated fields
            'from_date': from_date.strftime('%Y-%m-%d'),
            'to_date': to_date.strftime('%Y-%m-%d'),
            'no_of_days': no_of_days,

            # Initialize service collections
            'hotels': [],
            'transports': [],
            'meals': [],
            'guides': []
        }
        session.modified = True
        return redirect(url_for('inbound.wizard_step2'))

    # GET request - pass any existing wizard data to template
    wizard_data = session.get('wizard_data', {})
    return render_template('inbound/wizard_step1.html', wizard_data=wizard_data)


@inbound_bp.route('/wizard/step2', methods=['GET', 'POST'])
@login_required
def wizard_step2():
    """Wizard Step 2: Add All Services"""
    from flask import session

    if 'wizard_data' not in session:
        flash('Please start from step 1', 'warning')
        return redirect(url_for('inbound.wizard_step1'))

    if request.method == 'POST':
        # Parse services from form (handles both 2-level and 3-level nested structures)
        services_data = {}
        for key, value in request.form.items():
            if key.startswith('services['):
                parts = key.split('[')
                index = parts[1].split(']')[0]

                if index not in services_data:
                    services_data[index] = {}

                # Check if this is a nested structure like services[0][rooms][0][field]
                if len(parts) > 3 and 'rooms' in key:
                    # This is a hotel room field: services[INDEX][rooms][ROOM_INDEX][FIELD]
                    room_index = parts[3].split(']')[0]
                    field_name = parts[4].split(']')[0]

                    if 'rooms' not in services_data[index]:
                        services_data[index]['rooms'] = {}
                    if room_index not in services_data[index]['rooms']:
                        services_data[index]['rooms'][room_index] = {}

                    services_data[index]['rooms'][room_index][field_name] = value
                else:
                    # Simple field: services[INDEX][FIELD]
                    field = parts[2].split(']')[0]
                    services_data[index][field] = value

        # Validate at least one service
        if not services_data:
            flash('Please add at least one service before continuing to review', 'warning')
            wizard_data = session.get('wizard_data', {})
            return render_template('inbound/wizard_step2.html', wizard_data=wizard_data)

        # Store services in session
        session['wizard_data']['services'] = services_data
        session.modified = True

        # Redirect to step 3 (Review)
        return redirect(url_for('inbound.wizard_step3'))

    # GET request - pass wizard data to template
    wizard_data = session.get('wizard_data', {})
    return render_template('inbound/wizard_step2.html', wizard_data=wizard_data)


@inbound_bp.route('/wizard/step3', methods=['GET', 'POST'])
@login_required
def wizard_step3():
    """Wizard Step 3: Review & Create"""
    from flask import session

    if 'wizard_data' not in session:
        flash('Please start from step 1', 'warning')
        return redirect(url_for('inbound.wizard_step1'))

    if request.method == 'POST':
        wizard_data = session['wizard_data']

        # Helper functions
        def safe_int(value, default=0):
            """Safely convert to int, handling empty strings"""
            if not value or value == '':
                return default
            try:
                return int(value)
            except (ValueError, TypeError):
                return default

        def safe_float(value, default=0.0):
            """Safely convert to float, handling empty strings"""
            if not value or value == '':
                return default
            try:
                return float(value)
            except (ValueError, TypeError):
                return default

        def date_range(start_date, end_date):
            """Generate list of dates between start and end (inclusive for start, exclusive for end for hotel nights)"""
            from datetime import timedelta
            dates = []
            current = start_date
            while current < end_date:
                dates.append(current)
                current += timedelta(days=1)
            return dates

        # Get services from session
        services_data = wizard_data.get('services', {})

        # Validate that we have at least one service
        if not services_data:
            flash('Please add at least one service before creating the tour', 'error')
            return redirect(url_for('inbound.wizard_step2'))

        # Start transaction
        try:
            from datetime import timedelta
            from_date = datetime.strptime(wizard_data['from_date'], '%Y-%m-%d').date()
            to_date = datetime.strptime(wizard_data['to_date'], '%Y-%m-%d').date()

            # Create InboundRequest
            request_obj = InboundRequest(
                request_number=InboundRequest.generate_request_number(from_date),
                from_date=from_date,
                to_date=to_date,
                no_of_days=wizard_data['no_of_days'],
                customer_type=wizard_data['customer_type'],
                contact_name=wizard_data['contact_name'],
                customer_id=wizard_data.get('customer_id'),
                agent_ref=wizard_data.get('agent_ref', ''),
                nationality=wizard_data['nationality'],
                pax=wizard_data['pax'],
                special_note=wizard_data.get('special_note', ''),
                user_id=1,
                status=STATUS_REQUEST
            )

            db.session.add(request_obj)
            db.session.flush()  # Get the ID

            # Create arrival transport if driver/vehicle specified
            if wizard_data.get('arrival_driver') or wizard_data.get('arrival_vehicle'):
                arrival_time_str = wizard_data.get('arrival_time')
                arrival_time = None
                if arrival_time_str:
                    try:
                        arrival_time = datetime.strptime(arrival_time_str, '%H:%M').time()
                    except:
                        pass

                arrival_transport = InboundTransport(
                    request_id=request_obj.id,
                    date=from_date,
                    vehicle_type=wizard_data.get('arrival_vehicle'),
                    driver_name=wizard_data.get('arrival_driver'),
                    pickup_location=wizard_data.get('arrival_point', ''),
                    dropoff_location='Hotel',  # Default dropoff
                    pickup_time=arrival_time,
                    is_airport_transfer=True,
                    is_arrival=True,
                    cost=0.0,
                    currency='USD'
                )
                db.session.add(arrival_transport)

            # Create departure transport if driver/vehicle specified
            if wizard_data.get('departure_driver') or wizard_data.get('departure_vehicle'):
                departure_time_str = wizard_data.get('departure_time')
                departure_time = None
                if departure_time_str:
                    try:
                        departure_time = datetime.strptime(departure_time_str, '%H:%M').time()
                    except:
                        pass

                departure_transport = InboundTransport(
                    request_id=request_obj.id,
                    date=to_date,
                    vehicle_type=wizard_data.get('departure_vehicle'),
                    driver_name=wizard_data.get('departure_driver'),
                    pickup_location='Hotel',  # Default pickup
                    dropoff_location=wizard_data.get('departure_point', ''),
                    pickup_time=departure_time,
                    is_airport_transfer=True,
                    is_departure=True,
                    cost=0.0,
                    currency='USD'
                )
                db.session.add(departure_transport)

            # Track itinerary rows by date to merge services on same dates
            itinerary_by_date = {}

            # Process each service and generate itinerary rows
            for index in sorted(services_data.keys(), key=int):
                service = services_data[index]
                service_type = service.get('type')

                if service_type == 'hotel':
                    # Hotel: create rows for each night with inherited rooming
                    check_in = datetime.strptime(service['check_in_date'], '%Y-%m-%d').date()
                    check_out = datetime.strptime(service['check_out_date'], '%Y-%m-%d').date()

                    # Validate hotel dates
                    if check_out <= check_in:
                        raise ValueError("Hotel check-out date must be after check-in date")

                    hotel_name = service.get('hotel_name', '')
                    location = service.get('location', '')

                    # Room distribution (inherited across all nights)
                    single = safe_int(service.get('single_rooms'))
                    double = safe_int(service.get('double_rooms'))
                    triple = safe_int(service.get('triple_rooms'))
                    other = safe_int(service.get('other_rooms'))

                    cost = safe_float(service.get('cost'))
                    cost_unit = service.get('cost_unit', COST_UNIT_PER_PERSON)

                    # Generate description
                    desc = f"Hotel: {hotel_name or 'TBD'}"
                    if location:
                        desc += f" ({location})"

                    # Create itinerary row for each night
                    for night_date in date_range(check_in, check_out):
                        if night_date not in itinerary_by_date:
                            itinerary_by_date[night_date] = {
                                'date': night_date,
                                'description': desc,
                                'base_cost': cost,
                                'cost_unit': cost_unit,
                                'flag_hotel': True,
                                'hotel_single_rooms': single,
                                'hotel_double_rooms': double,
                                'hotel_triple_rooms': triple,
                                'hotel_other_rooms': other,
                                'flag_transport': False,
                                'flag_meal': False,
                                'flag_guide': False
                            }
                        else:
                            # Merge with existing
                            itinerary_by_date[night_date]['description'] += f" | {desc}"
                            itinerary_by_date[night_date]['base_cost'] += cost
                            itinerary_by_date[night_date]['flag_hotel'] = True
                            itinerary_by_date[night_date]['hotel_single_rooms'] = single
                            itinerary_by_date[night_date]['hotel_double_rooms'] = double
                            itinerary_by_date[night_date]['hotel_triple_rooms'] = triple
                            itinerary_by_date[night_date]['hotel_other_rooms'] = other

                elif service_type == 'transport':
                    # Transport: single date
                    transport_date = datetime.strptime(service['date'], '%Y-%m-%d').date()
                    pickup = service.get('pickup_location', 'TBD')
                    dropoff = service.get('dropoff_location', 'TBD')
                    vehicle = service.get('vehicle_type', '')
                    cost = safe_float(service.get('cost'))
                    cost_unit = service.get('cost_unit', COST_UNIT_PER_GROUP)

                    desc = f"Transport: {pickup} → {dropoff}"
                    if vehicle:
                        desc += f" ({vehicle})"

                    if transport_date not in itinerary_by_date:
                        itinerary_by_date[transport_date] = {
                            'date': transport_date,
                            'description': desc,
                            'base_cost': cost,
                            'cost_unit': cost_unit,
                            'flag_hotel': False,
                            'hotel_single_rooms': 0,
                            'hotel_double_rooms': 0,
                            'hotel_triple_rooms': 0,
                            'hotel_other_rooms': 0,
                            'flag_transport': True,
                            'flag_meal': False,
                            'flag_guide': False
                        }
                    else:
                        # Merge with existing
                        itinerary_by_date[transport_date]['description'] += f" | {desc}"
                        itinerary_by_date[transport_date]['base_cost'] += cost
                        itinerary_by_date[transport_date]['flag_transport'] = True

                elif service_type == 'meal':
                    # Meal: single date only
                    meal_from = datetime.strptime(service['from_date'], '%Y-%m-%d').date()
                    meal_type = service.get('meal_type', 'Meal')
                    restaurant = service.get('restaurant', 'TBD')
                    cost = safe_float(service.get('cost'))
                    cost_unit = service.get('cost_unit', COST_UNIT_PER_PERSON)

                    desc = f"{meal_type} at {restaurant}"

                    # Single date only
                    meal_date = meal_from
                    if meal_date not in itinerary_by_date:
                        itinerary_by_date[meal_date] = {
                            'date': meal_date,
                            'description': desc,
                            'base_cost': cost,
                            'cost_unit': cost_unit,
                            'flag_hotel': False,
                            'hotel_single_rooms': 0,
                            'hotel_double_rooms': 0,
                            'hotel_triple_rooms': 0,
                            'hotel_other_rooms': 0,
                            'flag_transport': False,
                            'flag_meal': True,
                            'flag_guide': False
                        }
                    else:
                        itinerary_by_date[meal_date]['description'] += f" | {desc}"
                        itinerary_by_date[meal_date]['base_cost'] += cost
                        itinerary_by_date[meal_date]['flag_meal'] = True

                elif service_type == 'guide':
                    # Guide: single date or date range (optional TO date)
                    guide_from = datetime.strptime(service['from_date'], '%Y-%m-%d').date()
                    guide_to_str = service.get('to_date', '')
                    guide_to = datetime.strptime(guide_to_str, '%Y-%m-%d').date() if guide_to_str else guide_from
                    guide_type = service.get('guide_type', 'Guide Service')
                    language = service.get('language', '')
                    cost = safe_float(service.get('cost'))
                    cost_unit = service.get('cost_unit', COST_UNIT_PER_GROUP)

                    desc = f"{guide_type}"
                    if language:
                        desc += f" ({language})"

                    # Create row for each day in range (inclusive)
                    guide_to_inclusive = guide_to + timedelta(days=1)
                    for guide_date in date_range(guide_from, guide_to_inclusive):
                        if guide_date not in itinerary_by_date:
                            itinerary_by_date[guide_date] = {
                                'date': guide_date,
                                'description': desc,
                                'base_cost': cost,
                                'cost_unit': cost_unit,
                                'flag_hotel': False,
                                'hotel_single_rooms': 0,
                                'hotel_double_rooms': 0,
                                'hotel_triple_rooms': 0,
                                'hotel_other_rooms': 0,
                                'flag_transport': False,
                                'flag_meal': False,
                                'flag_guide': True
                            }
                        else:
                            # Merge with existing
                            itinerary_by_date[guide_date]['description'] += f" | {desc}"
                            itinerary_by_date[guide_date]['base_cost'] += cost
                            itinerary_by_date[guide_date]['flag_guide'] = True

            # Create ItineraryRow objects from merged data
            for row_date in sorted(itinerary_by_date.keys()):
                row_data = itinerary_by_date[row_date]

                row = ItineraryRow(
                    request_id=request_obj.id,
                    date=row_data['date'],
                    description=row_data['description'],
                    base_cost=row_data['base_cost'],
                    cost_unit=row_data['cost_unit'],
                    currency='USD',
                    flag_hotel=row_data['flag_hotel'],
                    flag_transport=row_data['flag_transport'],
                    flag_meal=row_data['flag_meal'],
                    flag_guide=row_data['flag_guide'],
                    hotel_single_rooms=row_data['hotel_single_rooms'],
                    hotel_double_rooms=row_data['hotel_double_rooms'],
                    hotel_triple_rooms=row_data['hotel_triple_rooms'],
                    hotel_other_rooms=row_data['hotel_other_rooms']
                )

                db.session.add(row)

            # Create service records directly from service data
            for service_idx, service in services_data.items():
                service_type = service.get('type')

                if service_type == 'transport':
                    # Create InboundTransport record with arrival/departure flags
                    transport_date = datetime.strptime(service['date'], '%Y-%m-%d').date()
                    from_date_str = service.get('from_date')
                    to_date_str = service.get('to_date')

                    from_date = datetime.strptime(from_date_str, '%Y-%m-%d').date() if from_date_str else transport_date
                    to_date = datetime.strptime(to_date_str, '%Y-%m-%d').date() if to_date_str else transport_date

                    pickup_time_str = service.get('time')
                    pickup_time = None
                    if pickup_time_str:
                        try:
                            pickup_time = datetime.strptime(pickup_time_str, '%H:%M').time()
                        except:
                            pass

                    transport = InboundTransport(
                        request_id=request_obj.id,
                        date=from_date,
                        end_date=to_date if to_date != from_date else None,
                        vehicle_type=service.get('vehicle_type'),
                        driver_name=service.get('driver_name'),
                        pickup_location=service.get('pickup_location'),
                        dropoff_location=service.get('dropoff_location'),
                        pickup_time=pickup_time,
                        is_arrival=service.get('is_arrival') == '1',
                        is_departure=service.get('is_departure') == '1',
                        cost=safe_float(service.get('cost', 0)),
                        currency='USD'
                    )
                    db.session.add(transport)

                elif service_type == 'meal':
                    # Create InboundMeal record (single date only)
                    meal_date = datetime.strptime(service['from_date'], '%Y-%m-%d').date()

                    meal = InboundMeal(
                        request_id=request_obj.id,
                        date=meal_date,
                        end_date=None,
                        meal_type=service.get('meal_type'),
                        restaurant=service.get('restaurant'),
                        cost_per_person=safe_float(service.get('cost', 0)),
                        currency='USD'
                    )
                    db.session.add(meal)

                elif service_type == 'guide':
                    # Create InboundGuide records (single date or date range)
                    guide_from = datetime.strptime(service['from_date'], '%Y-%m-%d').date()
                    guide_to_str = service.get('to_date', '')
                    guide_to = datetime.strptime(guide_to_str, '%Y-%m-%d').date() if guide_to_str else guide_from

                    guide = InboundGuide(
                        request_id=request_obj.id,
                        date=guide_from,
                        end_date=guide_to if guide_to != guide_from else None,
                        service_type=service.get('guide_type'),
                        language=service.get('language', 'English'),
                        cost=safe_float(service.get('cost', 0)),
                        currency='USD'
                    )
                    db.session.add(guide)

                elif service_type == 'hotel':
                    # Create InboundHotel record
                    check_in = datetime.strptime(service['check_in_date'], '%Y-%m-%d').date()
                    check_out = datetime.strptime(service['check_out_date'], '%Y-%m-%d').date()
                    nights = (check_out - check_in).days

                    hotel = InboundHotel(
                        request_id=request_obj.id,
                        hotel_name=service.get('hotel_name'),
                        location=service.get('location'),
                        check_in_date=check_in,
                        check_out_date=check_out,
                        nights=nights,
                        total_cost=safe_float(service.get('cost', 0)),
                        currency='USD'
                    )
                    db.session.add(hotel)

            # Calculate total
            db.session.flush()
            request_obj.calculate_total()

            # Commit transaction
            db.session.commit()

            # Clear wizard data from session
            session.pop('wizard_data', None)

            flash(f'Tour itinerary {request_obj.request_number} created successfully!', 'success')
            return redirect(url_for('inbound.view_request', id=request_obj.id))

        except Exception as e:
            db.session.rollback()
            flash(f'Error creating itinerary: {str(e)}', 'error')
            return redirect(url_for('inbound.wizard_step2'))

    # GET request - show review page
    wizard_data = session.get('wizard_data', {})
    return render_template('inbound/wizard_step3.html', wizard_data=wizard_data)
@inbound_bp.route('/api/<int:request_id>/export-expense-report')

def api_export_expense_report(request_id):
    """Export cash expense report in Windows of Jordan Excel format"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    import os
    import tempfile

    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    try:
        wb = Workbook()
        ws = cast(Any, wb.active)
        ws.title = "Sheet1"

        # Header
        ws['B2'] = 'Windows of Jordan'
        ws['B2'].font = Font(size=16, bold=True)
        ws['B3'] = ' Actual Expense Sheet'
        ws['B3'].font = Font(size=22)

        # File info
        ws['B5'] = request_obj.request_number
        ws['E5'] = 'Date'
        ws['F5'] = request_obj.from_date if request_obj.from_date else datetime.now()
        ws['F5'].number_format = 'DD-MMM-YY'

        ws['B8'] = f'File Expense {request_obj.agent or "N/A"}'
        ws['E8'] = 'Ref:'
        ws['F8'] = request_obj.contact_name or 'N/A'
        ws['E9'] = 'Pax:'
        ws['F9'] = str(request_obj.pax)

        # Table header
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        ws['B11'] = 'Item'
        ws['C11'] = 'Driver'
        ws['D11'] = 'Cost PP'
        ws['E11'] = 'Pax'
        ws['F11'] = 'Total'

        for cell in ['B11', 'C11', 'D11', 'E11', 'F11']:
            ws[cell].font = Font(bold=True)
            ws[cell].fill = header_fill
            ws[cell].alignment = Alignment(horizontal='center')

        # Add expense items
        row = 12
        for expense in sorted(request_obj.inbound_cash_expenses, key=lambda x: x.date):
            ws[f'B{row}'] = expense.description
            ws[f'C{row}'] = expense.driver_name or '-'
            ws[f'D{row}'] = expense.amount
            ws[f'E{row}'] = request_obj.pax if expense.is_per_person else 1
            ws[f'F{row}'] = f'=SUM(D{row}*E{row})'
            row += 1

        # Totals
        if row > 12:
            ws[f'F{row+1}'] = f'=SUM(F12:F{row-1})'
            ws[f'F{row+1}'].font = Font(bold=True)

            ws[f'D{row+2}'] = 'Advance Payment'
            ws[f'D{row+2}'].font = Font(bold=True)
            ws[f'F{row+2}'] = 0

            ws[f'D{row+3}'] = 'Total'
            ws[f'D{row+3}'].font = Font(bold=True, size=12)
            ws[f'F{row+3}'] = f'=F{row+1}-F{row+2}'
            ws[f'F{row+3}'].font = Font(bold=True, size=12)

            # Signature lines
            ws[f'B{row+7}'] = 'Authorization:…................................................'
            ws[f'D{row+7}'] = 'Guide\\Driver:…............................'

        # Save to temp file
        output_dir = tempfile.gettempdir()
        output_path = os.path.join(output_dir, f'Expense_Report_{request_obj.request_number}.xlsx')
        wb.save(output_path)

        return send_file(
            output_path,
            as_attachment=True,
            download_name=f'Expense_Report_{request_obj.request_number}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error generating expense report: {str(e)}', 'error')
        return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/<int:request_id>/add-cash-expense', methods=['POST'])
@login_required
def add_cash_expense(request_id):
    """Add a cash expense item and create itinerary row"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    try:
        from app.models.inbound import ItineraryRow

        expense_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        description = request.form['description']
        driver_name = request.form.get('driver_name', '')
        amount = float(request.form['amount'])

        # Create the cash expense record
        expense = InboundCashExpense(
            request_id=request_obj.id,
            date=expense_date,
            description=description,
            driver_name=driver_name,
            amount=amount,
            currency='USD',
            is_per_person=False
        )
        db.session.add(expense)
        db.session.flush()

        # Create or update itinerary row for this date
        existing_row = ItineraryRow.query.filter_by(
            request_id=request_obj.id,
            date=expense_date
        ).first()

        if existing_row:
            # Add expense to existing row description
            expense_text = f"Cash: {description} (${amount:.2f})"
            if existing_row.description:
                existing_row.description += f" | {expense_text}"
            else:
                existing_row.description = expense_text
            # Add to base cost
            existing_row.base_cost += amount
        else:
            # Create new itinerary row
            expense_text = f"Cash: {description} (${amount:.2f})"

            itinerary_row = ItineraryRow(
                request_id=request_obj.id,
                date=expense_date,
                description=expense_text,
                base_cost=amount,
                currency='USD'
            )
            db.session.add(itinerary_row)

        db.session.commit()
        flash('Cash expense added to itinerary successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding cash expense: {str(e)}', 'error')

    return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/<int:request_id>/update-cash-expense/<int:expense_id>', methods=['POST'])
@login_required
def update_cash_expense(request_id, expense_id):
    """Update a cash expense item"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    expense = InboundCashExpense.query.get_or_404(expense_id)

    if expense.request_id != request_id:
        abort(403)

    try:
        if 'date' in request.form:
            expense.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        if 'description' in request.form:
            expense.description = request.form['description']
        if 'driver_name' in request.form:
            expense.driver_name = request.form['driver_name']
        if 'amount' in request.form:
            expense.amount = float(request.form['amount'])

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error updating cash expense: {str(e)}', 'error')

    return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/<int:request_id>/delete-cash-expense/<int:expense_id>', methods=['POST'])
@login_required
def delete_cash_expense(request_id, expense_id):
    """Delete a cash expense item"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    expense = InboundCashExpense.query.get_or_404(expense_id)

    if expense.request_id != request_id:
        abort(403)

    try:
        db.session.delete(expense)
        db.session.commit()
        flash('Cash expense deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting cash expense: {str(e)}', 'error')

    return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/<int:request_id>/add-meal', methods=['POST'])
@login_required
def add_meal(request_id):
    """Add a meal item and create itinerary row"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    try:
        from app.models.inbound import ItineraryRow

        meal_date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        meal_type = request.form['meal_type']
        restaurant = request.form.get('restaurant', '')
        location = request.form.get('location', '')

        # Create the meal record
        meal = InboundMeal(
            request_id=request_obj.id,
            date=meal_date,
            meal_type=meal_type,
            restaurant=restaurant,
            location=location,
            cost_per_person=0.0,
            currency='USD',
            status='CONFIRMED'
        )
        db.session.add(meal)
        db.session.flush()

        # Create or update itinerary row for this date
        existing_row = ItineraryRow.query.filter_by(
            request_id=request_obj.id,
            date=meal_date
        ).first()

        if existing_row:
            # Add meal flag to existing row
            existing_row.has_meal = True
            if existing_row.description:
                existing_row.description += f" | {meal_type}"
            else:
                existing_row.description = meal_type
        else:
            # Create new itinerary row
            description = f"{meal_type}"
            if restaurant:
                description += f" at {restaurant}"
            if location:
                description += f" ({location})"

            itinerary_row = ItineraryRow(
                request_id=request_obj.id,
                date=meal_date,
                description=description,
                has_meal=True,
                base_cost=0.0,
                currency='USD'
            )
            db.session.add(itinerary_row)
            db.session.flush()

            # Link meal to itinerary row
            meal.source_itinerary_id = itinerary_row.id

        db.session.commit()
        flash('Meal added to itinerary successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error adding meal: {str(e)}', 'error')

    return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/<int:request_id>/delete-meal/<int:meal_id>', methods=['POST'])
@login_required
def delete_meal(request_id, meal_id):
    """Delete a meal item"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        abort(403)

    meal = InboundMeal.query.get_or_404(meal_id)

    if meal.request_id != request_id:
        abort(403)

    try:
        db.session.delete(meal)
        db.session.commit()
        flash('Meal deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting meal: {str(e)}', 'error')

    return redirect(url_for('inbound.view_request', id=request_id))

@inbound_bp.route('/api/hotels/search')

def api_search_hotels():
    """API endpoint to search hotels for autocomplete"""
    query = request.args.get('query', '').strip()
    limit = request.args.get('limit', 20, type=int)

    # Query distinct hotel names from InboundHotel table
    hotels_query = db.session.query(InboundHotel.hotel_name, InboundHotel.location).filter(
        InboundHotel.hotel_name.isnot(None),
        InboundHotel.hotel_name != ''
    )

    # Apply search filter if query provided
    if query:
        hotels_query = hotels_query.filter(
            db.or_(
                InboundHotel.hotel_name.ilike(f'%{query}%'),
                InboundHotel.location.ilike(f'%{query}%')
            )
        )

    # Get distinct hotel names with their most recent location
    hotels_query = hotels_query.distinct(InboundHotel.hotel_name).order_by(
        InboundHotel.hotel_name
    ).limit(limit)

    hotels = hotels_query.all()

    # Format for Select2
    results = []
    for hotel_name, location in hotels:
        results.append({
            'id': hotel_name,
            'text': f"{hotel_name}" + (f" ({location})" if location else ""),
            'name': hotel_name,
            'location': location or ''
        })

    return jsonify({'results': results})

@inbound_bp.route('/api/<int:request_id>/update-itinerary-bulk', methods=['POST'])
@csrf.exempt

def api_update_itinerary_bulk(request_id):
    """API endpoint to bulk update itinerary rows"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()
        updates = data.get('updates', [])

        for update in updates:
            row_id = update.get('row_id')
            row = ItineraryRow.query.filter_by(id=row_id, request_id=request_id).first()

            if row:
                row.description = update.get('description', '')
                row.restaurant = update.get('restaurant', '')
                row.cash_expense = float(update.get('cash_expense', 0))
                row.comment = update.get('comment', '')

        db.session.commit()

        return jsonify({'success': True, 'message': 'Trip itinerary updated successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/add-itinerary-row', methods=['POST'])
@csrf.exempt

def api_add_itinerary_row(request_id):
    """API endpoint to add a new itinerary row"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        data = request.get_json()

        # Parse date
        date_str = data.get('date')
        if not date_str:
            return jsonify({'success': False, 'message': 'Date is required'}), 400

        date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()

        # Create new itinerary row (for tours and restaurants only)
        new_row = ItineraryRow(
            request_id=request_id,
            date=date_obj,
            description=data.get('description', ''),
            restaurant=data.get('restaurant', ''),
            cash_expense=float(data.get('cash_expense', 0)),
            comment=data.get('comment', '')
        )

        db.session.add(new_row)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Itinerary item added successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/delete-itinerary-row/<int:row_id>', methods=['DELETE'])
@csrf.exempt

def api_delete_itinerary_row(request_id, row_id):
    """API endpoint to delete an itinerary row"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if request_obj.user_id != 1:
        return jsonify({'success': False, 'message': 'Unauthorized'}), 403

    try:
        row = ItineraryRow.query.filter_by(id=row_id, request_id=request_id).first_or_404()

        # Only manually added rows (children created via the "+" button) can be deleted
        is_manual_row = False
        if row.comment:
            try:
                is_manual_row = bool(json.loads(row.comment).get('parent_row_id'))
            except (ValueError, TypeError, AttributeError):
                is_manual_row = False
        if not is_manual_row:
            return jsonify({'success': False, 'message': 'Original itinerary days cannot be deleted'}), 400

        # Delete the row
        db.session.delete(row)
        db.session.commit()

        return jsonify({'success': True, 'message': 'Itinerary item deleted successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

# Supplier Confirmation Status Endpoints
@inbound_bp.route('/api/hotel/<int:hotel_id>/mark-reserved', methods=['POST'])
@csrf.exempt

def api_mark_hotel_reserved(hotel_id):
    """Mark a single hotel as RESERVED (Supplier Confirmed)"""
    try:
        hotel = InboundHotel.query.get_or_404(hotel_id)
        request_obj = InboundRequest.query.get_or_404(hotel.request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Update hotel status to RESERVED
        hotel.status = STATUS_RESERVED
        db.session.commit()

        # Check if all services are now RESERVED, auto-update request to CONFIRMED
        check_and_update_request_status(request_obj.id)

        return jsonify({'success': True, 'message': 'Hotel marked as Supplier Confirmed'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/transport/<int:transport_id>/mark-reserved', methods=['POST'])
@csrf.exempt

def api_mark_transport_reserved(transport_id):
    """Mark a single transport as RESERVED (Supplier Confirmed)"""
    try:
        transport = InboundTransport.query.get_or_404(transport_id)
        request_obj = InboundRequest.query.get_or_404(transport.request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Update transport status to RESERVED
        transport.status = STATUS_RESERVED
        db.session.commit()

        # Check if all services are now RESERVED, auto-update request to CONFIRMED
        check_and_update_request_status(request_obj.id)

        return jsonify({'success': True, 'message': 'Transport marked as Supplier Confirmed'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/confirm-all-hotels', methods=['POST'])
@csrf.exempt

def api_confirm_all_hotels(request_id):
    """Mark ALL hotels in a request as RESERVED (Supplier Confirmed)"""
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Update all hotels to RESERVED
        hotels = InboundHotel.query.filter_by(request_id=request_id).all()
        count = 0
        for hotel in hotels:
            if hotel.status not in [STATUS_RESERVED, STATUS_CONFIRMED]:
                hotel.status = STATUS_RESERVED
                count += 1

        db.session.commit()

        # Check if all services are now RESERVED, auto-update request to CONFIRMED
        check_and_update_request_status(request_id)

        return jsonify({'success': True, 'count': count, 'message': f'{count} hotel(s) marked as Supplier Confirmed'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:hotel_id>', methods=['GET'])
@csrf.exempt
def api_get_hotel(hotel_id):
    """Get hotel details for editing"""
    try:
        hotel = InboundHotel.query.get_or_404(hotel_id)
        request_obj = InboundRequest.query.get_or_404(hotel.request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        return jsonify({
            'success': True,
            'hotel': {
                'id': hotel.id,
                'name': hotel.name or 'Hotel',
                'supplier_name': hotel.supplier_name or '',
                'status': hotel.status or 'REQUEST',
                'total_cost': float(hotel.total_cost) if hotel.total_cost else 0,
                'currency': hotel.currency or 'USD',
                'notes': hotel.notes or '',
                'check_in_date': hotel.check_in_date.strftime('%Y-%m-%d') if hotel.check_in_date else '',
                'check_out_date': hotel.check_out_date.strftime('%Y-%m-%d') if hotel.check_out_date else '',
                'nights': hotel.nights or 0
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/hotel/<int:hotel_id>/update', methods=['POST'])
@csrf.exempt
def api_update_hotel(hotel_id):
    """Update hotel details including check-in/check-out dates"""
    try:
        hotel = InboundHotel.query.get_or_404(hotel_id)
        request_obj = InboundRequest.query.get_or_404(hotel.request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        data = request.get_json() or {}

        # Update hotel fields
        if 'supplier_name' in data:
            hotel.supplier_name = data['supplier_name']
        if 'status' in data:
            hotel.status = data['status']
        if 'total_cost' in data:
            hotel.total_cost = float(data['total_cost']) if data['total_cost'] else 0
        if 'currency' in data:
            hotel.currency = data['currency']
        if 'notes' in data:
            hotel.notes = data['notes']
        if 'voucher_notes' in data:
            hotel.voucher_notes = data['voucher_notes']

        # Update check-in/check-out dates
        if 'check_in_date' in data and data['check_in_date']:
            hotel.check_in_date = datetime.strptime(data['check_in_date'], '%Y-%m-%d').date()
        if 'check_out_date' in data and data['check_out_date']:
            hotel.check_out_date = datetime.strptime(data['check_out_date'], '%Y-%m-%d').date()

            # Recalculate nights if both dates are set
            if hotel.check_in_date and hotel.check_out_date:
                hotel.nights = (hotel.check_out_date - hotel.check_in_date).days

        db.session.commit()

        # Check if all services are confirmed
        check_and_update_request_status(request_obj.id)

        return jsonify({'success': True, 'message': 'Hotel updated successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/meal/<int:meal_id>/update', methods=['POST'])
@csrf.exempt
def api_update_meal(meal_id):
    """Update meal voucher_notes for restaurant voucher"""
    try:
        data = request.get_json() or {}
        req_id = data.get('request_id')
        meal = InboundMeal.query.filter_by(id=meal_id).first()
        if not meal:
            from sqlalchemy import text
            r = db.session.execute(text("SELECT id, request_id FROM inbound_meal WHERE id = :mid"), {"mid": meal_id}).fetchone()
            print(f"[MEAL UPDATE] 404 meal_id={meal_id} direct_sql={r}")
            return jsonify({'success': False, 'message': 'Meal not found'}), 404
        
        if req_id and int(req_id) != meal.request_id:
            return jsonify({'success': False, 'message': 'Meal does not belong to this request'}), 403

        request_obj = InboundRequest.query.get(meal.request_id)
        if not request_obj:
            return jsonify({'success': False, 'message': 'Request not found'}), 404

        # Allow when user_id is 1 or None (legacy/unassigned requests)
        if request_obj.user_id is not None and request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        if 'voucher_notes' in data:
            meal.voucher_notes = data['voucher_notes']

        db.session.commit()

        return jsonify({'success': True, 'message': 'Meal voucher notes saved successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@inbound_bp.route('/api/<int:request_id>/confirm-all-transports', methods=['POST'])
@csrf.exempt

def api_confirm_all_transports(request_id):
    """Mark ALL transports in a request as RESERVED (Supplier Confirmed)"""
    try:
        request_obj = InboundRequest.query.get_or_404(request_id)

        if request_obj.user_id != 1:
            return jsonify({'success': False, 'message': 'Unauthorized'}), 403

        # Update all transports to RESERVED
        transports = InboundTransport.query.filter_by(request_id=request_id).all()
        count = 0
        for transport in transports:
            if transport.status not in [STATUS_RESERVED, STATUS_CONFIRMED]:
                transport.status = STATUS_RESERVED
                count += 1

        db.session.commit()

        # Check if all services are now RESERVED, auto-update request to CONFIRMED
        check_and_update_request_status(request_id)

        return jsonify({'success': True, 'count': count, 'message': f'{count} transport(s) marked as Supplier Confirmed'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

def check_and_update_request_status(request_id):
    """
    Check if all hotels are RESERVED.
    If yes, automatically update InboundRequest status to CONFIRMED (Supplier Confirmed).
    """
    try:
        request_obj = InboundRequest.query.get(request_id)
        if not request_obj:
            return

        # Get all hotels
        hotels = InboundHotel.query.filter_by(request_id=request_id).all()

        # If there are no hotels, don't auto-update
        if not hotels:
            return

        # Check if ALL hotels are RESERVED or CONFIRMED
        all_hotels_confirmed = all(
            hotel.status in [STATUS_RESERVED, STATUS_CONFIRMED] 
            for hotel in hotels
        )

        # If all hotels are confirmed, update request to CONFIRMED (Supplier Confirmed)
        if all_hotels_confirmed:
            if request_obj.status != STATUS_CONFIRMED:
                request_obj.status = STATUS_CONFIRMED
                db.session.commit()
                print(f"✅ All hotels confirmed! InboundRequest {request_id} auto-updated to CONFIRMED (Supplier Confirmed) status")

    except Exception as e:
        print(f"Error checking request status: {e}")
        db.session.rollback()

# ============================================================
# ANALYTICS DASHBOARD - COMPREHENSIVE RUN DOWNS
# ============================================================

@inbound_bp.route('/analytics')
@login_required
def analytics_dashboard():
    """Comprehensive analytics dashboard with search across all services"""
    from app.models.inbound import InboundOptional, ArrivalBatch, DepartureBatch
    from sqlalchemy.orm import joinedload
    import calendar

    # Get filter parameters
    search_query = request.args.get('search', '').strip()
    service_types = [st for st in request.args.getlist('service_type') if st]
    status_filter = request.args.get('status', '')
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')
    request_number = request.args.get('request_number', '')

    # Default to current month if dates not provided
    today = datetime.now().date()
    first_day_of_month = today.replace(day=1)
    last_day_of_month = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else first_day_of_month
    date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else last_day_of_month

    # Helper: apply common filters and eagerly load parent request (prevents N+1)
    def _base(model, date_col):
        q = model.query.join(InboundRequest).options(joinedload(model.request))
        if date_from_str:
            q = q.filter(date_col >= date_from)
        if date_to_str:
            q = q.filter(date_col <= date_to)
        return q

    def _apply_filters(q, model, status_col, search_filters):
        if status_filter:
            q = q.filter(status_col == status_filter)
        if request_number:
            q = q.filter(InboundRequest.request_number.contains(request_number))
        if search_query:
            q = q.filter(db.or_(*search_filters))
        return q

    # --- Single-pass: fetch each service type once, build both stats and rows ---
    service_stats = {
        'HOTEL': {'total': 0, 'confirmed': 0, 'pax': 0},
        'TRANSPORT': {'total': 0, 'confirmed': 0, 'pax': 0},
        'GUIDE': {'total': 0, 'confirmed': 0, 'pax': 0},
        'MEAL': {'total': 0, 'confirmed': 0, 'pax': 0},
        'OPTIONAL': {'total': 0, 'confirmed': 0, 'pax': 0},
        'GROUND_HANDLER': {'total': 0, 'confirmed': 0, 'pax': 0},
    }
    all_services = []
    CONFIRMED_STATUSES = {'CONFIRMED', 'RESERVED'}

    # 1. Hotels
    hotels_q = _base(InboundHotel, InboundHotel.check_in_date)
    hotels_q = _apply_filters(hotels_q, InboundHotel, InboundHotel.status, [
        InboundHotel.hotel_name.contains(search_query),
        InboundHotel.location.contains(search_query),
        InboundHotel.hotel_category.contains(search_query),
        InboundRequest.contact_name.contains(search_query),
    ])
    for hotel in hotels_q.all():
        req = hotel.request
        service_stats['HOTEL']['total'] += 1
        if hotel.status in CONFIRMED_STATUSES:
            service_stats['HOTEL']['confirmed'] += 1
        service_stats['HOTEL']['pax'] += req.pax or 0
        if not service_types or 'HOTEL' in service_types:
            # Get category from hotel record, fallback to supplier if empty
            hotel_category = hotel.hotel_category
            if not hotel_category and hotel.hotel_name:
                supplier = Supplier.query.filter_by(name=hotel.hotel_name, supplier_type='HOTEL').first()
                if supplier:
                    hotel_category = supplier.accommodation_category

            all_services.append({
                'date': hotel.check_in_date,
                'service_type': 'HOTEL',
                'request_number': req.request_number,
                'request_id': hotel.request_id,
                'contact_name': req.contact_name,
                'pax': req.pax,
                'description': f"{hotel.hotel_name} - {hotel.location or ''} ({hotel.nights} nights)",
                'details': f"Category: {hotel_category or 'N/A'}, Meal: {hotel.meal_plan}",
                'status': hotel.status,
                'cost': hotel.total_cost,
                'currency': hotel.currency,
            })

    # 2. Transport
    transports_q = _base(InboundTransport, InboundTransport.date)
    transports_q = _apply_filters(transports_q, InboundTransport, InboundTransport.status, [
        InboundTransport.vehicle_type.contains(search_query),
        InboundTransport.supplier.contains(search_query),
        InboundTransport.driver_name.contains(search_query),
        InboundTransport.pickup_location.contains(search_query),
        InboundTransport.dropoff_location.contains(search_query),
        InboundRequest.contact_name.contains(search_query),
    ])
    for transport in transports_q.all():
        req = transport.request
        service_stats['TRANSPORT']['total'] += 1
        if transport.status in CONFIRMED_STATUSES:
            service_stats['TRANSPORT']['confirmed'] += 1
        service_stats['TRANSPORT']['pax'] += transport.pax or req.pax or 0
        if not service_types or 'TRANSPORT' in service_types:
            all_services.append({
                'date': transport.date,
                'service_type': 'TRANSPORT',
                'request_number': req.request_number,
                'request_id': transport.request_id,
                'contact_name': req.contact_name,
                'pax': transport.pax or req.pax,
                'description': f"{transport.vehicle_type or 'Transport'} - {transport.supplier or 'TBA'}",
                'details': f"From: {transport.pickup_location or 'N/A'}, To: {transport.dropoff_location or 'N/A'}, Driver: {transport.driver_name or 'TBA'}",
                'status': transport.status,
                'cost': transport.cost,
                'currency': transport.currency,
            })

    # 3. Guides
    guides_q = _base(InboundGuide, InboundGuide.date)
    guides_q = _apply_filters(guides_q, InboundGuide, InboundGuide.status, [
        InboundGuide.guide_name.contains(search_query),
        InboundGuide.language.contains(search_query),
        InboundGuide.service_type.contains(search_query),
        InboundRequest.contact_name.contains(search_query),
    ])
    for guide in guides_q.all():
        req = guide.request
        service_stats['GUIDE']['total'] += 1
        if guide.status in CONFIRMED_STATUSES:
            service_stats['GUIDE']['confirmed'] += 1
        service_stats['GUIDE']['pax'] += req.pax or 0
        if not service_types or 'GUIDE' in service_types:
            all_services.append({
                'date': guide.date,
                'service_type': 'GUIDE',
                'request_number': req.request_number,
                'request_id': guide.request_id,
                'contact_name': req.contact_name,
                'pax': req.pax,
                'description': f"{guide.guide_name or 'Guide'} - {guide.language or 'N/A'}",
                'details': f"Service: {guide.service_type or 'N/A'}, Tel: {guide.telephone_number or 'N/A'}",
                'status': guide.status,
                'cost': guide.cost,
                'currency': guide.currency,
            })

    # 4. Meals
    meals_q = _base(InboundMeal, InboundMeal.date)
    meals_q = _apply_filters(meals_q, InboundMeal, InboundMeal.status, [
        InboundMeal.restaurant.contains(search_query),
        InboundMeal.meal_type.contains(search_query),
        InboundMeal.location.contains(search_query),
        InboundRequest.contact_name.contains(search_query),
    ])
    for meal in meals_q.all():
        req = meal.request
        service_stats['MEAL']['total'] += 1
        if meal.status in CONFIRMED_STATUSES:
            service_stats['MEAL']['confirmed'] += 1
        service_stats['MEAL']['pax'] += req.pax or 0
        if not service_types or 'MEAL' in service_types:
            all_services.append({
                'date': meal.date,
                'service_type': 'MEAL',
                'request_number': req.request_number,
                'request_id': meal.request_id,
                'contact_name': req.contact_name,
                'pax': req.pax,
                'description': f"{meal.meal_type or 'Meal'} at {meal.restaurant or 'TBA'}",
                'details': f"Location: {meal.location or 'N/A'}",
                'status': meal.status,
                'cost': meal.total_cost,
                'currency': meal.currency,
            })

    # 5. Optional Services
    optionals_q = InboundOptional.query.join(InboundRequest).options(joinedload(InboundOptional.request))
    if date_from_str and date_to_str:
        optionals_q = optionals_q.filter(db.or_(
            InboundOptional.date == None,
            db.and_(InboundOptional.date >= date_from, InboundOptional.date <= date_to)
        ))
    optionals_q = _apply_filters(optionals_q, InboundOptional, InboundOptional.status, [
        InboundOptional.service_name.contains(search_query),
        InboundOptional.description.contains(search_query),
        InboundOptional.supplier.contains(search_query),
        InboundRequest.contact_name.contains(search_query),
    ])
    for optional in optionals_q.all():
        req = optional.request
        service_stats['OPTIONAL']['total'] += 1
        if optional.status in CONFIRMED_STATUSES:
            service_stats['OPTIONAL']['confirmed'] += 1
        service_stats['OPTIONAL']['pax'] += req.pax or 0
        if not service_types or 'OPTIONAL' in service_types:
            all_services.append({
                'date': optional.date or date_from,
                'service_type': 'OPTIONAL',
                'request_number': req.request_number,
                'request_id': optional.request_id,
                'contact_name': req.contact_name,
                'pax': req.pax,
                'description': optional.service_name,
                'details': f"{optional.description or ''}, Supplier: {optional.supplier or 'N/A'}",
                'status': optional.status,
                'cost': optional.total_cost,
                'currency': optional.currency,
            })

    # 6. Ground handler (ArrivalBatch + DepartureBatch) — stats only, no table rows
    arr_q = ArrivalBatch.query.join(InboundRequest)
    dep_q = DepartureBatch.query.join(InboundRequest)
    if date_from_str:
        arr_q = arr_q.filter(ArrivalBatch.arrival_date >= date_from)
        dep_q = dep_q.filter(DepartureBatch.departure_date >= date_from)
    if date_to_str:
        arr_q = arr_q.filter(ArrivalBatch.arrival_date <= date_to)
        dep_q = dep_q.filter(DepartureBatch.departure_date <= date_to)
    arr_all = arr_q.all()
    dep_all = dep_q.all()
    service_stats['GROUND_HANDLER']['total'] = len(arr_all) + len(dep_all)
    service_stats['GROUND_HANDLER']['pax'] = sum(r.pax_count or 0 for r in arr_all + dep_all)

    # Calculate share/confirmed percentages
    total_services = sum(s['total'] for s in service_stats.values())
    for stype, s in service_stats.items():
        s['share'] = round((s['total'] / total_services) * 100) if total_services > 0 else 0
        s['confirmed_pct'] = round((s['confirmed'] / s['total']) * 100) if s['total'] > 0 else 0

    # Sort by date
    all_services.sort(key=lambda x: x['date'])

    # Single GROUP BY query for status counts instead of 5 separate COUNTs
    from sqlalchemy import func
    status_rows = db.session.query(InboundRequest.status, func.count(InboundRequest.id))\
        .group_by(InboundRequest.status).all()
    status_counts = {row[0]: row[1] for row in status_rows}
    for s in ('REQUEST', 'QUOTED', 'CONFIRMED', 'PROCESSING', 'COMPLETED'):
        status_counts.setdefault(s, 0)

    return render_template('inbound/analytics.html',
                         services=all_services,
                         status_counts=status_counts,
                         service_stats=service_stats,
                         total_services=total_services,
                         search_query=search_query,
                         service_types=service_types,
                         status_filter=status_filter,
                         date_from=date_from,
                         date_to=date_to,
                         request_number=request_number)

@inbound_bp.route('/analytics/export-excel', methods=['POST'])
@login_required
def analytics_export_excel():
    """Export the currently displayed Supplier Analytics table as Excel.
    Accepts table data directly from frontend to ensure exact match."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400

        headers = data.get('headers', [])
        rows = data.get('rows', [])
        title = data.get('title', 'Supplier Analytics Export')

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = cast(Any, wb.active)
        ws.title = "Analytics"

        # Styling
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        data_alignment = Alignment(horizontal="left", vertical="center")
        data_alignment_numeric = Alignment(horizontal="right", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title row
        ws.merge_cells(f'A1:{chr(64 + len(headers))}1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Headers (row 3)
        for col_idx, header_text in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header_text
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows (starting row 4)
        for row_idx, row_data in enumerate(rows, start=4):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                cell.border = thin_border
                # Numeric cells right-aligned
                if isinstance(cell_value, (int, float)):
                    cell.alignment = data_alignment_numeric
                else:
                    cell.alignment = data_alignment

        # Auto-adjust column widths based on header length
        for col_idx, header_text in enumerate(headers, start=1):
            col_letter = chr(64 + col_idx)
            # Base width on header length, with minimum
            width = max(len(str(header_text)) + 2, 12)
            ws.column_dimensions[col_letter].width = width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"SupplierAnalytics_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@inbound_bp.route('/analytics/supplier-analytics')
@login_required
def supplier_analytics_api():
    """JSON API: supplier analytics.
    Always returns: key (name) | [attr_val if attribute selected] | total/confirmed/requested/invoiced
    Also returns attr_values list (distinct values) for filter chips.
    """
    from sqlalchemy import func, distinct
    from sqlalchemy import case as sa_case

    supplier_type = request.args.get('supplier_type', 'GUIDE').upper()
    attribute = request.args.get('attribute', '').lower()
    attr_filter = request.args.get('attr_filter', '').strip()
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    today = datetime.now().date()
    first_day = today.replace(day=1)
    last_day = today.replace(day=calendar.monthrange(today.year, today.month)[1])

    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else first_day
    date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else last_day

    def _when(col, val):
        return sa_case((col == val, 1), else_=0)

    def _attr_query(name_col, attr_col, base_model, join_model, join_cond, base_filters, af):
        """Query grouping by name + attribute, returning standard fields."""
        return (
            db.session.query(
                name_col.label('name'),
                func.coalesce(attr_col, 'Unspecified').label('attr_val'),
                func.count(base_model.id).label('total'),
                func.sum(_when(join_model.status, 'CONFIRMED')).label('confirmed'),
                func.sum(_when(join_model.status, 'REQUEST')).label('requested'),
                func.sum(_when(join_model.status, 'INVOICED')).label('invoiced'),
            )
            .join(join_model, join_cond)
            .filter(*base_filters, *af)
            .group_by(name_col, attr_col)
            .order_by(func.count(base_model.id).desc())
        )

    def _name_query(name_col, base_model, join_model, join_cond, base_filters):
        """Query grouping by name only."""
        return (
            db.session.query(
                name_col.label('name'),
                func.count(base_model.id).label('total'),
                func.sum(_when(join_model.status, 'CONFIRMED')).label('confirmed'),
                func.sum(_when(join_model.status, 'REQUEST')).label('requested'),
                func.sum(_when(join_model.status, 'INVOICED')).label('invoiced'),
            )
            .join(join_model, join_cond)
            .filter(*base_filters)
            .group_by(name_col)
            .order_by(func.count(base_model.id).desc())
        )

    def _to_items(rows, has_attr):
        return [{'key': r.name or 'Unknown',
                 **({'attr_val': r.attr_val} if has_attr else {}),
                 'total': r.total,
                 'confirmed': r.confirmed or 0,
                 'requested': r.requested or 0,
                 'invoiced': r.invoiced or 0}
                for r in rows]

    def _attr_values(model, attr_col, attr_value_filters):
        rows = db.session.query(
            func.coalesce(attr_col, 'Unspecified').label('v')
        ).select_from(model).filter(*attr_value_filters).distinct().order_by('v').all()
        return [r.v for r in rows if r.v]

    def _attr_values_from_master(master_model):
        """Query attribute values from master data table (for categories, types, etc.)."""
        rows = db.session.query(master_model.name).filter(master_model.is_active == True).order_by(master_model.sort_order, master_model.name).all()
        return [r[0] for r in rows if r[0]]

    items = []
    attr_values = []
    attr_label = ''

    # ── GUIDE ──────────────────────────────────────────────────────────────────
    if supplier_type == 'GUIDE':
        base_filters = [
            InboundGuide.date >= date_from,
            InboundGuide.date <= date_to,
            InboundGuide.is_cancelled == False,
        ]
        attr_value_filters = [
            InboundGuide.is_cancelled == False,
        ]
        ATTR_COLS = {
            'language': (InboundGuide.language, 'Language'),
        }
        if attribute in ATTR_COLS:
            attr_col, attr_label = ATTR_COLS[attribute]
            # For language, show all available languages from the creation form
            if attribute == 'language':
                attr_values = GUIDE_LANGUAGES
            else:
                attr_values = _attr_values(InboundGuide, attr_col, attr_value_filters)
            af = [attr_col == attr_filter] if attr_filter else []
            q = _attr_query(InboundGuide.guide_name, attr_col, InboundGuide, InboundRequest,
                            InboundGuide.request_id == InboundRequest.id, base_filters, af)
            items = _to_items(q.all(), True)
        else:
            # For default "By Name" view, use Supplier FK when available, fallback to text field
            from app.models.supplier import Supplier as _Supplier
            supplier_name_col = func.coalesce(_Supplier.name, InboundGuide.guide_name).label('guide_name')
            q = (
                db.session.query(
                    supplier_name_col.label('name'),
                    func.count(InboundGuide.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, InboundGuide.request_id == InboundRequest.id)
                .outerjoin(_Supplier, InboundGuide.supplier_id == _Supplier.id)
                .filter(*base_filters)
                .group_by(supplier_name_col)
                .order_by(func.count(InboundGuide.id).desc())
            )
            items = _to_items(q.all(), False)

    # ── TRANSPORT ──────────────────────────────────────────────────────────────
    elif supplier_type == 'TRANSPORT':
        from app.models.supplier import Supplier

        base_filters = [
            InboundTransport.date >= date_from,
            InboundTransport.date <= date_to,
        ]
        attr_value_filters = []
        ATTR_COLS = {
            'vehicle_type': (InboundTransport.vehicle_type, 'Vehicle Type'),
            'supplier_type': (Supplier.entity_type, 'Supplier Type'),
        }

        if attribute == 'supplier_type':
            attr_label = 'Supplier Type'
            # Get distinct supplier types from Supplier table for TRANSPORT suppliers
            attr_values = db.session.query(
                func.coalesce(Supplier.entity_type, 'COMPANY').label('v')
            ).select_from(Supplier).filter(
                Supplier.supplier_type == 'TRANSPORT',
                Supplier.is_active == True
            ).distinct().order_by('v').all()
            attr_values = [r.v for r in attr_values if r.v]

            # Query with Supplier join for entity_type filtering
            af = [Supplier.entity_type == attr_filter] if attr_filter else []
            q = (
                db.session.query(
                    InboundTransport.supplier.label('name'),
                    func.coalesce(Supplier.entity_type, 'COMPANY').label('attr_val'),
                    func.count(InboundTransport.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, InboundTransport.request_id == InboundRequest.id)
                .join(Supplier, Supplier.name == InboundTransport.supplier)
                .filter(*base_filters, *af)
                .group_by(InboundTransport.supplier, Supplier.entity_type)
                .order_by(func.count(InboundTransport.id).desc())
            )
            items = _to_items(q.all(), True)
        elif attribute in ATTR_COLS:
            attr_col, attr_label = ATTR_COLS[attribute]
            attr_values = _attr_values(InboundTransport, attr_col, attr_value_filters)
            af = [attr_col == attr_filter] if attr_filter else []
            q = _attr_query(InboundTransport.supplier, attr_col, InboundTransport, InboundRequest,
                            InboundTransport.request_id == InboundRequest.id, base_filters, af)
            items = _to_items(q.all(), True)
        else:
            # For default "By Name" view, use Supplier FK when available, fallback to text field
            supplier_name_col = func.coalesce(Supplier.name, InboundTransport.supplier).label('supplier_name')
            q = (
                db.session.query(
                    supplier_name_col.label('name'),
                    func.count(InboundTransport.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, InboundTransport.request_id == InboundRequest.id)
                .outerjoin(Supplier, InboundTransport.supplier_id == Supplier.id)
                .filter(*base_filters)
                .group_by(supplier_name_col)
                .order_by(func.count(InboundTransport.id).desc())
            )
            items = _to_items(q.all(), False)

    # ── RESTAURANT ─────────────────────────────────────────────────────────────
    elif supplier_type == 'RESTAURANT':
        base_filters = [
            InboundMeal.date >= date_from,
            InboundMeal.date <= date_to,
        ]
        attr_value_filters = []
        ATTR_COLS = {
            'meal_type': (InboundMeal.meal_type, 'Meal Type'),
            'location':  (InboundMeal.location,  'Location'),
        }
        if attribute in ATTR_COLS:
            attr_col, attr_label = ATTR_COLS[attribute]
            attr_values = _attr_values(InboundMeal, attr_col, attr_value_filters)
            af = [attr_col == attr_filter] if attr_filter else []
            q = _attr_query(InboundMeal.restaurant, attr_col, InboundMeal, InboundRequest,
                            InboundMeal.request_id == InboundRequest.id, base_filters, af)
            items = _to_items(q.all(), True)
        else:
            # For default "By Name" view, use Supplier FK when available, fallback to text field
            from app.models.supplier import Supplier as _Supplier
            supplier_name_col = func.coalesce(_Supplier.name, InboundMeal.restaurant).label('restaurant_name')
            q = (
                db.session.query(
                    supplier_name_col.label('name'),
                    func.count(InboundMeal.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, InboundMeal.request_id == InboundRequest.id)
                .outerjoin(_Supplier, InboundMeal.supplier_id == _Supplier.id)
                .filter(*base_filters)
                .group_by(supplier_name_col)
                .order_by(func.count(InboundMeal.id).desc())
            )
            items = _to_items(q.all(), False)

    # ── HOTEL / ACCOMMODATION ──────────────────────────────────────────────────
    elif supplier_type == 'HOTEL':
        base_filters = [
            InboundHotel.check_in_date >= date_from,
            InboundHotel.check_in_date <= date_to,
        ]
        attr_value_filters = []
        ATTR_COLS = {
            'hotel_category': (InboundHotel.hotel_category, 'Category'),
            'meal_plan':      (InboundHotel.meal_plan,      'Meal Plan'),
            'location':       (InboundHotel.location,       'Location'),
        }
        if attribute in ATTR_COLS:
            attr_col, attr_label = ATTR_COLS[attribute]
            # Use master data for hotel_category, but also include categories actually in use
            if attribute == 'hotel_category':
                master_values = _attr_values_from_master(HotelCategory)
                # Also get categories from actual InboundHotel records
                used_values = _attr_values(InboundHotel, attr_col, attr_value_filters)
                # Merge and deduplicate, keeping master order first then adding new ones
                attr_values = list(dict.fromkeys(master_values + used_values))
            else:
                attr_values = _attr_values(InboundHotel, attr_col, attr_value_filters)
            af = [attr_col == attr_filter] if attr_filter else []
            q = _attr_query(InboundHotel.hotel_name, attr_col, InboundHotel, InboundRequest,
                            InboundHotel.request_id == InboundRequest.id, base_filters, af)
            items = _to_items(q.all(), True)
        else:
            q = _name_query(InboundHotel.hotel_name, InboundHotel, InboundRequest,
                            InboundHotel.request_id == InboundRequest.id, base_filters)
            items = _to_items(q.all(), False)

    # ── MEET & ASSIST ──────────────────────────────────────────────────────────
    elif supplier_type == 'MEET_ASSIST':
        from app.models.supplier import Supplier

        attr_label = ''

        if attribute == 'supplier_type':
            attr_label = 'Supplier Type'
            # Get distinct supplier types from GROUND_HANDLER suppliers
            attr_values = db.session.query(
                func.coalesce(Supplier.entity_type, 'COMPANY').label('v')
            ).select_from(Supplier).filter(
                Supplier.supplier_type == 'GROUND_HANDLER',
                Supplier.is_active == True
            ).distinct().order_by('v').all()
            attr_values = [r.v for r in attr_values if r.v]

            # Query arrivals with supplier_type
            arrival_rows = (
                db.session.query(
                    Supplier.name.label('name'),
                    func.coalesce(Supplier.entity_type, 'COMPANY').label('attr_val'),
                    func.count(ArrivalBatch.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, ArrivalBatch.request_id == InboundRequest.id)
                .join(Supplier, ArrivalBatch.supplier_id == Supplier.id)
                .filter(
                    ArrivalBatch.arrival_date >= date_from,
                    ArrivalBatch.arrival_date <= date_to,
                    Supplier.supplier_type == 'GROUND_HANDLER',
                    *([Supplier.entity_type == attr_filter] if attr_filter else [])
                )
                .group_by(Supplier.name, Supplier.entity_type)
            ).all()

            # Query departures with supplier_type
            departure_rows = (
                db.session.query(
                    Supplier.name.label('name'),
                    func.coalesce(Supplier.entity_type, 'COMPANY').label('attr_val'),
                    func.count(DepartureBatch.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, DepartureBatch.request_id == InboundRequest.id)
                .join(Supplier, DepartureBatch.supplier_id == Supplier.id)
                .filter(
                    DepartureBatch.departure_date >= date_from,
                    DepartureBatch.departure_date <= date_to,
                    Supplier.supplier_type == 'GROUND_HANDLER',
                    *([Supplier.entity_type == attr_filter] if attr_filter else [])
                )
                .group_by(Supplier.name, Supplier.entity_type)
            ).all()

            # Merge arrival and departure data by supplier and supplier_type
            merged_data = {}
            for row in arrival_rows:
                key = (row.name, row.attr_val)
                if key not in merged_data:
                    merged_data[key] = {'key': row.name, 'attr_val': row.attr_val, 'total': 0, 'confirmed': 0, 'requested': 0, 'invoiced': 0}
                merged_data[key]['total'] += row.total or 0
                merged_data[key]['confirmed'] += row.confirmed or 0
                merged_data[key]['requested'] += row.requested or 0
                merged_data[key]['invoiced'] += row.invoiced or 0

            for row in departure_rows:
                key = (row.name, row.attr_val)
                if key not in merged_data:
                    merged_data[key] = {'key': row.name, 'attr_val': row.attr_val, 'total': 0, 'confirmed': 0, 'requested': 0, 'invoiced': 0}
                merged_data[key]['total'] += row.total or 0
                merged_data[key]['confirmed'] += row.confirmed or 0
                merged_data[key]['requested'] += row.requested or 0
                merged_data[key]['invoiced'] += row.invoiced or 0

            items = [v for v in merged_data.values() if v['total'] > 0]
            items.sort(key=lambda x: x['total'], reverse=True)
        else:
            # Query by supplier name only (no attribute)
            arrival_rows = (
                db.session.query(
                    Supplier.name.label('name'),
                    func.count(ArrivalBatch.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, ArrivalBatch.request_id == InboundRequest.id)
                .join(Supplier, ArrivalBatch.supplier_id == Supplier.id)
                .filter(
                    ArrivalBatch.arrival_date >= date_from,
                    ArrivalBatch.arrival_date <= date_to,
                    Supplier.supplier_type == 'GROUND_HANDLER',
                )
                .group_by(Supplier.name)
            ).all()

            departure_rows = (
                db.session.query(
                    Supplier.name.label('name'),
                    func.count(DepartureBatch.id).label('total'),
                    func.sum(_when(InboundRequest.status, 'CONFIRMED')).label('confirmed'),
                    func.sum(_when(InboundRequest.status, 'REQUEST')).label('requested'),
                    func.sum(_when(InboundRequest.status, 'INVOICED')).label('invoiced'),
                )
                .join(InboundRequest, DepartureBatch.request_id == InboundRequest.id)
                .join(Supplier, DepartureBatch.supplier_id == Supplier.id)
                .filter(
                    DepartureBatch.departure_date >= date_from,
                    DepartureBatch.departure_date <= date_to,
                    Supplier.supplier_type == 'GROUND_HANDLER',
                )
                .group_by(Supplier.name)
            ).all()

            # Merge by supplier name
            merged_data = {}
            for row in arrival_rows:
                if row.name not in merged_data:
                    merged_data[row.name] = {'key': row.name, 'total': 0, 'confirmed': 0, 'requested': 0, 'invoiced': 0}
                merged_data[row.name]['total'] += row.total or 0
                merged_data[row.name]['confirmed'] += row.confirmed or 0
                merged_data[row.name]['requested'] += row.requested or 0
                merged_data[row.name]['invoiced'] += row.invoiced or 0

            for row in departure_rows:
                if row.name not in merged_data:
                    merged_data[row.name] = {'key': row.name, 'total': 0, 'confirmed': 0, 'requested': 0, 'invoiced': 0}
                merged_data[row.name]['total'] += row.total or 0
                merged_data[row.name]['confirmed'] += row.confirmed or 0
                merged_data[row.name]['requested'] += row.requested or 0
                merged_data[row.name]['invoiced'] += row.invoiced or 0

            items = [v for v in merged_data.values() if v['total'] > 0]
            items.sort(key=lambda x: x['total'], reverse=True)

    return jsonify({
        'supplier_type': supplier_type,
        'attribute': attribute,
        'attr_label': attr_label,
        'attr_values': attr_values,
        'attr_filter': attr_filter,
        'items': items,
        'total': len(items),
        'date_from': str(date_from),
        'date_to': str(date_to),
    })


@inbound_bp.route('/analytics/kpi-data')
@login_required
def analytics_kpi_data():
    """JSON: KPI summary statistics aligned with /finance/suppliers business categories."""
    from app.models.inbound import InboundOptional, ArrivalBatch, DepartureBatch

    today = datetime.now().date()
    date_from_str = request.args.get('date_from', '')
    date_to_str  = request.args.get('date_to', '')
    first_day = today.replace(day=1)
    last_day  = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else first_day
    date_to   = datetime.strptime(date_to_str,  '%Y-%m-%d').date() if date_to_str  else last_day

    def _stats(model, date_col, has_status=True):
        rows = (
            model.query.join(InboundRequest)
            .filter(date_col >= date_from, date_col <= date_to)
            .all()
        )
        total = len(rows)
        confirmed = sum(1 for r in rows if has_status and r.status in ('CONFIRMED', 'RESERVED'))
        pax = sum((r.request.pax or 0) for r in rows)
        return {'total': total, 'confirmed': confirmed, 'pax': pax}

    # Core service categories matching /finance/suppliers business names
    stats = {
        'HOTEL':     _stats(InboundHotel,    InboundHotel.check_in_date),
        'TRANSPORT': _stats(InboundTransport, InboundTransport.date),
        'GUIDE':     _stats(InboundGuide,     InboundGuide.date),
        'MEAL':      _stats(InboundMeal,      InboundMeal.date),
    }

    # Meet & Assist: ArrivalBatch + DepartureBatch (GROUND_HANDLER suppliers)
    arr_rows = (
        ArrivalBatch.query.join(InboundRequest)
        .filter(ArrivalBatch.arrival_date >= date_from, ArrivalBatch.arrival_date <= date_to)
        .all()
    )
    dep_rows = (
        DepartureBatch.query.join(InboundRequest)
        .filter(DepartureBatch.departure_date >= date_from, DepartureBatch.departure_date <= date_to)
        .all()
    )
    all_batches = arr_rows + dep_rows
    stats['GROUND_HANDLER'] = {
        'total':     len(all_batches),
        'confirmed': 0,
        'pax':       sum((r.pax_count or 0) for r in all_batches),
    }

    opt_rows = (
        InboundOptional.query.join(InboundRequest)
        .filter(db.or_(
            InboundOptional.date == None,
            db.and_(InboundOptional.date >= date_from, InboundOptional.date <= date_to)
        )).all()
    )
    stats['OPTIONAL'] = {
        'total':     len(opt_rows),
        'confirmed': sum(1 for o in opt_rows if o.status in ('CONFIRMED', 'RESERVED')),
        'pax':       sum(o.request.pax or 0 for o in opt_rows),
    }

    total_services = sum(s['total'] for s in stats.values())
    for s in stats.values():
        s['share']         = round(s['total'] / total_services * 100) if total_services else 0
        s['confirmed_pct'] = round(s['confirmed'] / s['total'] * 100) if s['total'] else 0

    return jsonify({'stats': stats, 'total_services': total_services})


@inbound_bp.route('/analytics/run-down-data')
@login_required
def analytics_run_down_data():
    """JSON: Filtered run-down service rows for the detail table."""
    from app.models.inbound import InboundOptional

    service_types  = [st for st in request.args.getlist('service_type') if st]
    status_filter  = request.args.get('status', '')
    search_query   = request.args.get('search', '').strip()
    date_from_str  = request.args.get('date_from', '')
    date_to_str    = request.args.get('date_to', '')

    today     = datetime.now().date()
    first_day = today.replace(day=1)
    last_day  = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else first_day
    date_to   = datetime.strptime(date_to_str,  '%Y-%m-%d').date() if date_to_str  else last_day

    rows = []

    if not service_types or 'HOTEL' in service_types:
        q = InboundHotel.query.join(InboundRequest).filter(
            InboundHotel.check_in_date >= date_from,
            InboundHotel.check_in_date <= date_to,
        )
        if status_filter: q = q.filter(InboundHotel.status == status_filter)
        if search_query:
            q = q.filter(db.or_(
                InboundHotel.hotel_name.contains(search_query),
                InboundHotel.location.contains(search_query),
                InboundRequest.contact_name.contains(search_query),
            ))
        for h in q.all():
            # Get category from hotel record, fallback to supplier if empty
            hotel_category = h.hotel_category
            if not hotel_category and h.hotel_name:
                supplier = Supplier.query.filter_by(name=h.hotel_name, supplier_type='HOTEL').first()
                if supplier:
                    hotel_category = supplier.accommodation_category

            rows.append({
                'date': h.check_in_date.strftime('%Y-%m-%d'),
                'date_display': h.check_in_date.strftime('%d %b %Y'),
                'service_type': 'HOTEL',
                'request_id': h.request_id,
                'request_number': h.request.request_number,
                'contact_name': h.request.contact_name,
                'pax': h.request.pax,
                'description': f"{h.hotel_name or ''} – {h.location or ''} ({h.nights}n)",
                'hotel_category': hotel_category or '',
                'status': h.status,
                'cost': h.total_cost,
                'currency': h.currency,
            })

    if not service_types or 'TRANSPORT' in service_types:
        q = InboundTransport.query.join(InboundRequest).filter(
            InboundTransport.date >= date_from,
            InboundTransport.date <= date_to,
        )
        if status_filter: q = q.filter(InboundTransport.status == status_filter)
        if search_query:
            q = q.filter(db.or_(
                InboundTransport.vehicle_type.contains(search_query),
                InboundTransport.supplier.contains(search_query),
                InboundTransport.driver_name.contains(search_query),
                InboundRequest.contact_name.contains(search_query),
            ))
        for t in q.all():
            rows.append({
                'date': t.date.strftime('%Y-%m-%d'),
                'date_display': t.date.strftime('%d %b %Y'),
                'service_type': 'TRANSPORT',
                'request_id': t.request_id,
                'request_number': t.request.request_number,
                'contact_name': t.request.contact_name,
                'pax': t.pax or t.request.pax,
                'description': f"{t.vehicle_type or 'Transport'} – {t.supplier_name or 'TBA'}",
                'status': t.status,
                'cost': t.cost,
                'currency': t.currency,
            })

    if not service_types or 'GUIDE' in service_types:
        q = InboundGuide.query.join(InboundRequest).filter(
            InboundGuide.date >= date_from,
            InboundGuide.date <= date_to,
            InboundGuide.is_cancelled == False,
        )
        if status_filter: q = q.filter(InboundGuide.status == status_filter)
        if search_query:
            q = q.filter(db.or_(
                InboundGuide.guide_name.contains(search_query),
                InboundGuide.language.contains(search_query),
                InboundRequest.contact_name.contains(search_query),
            ))
        for g in q.all():
            rows.append({
                'date': g.date.strftime('%Y-%m-%d'),
                'date_display': g.date.strftime('%d %b %Y'),
                'service_type': 'GUIDE',
                'request_id': g.request_id,
                'request_number': g.request.request_number,
                'contact_name': g.request.contact_name,
                'pax': g.request.pax,
                'description': f"{g.guide_name or 'Guide'} – {g.language or 'N/A'}",
                'status': g.status,
                'cost': g.cost,
                'currency': g.currency,
            })

    if not service_types or 'MEAL' in service_types:
        q = InboundMeal.query.join(InboundRequest).filter(
            InboundMeal.date >= date_from,
            InboundMeal.date <= date_to,
        )
        if status_filter: q = q.filter(InboundMeal.status == status_filter)
        if search_query:
            q = q.filter(db.or_(
                InboundMeal.restaurant.contains(search_query),
                InboundMeal.meal_type.contains(search_query),
                InboundRequest.contact_name.contains(search_query),
            ))
        for m in q.all():
            rows.append({
                'date': m.date.strftime('%Y-%m-%d'),
                'date_display': m.date.strftime('%d %b %Y'),
                'service_type': 'MEAL',
                'request_id': m.request_id,
                'request_number': m.request.request_number,
                'contact_name': m.request.contact_name,
                'pax': m.request.pax,
                'description': f"{m.meal_type or 'Meal'} – {m.supplier_name or 'TBA'}",
                'status': m.status,
                'cost': m.total_cost,
                'currency': m.currency,
            })

    if not service_types or 'OPTIONAL' in service_types:
        q = InboundOptional.query.join(InboundRequest).filter(
            db.or_(
                InboundOptional.date == None,
                db.and_(InboundOptional.date >= date_from, InboundOptional.date <= date_to)
            )
        )
        if status_filter: q = q.filter(InboundOptional.status == status_filter)
        if search_query:
            q = q.filter(db.or_(
                InboundOptional.service_name.contains(search_query),
                InboundRequest.contact_name.contains(search_query),
            ))
        for o in q.all():
            d = o.date or date_from
            rows.append({
                'date': d.strftime('%Y-%m-%d'),
                'date_display': d.strftime('%d %b %Y'),
                'service_type': 'OPTIONAL',
                'request_id': o.request_id,
                'request_number': o.request.request_number,
                'contact_name': o.request.contact_name,
                'pax': o.request.pax,
                'description': o.service_name,
                'status': o.status,
                'cost': o.total_cost,
                'currency': o.currency,
            })

    rows.sort(key=lambda r: r['date'])
    return jsonify({'services': rows, 'total': len(rows)})


@inbound_bp.route('/analytics/supplier-types')
@login_required
def analytics_supplier_types():
    """JSON: Business supplier categories matching /finance/suppliers (excludes Airline)."""
    from app.models.supplier import Supplier as _Supplier
    from sqlalchemy import or_ as _or

    # Mirrors _SUPPLIER_TYPE_PAGE_MAP in finance.py — same categories as the suppliers hub
    CATEGORIES = [
        {
            'key': 'accommodation', 'label': 'Accommodation', 'icon': 'fa-hotel',
            'patterns': ['HOTEL', 'ACCOMMODATION'], 'service': 'HOTEL',
        },
        {
            'key': 'guides', 'label': 'Guides', 'icon': 'fa-user-tie',
            'patterns': ['GUIDE'], 'service': 'GUIDE',
        },
        {
            'key': 'transportation', 'label': 'Transportation', 'icon': 'fa-bus',
            'patterns': ['TRANSPORT', 'TRANSPORTATION', 'TRANSFER'], 'service': 'TRANSPORT',
        },
        {
            'key': 'meet-assist', 'label': 'Meet & Assist', 'icon': 'fa-handshake',
            'patterns': ['GROUND_HANDLER', 'MEET', 'ASSIST'], 'service': 'GROUND_HANDLER',
        },
        {
            'key': 'restaurant', 'label': 'Restaurant', 'icon': 'fa-utensils',
            'patterns': ['RESTAURANT', 'MEAL', 'FOOD'], 'service': 'RESTAURANT',
        },
    ]

    active = []
    for cat in CATEGORIES:
        f = [_Supplier.supplier_type.ilike(f'%{p}%') for p in cat['patterns']]
        count = _Supplier.query.filter(_Supplier.is_active == True, _or(*f)).count() if f else 0
        if count > 0:
            active.append({**cat, 'supplier_count': count})

    return jsonify({'categories': active})


@inbound_bp.route('/analytics/suppliers-by-type')
@login_required
def analytics_suppliers_by_type():
    """JSON: Supplier records from the Supplier table with service usage counts.
    Accepts category key matching /finance/suppliers (e.g. 'guides', 'accommodation').
    """
    from app.models.supplier import Supplier as _Supplier
    from app.models.inbound import ArrivalBatch, DepartureBatch
    from sqlalchemy import func, case as sa_case, or_ as _or

    category_key  = request.args.get('category', 'guides').lower()
    statuses      = [s for s in request.args.getlist('status') if s]
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to', '')

    today     = datetime.now().date()
    first_day = today.replace(day=1)
    last_day  = today.replace(day=calendar.monthrange(today.year, today.month)[1])
    date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else first_day
    date_to   = datetime.strptime(date_to_str,  '%Y-%m-%d').date() if date_to_str  else last_day

    # Business category → (DB type patterns, service model, date col, FK col)
    CATEGORY_MAP = {
        'accommodation': (['HOTEL', 'ACCOMMODATION'], None,            None,                          None),
        'guides':        (['GUIDE'],                  InboundGuide,    InboundGuide.date,             InboundGuide.supplier_id),
        'transportation':(['TRANSPORT','TRANSPORTATION','TRANSFER'],
                                                      InboundTransport,InboundTransport.date,         InboundTransport.supplier_id),
        'meet-assist':   (['GROUND_HANDLER','MEET','ASSIST'],
                                                      None,            None,                          None),  # uses ArrivalBatch/DepartureBatch below
        'restaurant':    (['RESTAURANT','MEAL','FOOD'],InboundMeal,    InboundMeal.date,              InboundMeal.supplier_id),
    }

    if category_key not in CATEGORY_MAP:
        return jsonify({'category': category_key, 'items': [], 'total': 0,
                        'date_from': str(date_from), 'date_to': str(date_to)})

    patterns, SvcModel, date_col, fk_col = CATEGORY_MAP[category_key]

    def _when(col, val):
        return sa_case((col == val, 1), else_=0)

    # Build supplier filter using ilike patterns (matches /finance/suppliers logic)
    sup_filter = _or(*[_Supplier.supplier_type.ilike(f'%{p}%') for p in patterns])

    if SvcModel is not None and fk_col is not None:
        # Standard join: count service records per supplier
        sub_filters = [date_col >= date_from, date_col <= date_to]
        if statuses:
            status_col = getattr(SvcModel, 'status', None)
            if status_col is not None:
                sub_filters.append(_or(*[status_col == s for s in statuses]))

        sub = (
            db.session.query(
                fk_col.label('supplier_id'),
                func.count(SvcModel.id).label('total'),
                func.sum(_when(getattr(SvcModel, 'status'), 'CONFIRMED')).label('confirmed'),
                func.sum(_when(getattr(SvcModel, 'status'), 'REQUEST')).label('requested'),
                func.sum(_when(getattr(SvcModel, 'status'), 'INVOICED')).label('invoiced'),
            )
            .filter(*sub_filters)
            .group_by(fk_col)
            .subquery()
        )

        rows = (
            db.session.query(
                _Supplier.id, _Supplier.name, _Supplier.city,
                _Supplier.country, _Supplier.languages,
                func.coalesce(sub.c.total,     0).label('total'),
                func.coalesce(sub.c.confirmed, 0).label('confirmed'),
                func.coalesce(sub.c.requested, 0).label('requested'),
                func.coalesce(sub.c.invoiced,  0).label('invoiced'),
            )
            .outerjoin(sub, _Supplier.id == sub.c.supplier_id)
            .filter(_Supplier.is_active == True, sup_filter)
            .order_by(func.coalesce(sub.c.total, 0).desc(), _Supplier.name)
            .all()
        )
        items = [
            {'id': r.id, 'name': r.name, 'city': r.city or '—', 'country': r.country or '—',
             'languages': r.languages or '—', 'total': r.total, 'confirmed': r.confirmed,
             'requested': r.requested, 'invoiced': r.invoiced}
            for r in rows
        ]

    elif category_key == 'meet-assist':
        # Meet & Assist: count from ArrivalBatch + DepartureBatch per supplier
        arr_sub = (
            db.session.query(
                ArrivalBatch.supplier_id.label('supplier_id'),
                func.count(ArrivalBatch.id).label('total'),
            )
            .filter(ArrivalBatch.arrival_date >= date_from, ArrivalBatch.arrival_date <= date_to)
            .group_by(ArrivalBatch.supplier_id).subquery()
        )
        dep_sub = (
            db.session.query(
                DepartureBatch.supplier_id.label('supplier_id'),
                func.count(DepartureBatch.id).label('total'),
            )
            .filter(DepartureBatch.departure_date >= date_from, DepartureBatch.departure_date <= date_to)
            .group_by(DepartureBatch.supplier_id).subquery()
        )
        rows = (
            db.session.query(
                _Supplier.id, _Supplier.name, _Supplier.city, _Supplier.country,
                func.coalesce(arr_sub.c.total, 0).label('arrivals'),
                func.coalesce(dep_sub.c.total, 0).label('departures'),
            )
            .outerjoin(arr_sub, _Supplier.id == arr_sub.c.supplier_id)
            .outerjoin(dep_sub, _Supplier.id == dep_sub.c.supplier_id)
            .filter(_Supplier.is_active == True, sup_filter)
            .order_by(
                (func.coalesce(arr_sub.c.total, 0) + func.coalesce(dep_sub.c.total, 0)).desc(),
                _Supplier.name
            )
            .all()
        )
        items = [
            {'id': r.id, 'name': r.name, 'city': r.city or '—', 'country': r.country or '—',
             'arrivals': r.arrivals, 'departures': r.departures,
             'total': r.arrivals + r.departures}
            for r in rows
        ]

    else:
        # Accommodation: no FK on InboundHotel — show suppliers with 0 counts
        rows = (
            _Supplier.query
            .filter(_Supplier.is_active == True, sup_filter)
            .order_by(_Supplier.name)
            .all()
        )
        items = [
            {'id': r.id, 'name': r.name, 'city': r.city or '—', 'country': r.country or '—',
             'total': 0, 'confirmed': 0, 'requested': 0, 'invoiced': 0}
            for r in rows
        ]

    return jsonify({
        'category': category_key,
        'items': items,
        'total': len(items),
        'date_from': str(date_from),
        'date_to': str(date_to),
    })


# ==================== DOCUMENT MANAGEMENT ====================

ALLOWED_DOCUMENT_EXTENSIONS = {"pdf", "png", "jpg", "jpeg", "gif", "doc", "docx", "xls", "xlsx", "txt"}

def allowed_document_file(filename):
    """Check if file extension is allowed"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_DOCUMENT_EXTENSIONS

@inbound_bp.route("/api/<int:request_id>/documents", methods=["GET"])
def api_list_documents(request_id):
    """List all documents for a request"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    documents = InboundDocument.query.filter_by(request_id=request_id).order_by(InboundDocument.uploaded_at.desc()).all()

    return jsonify({
        "success": True,
        "documents": [{
            "id": doc.id,
            "document_type": doc.document_type,
            "original_filename": doc.original_filename,
            "file_size": doc.file_size,
            "mime_type": doc.mime_type,
            "description": doc.description,
            "uploaded_at": doc.uploaded_at.strftime("%Y-%m-%d %H:%M") if doc.uploaded_at else None,
            "is_image": doc.is_image,
            "is_pdf": doc.is_pdf
        } for doc in documents]
    })

@inbound_bp.route("/api/<int:request_id>/documents/upload", methods=["POST"])
@csrf.exempt
def api_upload_document(request_id):
    """Upload a document attachment"""
    request_obj = InboundRequest.query.get_or_404(request_id)

    if "file" not in request.files:
        return jsonify({"success": False, "message": "No file provided"}), 400

    file = request.files["file"]

    if file.filename == "":
        return jsonify({"success": False, "message": "No file selected"}), 400

    if not allowed_document_file(file.filename):
        return jsonify({"success": False, "message": "File type not allowed"}), 400

    try:
        # Generate unique filename
        original_filename = secure_filename(file.filename)
        file_ext = original_filename.rsplit(".", 1)[1].lower() if "." in original_filename else ""
        unique_filename = f"{uuid.uuid4().hex}_{original_filename}"

        # Create upload directory
        upload_folder = os.path.join("app", "static", "uploads", "inbound_documents", str(request_id))
        os.makedirs(upload_folder, exist_ok=True)

        # Save file
        filepath = os.path.join(upload_folder, unique_filename)
        file.save(filepath)

        # Get file info
        file_size = os.path.getsize(filepath)
        mime_type = file.content_type or "application/octet-stream"

        # Get document type from form
        document_type = request.form.get("document_type", "OTHER")
        description = request.form.get("description", "")

        # Create database record
        doc = InboundDocument(  # type: ignore[call-arg]
            request_id=request_id,
            document_type=document_type,
            filename=unique_filename,
            original_filename=original_filename,
            filepath=f"uploads/inbound_documents/{request_id}/{unique_filename}",
            file_size=file_size,
            mime_type=mime_type,
            description=description,
            uploaded_by=1  # Default user
        )
        db.session.add(doc)
        db.session.commit()

        return jsonify({
            "success": True,
            "message": "Document uploaded successfully",
            "document": {
                "id": doc.id,
                "document_type": doc.document_type,
                "original_filename": doc.original_filename,
                "filepath": doc.filepath,
                "file_size": doc.file_size,
                "uploaded_at": doc.uploaded_at.strftime("%d %b %Y") if doc.uploaded_at else "",
                "is_image": doc.is_image,
                "is_pdf": doc.is_pdf
            }
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@inbound_bp.route("/api/documents/<int:doc_id>/delete", methods=["POST"])
@csrf.exempt
def api_delete_document(doc_id):
    """Delete a document"""
    doc = InboundDocument.query.get_or_404(doc_id)

    try:
        # Delete file from filesystem
        full_path = os.path.join("app", "static", doc.filepath)
        if os.path.exists(full_path):
            os.remove(full_path)

        # Delete database record
        db.session.delete(doc)
        db.session.commit()

        return jsonify({"success": True, "message": "Document deleted successfully"})

    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "message": str(e)}), 500

@inbound_bp.route("/documents/<int:doc_id>/view")
@login_required
def view_document(doc_id):
    """View/download a document"""
    doc = InboundDocument.query.get_or_404(doc_id)

    full_path = os.path.join("app", "static", doc.filepath)
    if not os.path.exists(full_path):
        abort(404)

    return send_file(full_path, download_name=doc.original_filename)


# ============ Confirmation Email File Upload Endpoints ============

@inbound_bp.route('/api/<service_type>/<int:record_id>/upload-confirmation', methods=['POST'])
@csrf.exempt
def api_upload_confirmation_file(service_type, record_id):
    """Upload confirmation email file for hotel, transport, or meal service"""

    print(f"[UPLOAD CONFIRMATION] Started - service_type={service_type}, record_id={record_id}")

    # Validate service type
    if service_type not in ['hotel', 'transport', 'meal', 'arrival']:
        print(f"[UPLOAD CONFIRMATION] Invalid service type: {service_type}")
        return jsonify({'success': False, 'error': 'Invalid service type'}), 400

    # Get the service record
    try:
        if service_type == 'hotel':
            service = InboundHotel.query.get_or_404(record_id)
        elif service_type == 'transport':
            service = InboundTransport.query.get_or_404(record_id)
        elif service_type == 'arrival':
            service = ArrivalBatch.query.get_or_404(record_id)
        else:  # meal
            service = InboundMeal.query.get_or_404(record_id)
        print(f"[UPLOAD CONFIRMATION] Found service record: {service}")
    except Exception as e:
        print(f"[UPLOAD CONFIRMATION] Service record not found: {e}")
        return jsonify({'success': False, 'error': 'Service record not found'}), 404

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not allowed_document_file(file.filename):
        return jsonify({'success': False, 'error': 'File type not allowed'}), 400

    try:
        # Generate unique filename
        original_filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{original_filename}"

        # Create upload directory
        upload_folder = os.path.join('app', 'static', 'uploads', 'confirmations', service_type, str(record_id))
        os.makedirs(upload_folder, exist_ok=True)

        # Delete old file if it exists
        if service.confirmation_email_filepath:
            old_path = os.path.join('app', 'static', service.confirmation_email_filepath)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass  # Ignore errors deleting old file

        # Save file
        filepath = os.path.join(upload_folder, unique_filename)
        file.save(filepath)

        # Update service record
        relative_filepath = f"uploads/confirmations/{service_type}/{record_id}/{unique_filename}"
        print(f"[UPLOAD CONFIRMATION] Before update - filename={getattr(service, 'confirmation_email_filename', 'NOT_SET')}")

        service.confirmation_email_filename = original_filename
        service.confirmation_email_filepath = relative_filepath
        service.confirmation_email_uploaded_at = datetime.utcnow()

        print(f"[UPLOAD CONFIRMATION] After update - filename={service.confirmation_email_filename}, filepath={service.confirmation_email_filepath}")
        print(f"[UPLOAD CONFIRMATION] Updated service record - filename={original_filename}, filepath={relative_filepath}")

        # Force flush to ensure attributes are updated
        db.session.flush()
        print(f"[UPLOAD CONFIRMATION] Session flushed")

        # Commit the transaction
        db.session.commit()
        print(f"[UPLOAD CONFIRMATION] Database committed successfully")

        # Verify the data was saved by querying it back
        verified_service = None
        if service_type == 'hotel':
            verified_service = InboundHotel.query.get(record_id)
        elif service_type == 'transport':
            verified_service = InboundTransport.query.get(record_id)
        elif service_type == 'arrival':
            verified_service = ArrivalBatch.query.get(record_id)
        else:
            verified_service = InboundMeal.query.get(record_id)

        print(f"[UPLOAD CONFIRMATION] Verified saved data - filename={verified_service.confirmation_email_filename if verified_service else 'NOT_FOUND'}")

        return jsonify({
            'success': True,
            'message': 'File uploaded successfully',
            'filename': original_filename,
            'filepath': relative_filepath,
            'uploaded_at': service.confirmation_email_uploaded_at.strftime('%Y-%m-%d %H:%M') if service.confirmation_email_uploaded_at else ''
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<service_type>/<int:record_id>/delete-confirmation', methods=['POST'])
@csrf.exempt
def api_delete_confirmation_file(service_type, record_id):
    """Delete confirmation email file for hotel, transport, meal, or arrival service"""

    # Validate service type
    if service_type not in ['hotel', 'transport', 'meal', 'arrival']:
        return jsonify({'success': False, 'error': 'Invalid service type'}), 400

    # Get the service record
    try:
        if service_type == 'hotel':
            service = InboundHotel.query.get_or_404(record_id)
        elif service_type == 'transport':
            service = InboundTransport.query.get_or_404(record_id)
        elif service_type == 'arrival':
            service = ArrivalBatch.query.get_or_404(record_id)
        else:  # meal
            service = InboundMeal.query.get_or_404(record_id)
    except:
        return jsonify({'success': False, 'error': 'Service record not found'}), 404

    try:
        # Delete file from filesystem
        if service.confirmation_email_filepath:
            full_path = os.path.join('app', 'static', service.confirmation_email_filepath)
            if os.path.exists(full_path):
                os.remove(full_path)

        # Update service record
        service.confirmation_email_filename = None
        service.confirmation_email_filepath = None
        service.confirmation_email_uploaded_at = None

        db.session.commit()

        return jsonify({'success': True, 'message': 'File deleted successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<service_type>/<int:record_id>/confirmation-file')
def get_confirmation_file(service_type, record_id):
    """Get/download confirmation email file"""

    # Validate service type
    if service_type not in ['hotel', 'transport', 'meal', 'arrival']:
        abort(400)

    # Get the service record
    try:
        if service_type == 'hotel':
            service = InboundHotel.query.get_or_404(record_id)
        elif service_type == 'transport':
            service = InboundTransport.query.get_or_404(record_id)
        elif service_type == 'arrival':
            service = ArrivalBatch.query.get_or_404(record_id)
        else:  # meal
            service = InboundMeal.query.get_or_404(record_id)
    except:
        abort(404)

    if not service.confirmation_email_filepath:
        abort(404)

    full_path = os.path.join(current_app.static_folder, service.confirmation_email_filepath)
    if not os.path.exists(full_path):
        abort(404)

    return send_file(full_path, download_name=service.confirmation_email_filename)


# ==================== Rooming List Attachment (hotel only) ====================
# Mirrors the Confirmation Email attachment mechanism exactly, but stores into a
# separate file slot (rooming_list_*) so a hotel can carry both attachments.

@inbound_bp.route('/api/<service_type>/<int:record_id>/upload-rooming-list', methods=['POST'])
@csrf.exempt
def api_upload_rooming_list_file(service_type, record_id):
    """Upload rooming list file for a hotel service"""

    if service_type != 'hotel':
        return jsonify({'success': False, 'error': 'Invalid service type'}), 400

    try:
        service = InboundHotel.query.get_or_404(record_id)
    except Exception as e:
        return jsonify({'success': False, 'error': 'Service record not found'}), 404

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file provided'}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'}), 400

    if not allowed_document_file(file.filename):
        return jsonify({'success': False, 'error': 'File type not allowed'}), 400

    try:
        # Generate unique filename
        original_filename = secure_filename(file.filename)
        unique_filename = f"{uuid.uuid4().hex}_{original_filename}"

        # Create upload directory
        upload_folder = os.path.join('app', 'static', 'uploads', 'rooming_lists', service_type, str(record_id))
        os.makedirs(upload_folder, exist_ok=True)

        # Delete old file if it exists
        if service.rooming_list_filepath:
            old_path = os.path.join('app', 'static', service.rooming_list_filepath)
            if os.path.exists(old_path):
                try:
                    os.remove(old_path)
                except:
                    pass  # Ignore errors deleting old file

        # Save file
        filepath = os.path.join(upload_folder, unique_filename)
        file.save(filepath)

        # Update service record
        relative_filepath = f"uploads/rooming_lists/{service_type}/{record_id}/{unique_filename}"
        service.rooming_list_filename = original_filename
        service.rooming_list_filepath = relative_filepath
        service.rooming_list_uploaded_at = datetime.utcnow()

        db.session.flush()
        db.session.commit()

        return jsonify({
            'success': True,
            'message': 'File uploaded successfully',
            'filename': original_filename,
            'filepath': relative_filepath,
            'uploaded_at': service.rooming_list_uploaded_at.strftime('%Y-%m-%d %H:%M') if service.rooming_list_uploaded_at else ''
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<service_type>/<int:record_id>/delete-rooming-list', methods=['POST'])
@csrf.exempt
def api_delete_rooming_list_file(service_type, record_id):
    """Delete rooming list file for a hotel service"""

    if service_type != 'hotel':
        return jsonify({'success': False, 'error': 'Invalid service type'}), 400

    try:
        service = InboundHotel.query.get_or_404(record_id)
    except:
        return jsonify({'success': False, 'error': 'Service record not found'}), 404

    try:
        # Delete file from filesystem
        if service.rooming_list_filepath:
            full_path = os.path.join('app', 'static', service.rooming_list_filepath)
            if os.path.exists(full_path):
                os.remove(full_path)

        # Update service record
        service.rooming_list_filename = None
        service.rooming_list_filepath = None
        service.rooming_list_uploaded_at = None

        db.session.commit()

        return jsonify({'success': True, 'message': 'File deleted successfully'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@inbound_bp.route('/api/<service_type>/<int:record_id>/rooming-list-file')
def get_rooming_list_file(service_type, record_id):
    """Get/download rooming list file"""

    if service_type != 'hotel':
        abort(400)

    try:
        service = InboundHotel.query.get_or_404(record_id)
    except:
        abort(404)

    if not service.rooming_list_filepath:
        abort(404)

    full_path = os.path.join(current_app.static_folder, service.rooming_list_filepath)
    if not os.path.exists(full_path):
        abort(404)

    return send_file(full_path, download_name=service.rooming_list_filename)

